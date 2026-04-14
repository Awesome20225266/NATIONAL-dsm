# dsm_dashboard.py
# Modern DSM Dashboard with Solar Analytics style

from __future__ import annotations

import json
from datetime import datetime, timedelta

import os
from pathlib import Path
from uuid import uuid4
import time

import numpy as np
import pandas as pd
import plotly.express as px

import dash
from dash import Dash, html, dcc, dash_table, Input, Output, State, ctx, ALL
from dash.exceptions import PreventUpdate
import dash_bootstrap_components as dbc
import dash_mantine_components as dmc

# Battery Investment (modular tab)
from battery_investment_tab import battery_investment_layout, register_battery_callbacks

# =========================
# Models & light utils (drop-in)
# =========================
from dataclasses import dataclass
from typing import List, Tuple, Dict, Any
from io import BytesIO
import math
import statistics as stats

# =========================
# Plant master (Excel) for cascading filters (Analysis tab)
# =========================
_PLANT_MASTER_CACHE: pd.DataFrame | None = None
_PLANT_MASTER_MTIME: float | None = None
_PLANT_MASTER_PATH = Path(__file__).with_name("consolidated_plant_list.xlsx")
ALL_SENTINEL = "__ALL__"
REMAINING_SENTINEL = "__REMAINING__"

# =========================
# Server-side cache for large runs (prevents browser "Aw, Snap" from huge dcc.Store payloads)
# =========================
# IMPORTANT: Dash stores are JSON-serialized into the browser. Even ~50k rows can crash Chrome
# for wide tables or multi-run payloads. Keep the preview small and cache early.
MAX_CLIENT_DETAIL_ROWS = 5_000           # max rows we will send to browser for preview
_DF_CACHE_DIR = Path(__file__).with_name(".dash_df_cache")
_DF_CACHE_DIR.mkdir(exist_ok=True)


def _cache_prune(max_files: int = 30, max_age_hours: float = 24.0) -> None:
    """Best-effort cleanup of cached df files (disk safety)."""
    try:
        files = sorted(_DF_CACHE_DIR.glob("*.pkl"), key=lambda p: p.stat().st_mtime, reverse=True)
        now = time.time()
        for i, p in enumerate(files):
            age_h = (now - p.stat().st_mtime) / 3600.0
            if i >= max_files or age_h > max_age_hours:
                try:
                    p.unlink(missing_ok=True)
                except Exception:
                    pass
    except Exception:
        pass


def _cache_df_to_disk(df: pd.DataFrame) -> str:
    """Write df to disk and return cache key."""
    _cache_prune()
    key = uuid4().hex
    path = _DF_CACHE_DIR / f"{key}.pkl"
    df.to_pickle(path)
    return key


def _load_df_from_cache(key: str) -> pd.DataFrame:
    """Load df from disk cache; returns empty df if missing."""
    if not key:
        return pd.DataFrame()
    path = _DF_CACHE_DIR / f"{str(key)}.pkl"
    if not path.exists():
        return pd.DataFrame()
    try:
        return pd.read_pickle(path)
    except Exception:
        return pd.DataFrame()


def _detail_df_from_store(stored: dict) -> pd.DataFrame:
    """Load detail dataframe from store, using cache if present."""
    if not stored or not isinstance(stored, dict):
        return pd.DataFrame()
    # Prefer cached full df when present
    key = stored.get("df_full_key")
    if key:
        df_full = _load_df_from_cache(key)
        if not df_full.empty:
            return df_full
    # Fallback to inline json/list
    try:
        if isinstance(stored.get("df"), str):
            return pd.DataFrame(json.loads(stored.get("df", "[]")))
        return pd.DataFrame(stored.get("df", []))
    except Exception:
        return pd.DataFrame()


def _pack_df_for_store(df: pd.DataFrame) -> dict:
    """Return dict fields to store df safely in results-store without crashing browser."""
    if df is None or not isinstance(df, pd.DataFrame) or df.empty:
        return {"df": "[]", "df_is_truncated": False, "df_full_key": None, "df_full_rows": 0, "df_preview_rows": 0}

    n = int(len(df))
    # Cache early: if it exceeds the preview threshold, write full df to disk and send preview only
    if n > MAX_CLIENT_DETAIL_ROWS:
        key = _cache_df_to_disk(df)
        preview = df.head(MAX_CLIENT_DETAIL_ROWS)
        return {
            "df": preview.to_json(date_format="iso", orient="records"),
            "df_is_truncated": True,
            "df_full_key": key,
            "df_full_rows": n,
            "df_preview_rows": int(len(preview)),
        }

    # Small enough: keep inline
    return {
        "df": df.to_json(date_format="iso", orient="records"),
        "df_is_truncated": False,
        "df_full_key": None,
        "df_full_rows": n,
        "df_preview_rows": n,
    }


def _pack_df_for_store_cache_only(df: pd.DataFrame) -> dict:
    """Always cache df to disk and return only a cache key + counts (no preview)."""
    if df is None or not isinstance(df, pd.DataFrame) or df.empty:
        return {"df": "[]", "df_is_truncated": False, "df_full_key": None, "df_full_rows": 0, "df_preview_rows": 0}
    n = int(len(df))
    key = _cache_df_to_disk(df)
    return {
        "df": "[]",
        "df_is_truncated": True,
        "df_full_key": key,
        "df_full_rows": n,
        "df_preview_rows": 0,
    }


def _load_plant_master_df() -> pd.DataFrame:
    """Load consolidated plant master from Excel (single source of truth for Analysis filters).

    Expected columns (case-insensitive):
      - plantname
      - Region
      - State
      - Type of Renewable
      - QCA (Agency)
      - Pooling Station
    """
    global _PLANT_MASTER_CACHE, _PLANT_MASTER_MTIME
    try:
        mtime = os.path.getmtime(_PLANT_MASTER_PATH)
    except Exception:
        mtime = None

    if _PLANT_MASTER_CACHE is not None and _PLANT_MASTER_MTIME is not None and mtime is not None:
        if float(mtime) == float(_PLANT_MASTER_MTIME):
            return _PLANT_MASTER_CACHE

    try:
        df = pd.read_excel(_PLANT_MASTER_PATH)
    except Exception as e:
        # If Excel is temporarily locked/open, fall back to last cached copy if available
        print(f"WARNING - Failed to read plant master Excel: {e}")
        if _PLANT_MASTER_CACHE is not None:
            return _PLANT_MASTER_CACHE
        return pd.DataFrame(columns=["plant_name", "region", "state", "resource", "qca", "pooling_station"])

    # Normalize columns
    col_map = {}
    for c in df.columns:
        key = str(c).strip().lower()
        col_map[key] = c

    def _pick(*names: str):
        for n in names:
            if n.lower() in col_map:
                return col_map[n.lower()]
        return None

    c_plant = _pick("plantname", "plant_name", "plant")
    c_region = _pick("region")
    c_state = _pick("state")
    c_res = _pick("type of renewable", "resource", "resource type", "type")
    c_qca = _pick("qca (agency)", "qca", "agency")
    c_pool = _pick("pooling station", "pooling", "pooling_station")

    out = pd.DataFrame()
    out["plant_name"] = df[c_plant].astype(str).str.strip() if c_plant else ""
    out["region"] = df[c_region].astype(str).str.strip().str.upper() if c_region else ""
    out["state"] = df[c_state].astype(str).str.strip() if c_state else ""
    out["resource"] = df[c_res].astype(str).str.strip() if c_res else ""
    out["qca"] = df[c_qca].astype(str).str.strip() if c_qca else ""
    out["pooling_station"] = df[c_pool].astype(str).str.strip() if c_pool else ""

    # Drop blank plants/regions
    out = out[(out["plant_name"] != "") & (out["region"] != "")]
    out = out.dropna(subset=["plant_name", "region"])

    _PLANT_MASTER_CACHE = out
    _PLANT_MASTER_MTIME = float(mtime) if mtime is not None else None
    return out


def _normalize_multi(val):
    """Normalize Dash dropdown value into a list of strings (or empty list)."""
    if val is None:
        return []
    if isinstance(val, list):
        return [str(v) for v in val if v is not None]
    return [str(val)]


def _strip_all_sentinel(vals: list[str]) -> list[str]:
    """If ALL sentinel present with other items, drop the sentinel."""
    vals = [v for v in vals if v is not None]
    if ALL_SENTINEL in vals and len(vals) > 1:
        vals = [v for v in vals if v != ALL_SENTINEL]
    return vals


def _is_all(vals: list[str]) -> bool:
    return (not vals) or (ALL_SENTINEL in vals)


def _filter_master(
    regions,
    states,
    resources,
    qcas,
    pools,
) -> pd.DataFrame:
    df = _load_plant_master_df()
    if df.empty:
        return df

    reg = [r.upper() for r in _strip_all_sentinel(_normalize_multi(regions)) if str(r).strip()]
    st = [s for s in _strip_all_sentinel(_normalize_multi(states)) if str(s).strip()]
    res = [r for r in _strip_all_sentinel(_normalize_multi(resources)) if str(r).strip()]
    q = [x for x in _strip_all_sentinel(_normalize_multi(qcas)) if str(x).strip()]
    p = [x for x in _strip_all_sentinel(_normalize_multi(pools)) if str(x).strip()]

    out = df
    if reg:
        out = out[out["region"].isin(reg)]
    if st and not _is_all(st):
        out = out[out["state"].isin(st)]
    if res and not _is_all(res):
        out = out[out["resource"].isin(res)]
    if q and not _is_all(q):
        out = out[out["qca"].isin(q)]
    if p and not _is_all(p):
        out = out[out["pooling_station"].isin(p)]
    return out


def enrich_plant_summary_with_master(plant_summary_df: pd.DataFrame) -> pd.DataFrame:
    """Add Region/State/Resource/QCA/Pooling Station columns for each plant using the Excel master.

    Analysis tab requirement: show metadata columns in Plant Summary table driven by consolidated_plant_list.xlsx.
    """
    if plant_summary_df is None or plant_summary_df.empty:
        return plant_summary_df

    ps = plant_summary_df.copy()
    if "Plant name" not in ps.columns:
        return ps

    m = _load_plant_master_df()
    if m.empty:
        return ps

    # Normalize join keys (exact match against master, but robust to NBSP/double-spaces/casing)
    ps["_plant_key"] = (
        ps["Plant name"]
          .astype(str)
          .str.replace("\u00A0", " ", regex=False)
          .str.strip()
          .str.replace(r"\s+", " ", regex=True)
          .str.casefold()
    )
    m2 = m.copy()
    m2["_plant_key"] = (
        m2["plant_name"]
          .astype(str)
          .str.replace("\u00A0", " ", regex=False)
          .str.strip()
          .str.replace(r"\s+", " ", regex=True)
          .str.casefold()
    )

    # One row per plant in master (pick first)
    m2 = (
        m2.sort_values(["region", "state", "resource", "qca", "pooling_station"])
          .drop_duplicates(subset=["_plant_key"], keep="first")
          .loc[:, ["_plant_key", "region", "state", "resource", "qca", "pooling_station"]]
          .rename(columns={
              "region": "Region_master",
              "state": "State",
              "resource": "Type of Renewable",
              "qca": "QCA",
              "pooling_station": "Pooling Station",
          })
    )

    ps = ps.merge(m2, on="_plant_key", how="left")

    # Prefer existing Region if present; otherwise use master region
    if "Region" in ps.columns:
        ps["Region"] = ps["Region"].where(ps["Region"].astype(str).str.strip() != "", ps["Region_master"])
    else:
        ps["Region"] = ps["Region_master"]

    ps = ps.drop(columns=["_plant_key", "Region_master"], errors="ignore")

    # Put metadata columns right after Region and Plant name (if present)
    desired_insert = ["Region", "State", "Type of Renewable", "QCA", "Pooling Station", "Plant name"]
    cols = list(ps.columns)
    ordered = []
    for c in desired_insert:
        if c in cols and c not in ordered:
            ordered.append(c)
    for c in cols:
        if c not in ordered:
            ordered.append(c)
    return ps[ordered]

@dataclass
class Band:
    direction: str        # "UI" | "OI"
    lower_pct: float      # inclusive lower bound
    upper_pct: float      # exclusive upper bound (use 999 for open-ended)
    rate_type: str        # "FLAT" | "PPA_FRAC" | "PPA_MULT" | "SCALED"
    rate_value: float     # flat ₹/kWh OR fraction/multiple 'a' in scaled
    rate_slope: float     # slope 'b' for scaled, else 0
    loss_zone: bool       # True → goes to OI_Loss (only used when direction="OI")

RATE_FLAT = "FLAT"
RATE_FRAC = "PPA_FRAC"
RATE_MULT = "PPA_MULT"
RATE_SCALED = "SCALED"

MODE_DEFAULT = "DEFAULT"
MODE_DYNAMIC = "DYNAMIC"

def safe_mode(values: List[float]) -> float:
    """MODE with sensible fallback (median) when multimodal/empty."""
    vals = [v for v in values if pd.notna(v)]
    if not vals:
        return 0.0
    try:
        return stats.mode(vals)
    except Exception:
        return float(np.median(vals))

def denominator_and_basis(avc: float, sch: float, mode: str, dyn_x: float) -> float:
    """Return denominator (also used as basis for energy) as per rule."""
    if mode == MODE_DYNAMIC:
        return (dyn_x * avc) + ((1.0 - dyn_x) * sch)
    return avc

def direction_from(actual: float, scheduled: float) -> str:
    if actual < scheduled:
        return "UI"
    elif actual > scheduled:
        return "OI"
    return "FLAT"

def slice_pct(abs_err: float, lower: float, upper: float) -> float:
    return max(0.0, min(abs_err, upper) - lower)

def kwh_from_slice(slice_pct_val: float, basis_mw: float) -> float:
    # 15-min block → 0.25 h; MW → kW × 1000; energy = P(kw)*h
    return (slice_pct_val / 100.0) * basis_mw * 0.25 * 1000.0

def band_rate(ppa: float, rate_type: str, rate_value: float, rate_slope: float, abs_err: float) -> float:
    if rate_type == RATE_FLAT:
        return rate_value
    if rate_type in (RATE_FRAC, RATE_MULT):
        return rate_value * ppa
    if rate_type == RATE_SCALED:
        return rate_value + rate_slope * abs_err
    return 0.0


def get_bandwise_analysis_df(
    detail_df: pd.DataFrame,
    bands_rows: list[dict] | None,
    err_mode: str = "default",
    x_pct: float = 50.0,
) -> pd.DataFrame:
    """Authoritative per-block bandwise output used by Battery Investment.

    Core principle: this function reuses the same slicing + DSM logic as the Analysis/export engine.
    Battery Investment must NOT reslice/recompute; it should call this function and then:
      - filter Scheduled Power (MW) > 0
      - convert MW -> kWh by multiplying by 250
      - visualize + compute statistics

    Returns a dataframe with:
      - Base block columns (plant/date/block/times, Scheduled/AvC/Actual/PPA, error%, abs_err, direction, basis)
      - One *band energy* column per band: "{UI|OI} Energy {lo}-{up}% (MW)" / "{UI|OI} Energy >{lo}% (MW)"
      - One *band DSM* column per band: "{UI|OI} DSM {lo}-{up}% (₹)" / "{UI|OI} DSM >{lo}% (₹)"
      - "Total DSM (₹)" as sum of all band DSM columns (no new logic; purely additive)
    """
    if detail_df is None or detail_df.empty:
        return pd.DataFrame()

    bands_rows = bands_rows or []
    if not bands_rows:
        return pd.DataFrame()

    df = detail_df.copy()

    # Normalize plant column
    if "plant_name" not in df.columns and "Plant" in df.columns:
        df["plant_name"] = df["Plant"]
    if "plant_name" not in df.columns:
        df["plant_name"] = "UNKNOWN"

    # Ensure numeric inputs exist
    for c in ["AvC_MW", "Scheduled_MW", "Actual_MW", "PPA"]:
        if c not in df.columns:
            df[c] = 0.0
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0.0)

    # Prefer existing error_pct/basis_MW (already computed by pipeline); otherwise compute.
    if "error_pct" not in df.columns:
        df["error_pct"] = compute_error_pct(df, str(err_mode or "default").lower(), float(x_pct or 50.0))
    df["error_pct"] = pd.to_numeric(df["error_pct"], errors="coerce").fillna(0.0)

    if "abs_err" not in df.columns:
        df["abs_err"] = df["error_pct"].abs()
    df["abs_err"] = pd.to_numeric(df["abs_err"], errors="coerce").fillna(0.0).abs()

    if "direction" not in df.columns:
        df["direction"] = np.where(df["error_pct"] < 0, "UI", "OI")
    df["direction"] = df["direction"].astype(str).str.upper()

    if "basis_MW" not in df.columns:
        df["basis_MW"] = compute_basis_mw(df, str(err_mode or "default").lower(), float(x_pct or 50.0))
    df["basis_MW"] = pd.to_numeric(df["basis_MW"], errors="coerce").fillna(0.0)

    # Prepare output scaffold with key columns (keep whatever exists from source)
    base_cols = []
    for c in [
        "region",
        "plant_name",
        "date",
        "time_block",
        "from_time",
        "to_time",
        "Scheduled_MW",
        "AvC_MW",
        "Actual_MW",
        "PPA",
        "error_pct",
        "abs_err",
        "direction",
        "basis_MW",
        "Custom Setting",
    ]:
        if c in df.columns and c not in base_cols:
            base_cols.append(c)

    out = df[base_cols].copy()

    # Compute per-band energy (MW) and per-band DSM (₹) columns from authoritative band definitions
    bands_sorted = _band_rows_sorted(bands_rows)
    band_energy_cols: list[str] = []
    band_dsm_cols: list[str] = []

    abs_err = out["abs_err"].to_numpy()
    basis_mw = out["basis_MW"].to_numpy()
    ppa = pd.to_numeric(df.get("PPA", 0.0), errors="coerce").fillna(0.0).to_numpy()
    direction = out["direction"].astype(str).to_numpy()

    for r in bands_sorted:
        dirn = str(r.get("direction", "")).strip().upper()
        try:
            lo = float(r.get("lower_pct", 0) or 0.0)
        except Exception:
            lo = 0.0
        try:
            up = float(r.get("upper_pct", 0) or 0.0)
        except Exception:
            up = 0.0

        # Column headers expected by Battery Investment tab
        if up >= 999:
            energy_col = f"{dirn} Energy >{int(lo)}% (MW)"
            dsm_col = f"{dirn} DSM >{int(lo)}% (₹)"
        else:
            energy_col = f"{dirn} Energy {int(lo)}-{int(up)}% (MW)"
            dsm_col = f"{dirn} DSM {int(lo)}-{int(up)}% (₹)"

        # Compute slice in % for this band and direction
        mask_dir = (direction == dirn)
        sp = np.maximum(0.0, np.minimum(abs_err, up) - lo)
        sp = np.where(mask_dir, sp, 0.0)

        # Energy in MW is slice% of basis (no sign; UI/OI already separated)
        energy_mw = (sp / 100.0) * basis_mw

        # DSM loss for this band (₹): kWh_from_slice * band_rate
        # IMPORTANT: bands_rows uses legacy rate_type strings; band_rate expects RATE_* constants.
        legacy_type = str(r.get("rate_type", "")).strip().lower()
        type_map = {
            "flat_per_kwh": RATE_FLAT,
            "ppa_fraction": RATE_FRAC,
            "ppa_multiple": RATE_MULT,
            "flat_per_mwh": RATE_FLAT,
            "scaled_excess": RATE_SCALED,
        }
        rate_type = type_map.get(legacy_type, RATE_FLAT)
        try:
            rate_value = float(r.get("rate_value", 0) or 0.0)
        except Exception:
            rate_value = 0.0
        # Convert MWh -> kWh if needed
        if legacy_type == "flat_per_mwh":
            rate_value = rate_value / 1000.0
        try:
            slope = float(r.get("excess_slope_per_pct", 0) or 0.0)
        except Exception:
            slope = 0.0
        # NOTE: band_rate takes abs_err (full abs_err), same as export logic
        rate = np.where(mask_dir, np.vectorize(band_rate)(ppa, rate_type, rate_value, slope, abs_err), 0.0)
        kwh = (sp / 100.0) * basis_mw * 0.25 * 1000.0
        dsm_loss = kwh * rate

        out[energy_col] = energy_mw
        out[dsm_col] = dsm_loss
        band_energy_cols.append(energy_col)
        band_dsm_cols.append(dsm_col)

    # Total DSM (₹) is purely additive over bands
    if band_dsm_cols:
        out["Total DSM (₹)"] = pd.to_numeric(out[band_dsm_cols].sum(axis=1), errors="coerce").fillna(0.0)
    else:
        out["Total DSM (₹)"] = 0.0

    return out

def compute_data_quality_stats(df: pd.DataFrame, max_issue_rows: int = 2000) -> dict:
    """Compute slot-level data-quality stats for a time-series dataframe.

    Evaluation is per row (typically plant_name + date + time_block).

    Categories are mutually exclusive (no double counting) and ordered by priority:
    - ACTUAL_POWER_MISSING
    - PPA_OR_AVC_MISSING
    - AVAILABLE
    - OTHER
    """
    if df is None or len(df) == 0:
        return {
            "total_ts": 0,
            "available_count": 0,
            "actual_missing_count": 0,
            "ppa_avc_missing_count": 0,
            "other_count": 0,
            "dsm_skipped_zero_inputs_count": 0,
            "available_pct": 0.0,
            "actual_missing_pct": 0.0,
            "ppa_avc_missing_pct": 0.0,
            "other_pct": 0.0,
            "dsm_skipped_zero_inputs_pct": 0.0,
            "issues_total": 0,
            "issues_rows": [],
        }

    ppa = pd.to_numeric(df.get("PPA"), errors="coerce").fillna(0.0)
    avc = pd.to_numeric(df.get("AvC_MW"), errors="coerce").fillna(0.0)
    fc = pd.to_numeric(df.get("Scheduled_MW"), errors="coerce").fillna(0.0)
    act = pd.to_numeric(df.get("Actual_MW"), errors="coerce").fillna(0.0)

    total_ts = int(len(df))

    # DSM NOT CALCULATED (hard skip) when any critical input is zero
    if "dsm_skipped_zero_inputs" in df.columns:
        dsm_skipped = df["dsm_skipped_zero_inputs"].fillna(False).astype(bool)
    else:
        dsm_skipped = (fc == 0) | (avc == 0) | (ppa == 0)
    dsm_skipped_count = int(dsm_skipped.sum())

    # FINAL AUTHORITATIVE RULES (slot-level)
    # DATA AVAILABLE if:
    #   ppa > 0 AND avc > 0 AND (
    #     (fc == 0 AND act == 0)              # valid night block
    #     OR (fc >= 0 AND act != 0)           # telemetry exists (positive or negative actual)
    #   )
    # DATA AVAILABLE (authoritative):
    # ppa > 0 AND avc > 0 AND (
    #   (fc == 0 AND act == 0)   # valid night block
    #   OR (act > 0)             # positive actual
    #   OR (act < 0)             # negative actual allowed (telemetry exists)
    # )
    available_base = (ppa > 0) & (avc > 0) & (((fc == 0) & (act == 0)) | (act > 0) | (act < 0))

    # ACTUAL POWER MISSING:
    #   fc > 0 AND ppa > 0 AND avc > 0 AND act == 0
    actual_missing = (fc > 0) & (ppa > 0) & (avc > 0) & (act == 0)

    # PPA / AVC MISSING:
    #   fc > 0 AND act > 0 AND (ppa == 0 OR avc == 0)
    ppa_avc_missing = (fc > 0) & (act > 0) & ((ppa == 0) | (avc == 0))

    # Make categories exclusive
    available = available_base & ~(actual_missing | ppa_avc_missing)
    other = ~(available | actual_missing | ppa_avc_missing)

    available_count = int(available.sum())
    actual_missing_count = int(actual_missing.sum())
    ppa_avc_missing_count = int(ppa_avc_missing.sum())
    other_count = int(other.sum())

    def pct(x: int) -> float:
        return round((x / total_ts * 100.0), 2) if total_ts else 0.0

    # Build issues table rows (only the two requested issue types)
    issues_mask = actual_missing | ppa_avc_missing
    issues_total = int(issues_mask.sum())
    if issues_total > 0:
        issues_df = df.loc[issues_mask, ["region", "plant_name", "date", "time_block", "from_time", "to_time"]].copy() \
            if all(c in df.columns for c in ["region", "plant_name", "date", "time_block", "from_time", "to_time"]) \
            else df.loc[issues_mask].copy()

        issues_df["Issue Type"] = np.where(actual_missing.loc[issues_df.index], "ACTUAL_POWER_MISSING", "PPA_OR_AVC_MISSING")
        # Helpful values for debugging
        for c in ("AvC_MW", "Scheduled_MW", "Actual_MW", "PPA"):
            if c in df.columns and c not in issues_df.columns:
                issues_df[c] = df.loc[issues_df.index, c]

        issues_df = issues_df.sort_values(by=[c for c in ["date", "time_block", "plant_name"] if c in issues_df.columns]).reset_index(drop=True)
        issues_rows = issues_df.head(max_issue_rows).to_dict("records")
    else:
        issues_rows = []

    return {
        "total_ts": total_ts,
        "available_count": available_count,
        "actual_missing_count": actual_missing_count,
        "ppa_avc_missing_count": ppa_avc_missing_count,
        "other_count": other_count,
        "dsm_skipped_zero_inputs_count": dsm_skipped_count,
        "available_pct": pct(available_count),
        "actual_missing_pct": pct(actual_missing_count),
        "ppa_avc_missing_pct": pct(ppa_avc_missing_count),
        "other_pct": pct(other_count),
        "dsm_skipped_zero_inputs_pct": pct(dsm_skipped_count),
        "issues_total": issues_total,
        "issues_rows": issues_rows,
    }


def compute_plant_data_availability_pct(df_plant: pd.DataFrame) -> float:
    """Compute data availability percentage for a single plant's dataframe.
    
    Uses the same slot-level logic as compute_data_quality_stats().
    Returns the percentage of slots that are "Data Available" for this plant.
    """
    if df_plant is None or len(df_plant) == 0:
        return 0.0
    
    ppa = pd.to_numeric(df_plant.get("PPA"), errors="coerce").fillna(0.0)
    avc = pd.to_numeric(df_plant.get("AvC_MW"), errors="coerce").fillna(0.0)
    fc = pd.to_numeric(df_plant.get("Scheduled_MW"), errors="coerce").fillna(0.0)
    act = pd.to_numeric(df_plant.get("Actual_MW"), errors="coerce").fillna(0.0)
    
    total_ts = len(df_plant)
    if total_ts == 0:
        return 0.0
    
    # Same authoritative rules as compute_data_quality_stats()
    available_base = (ppa > 0) & (avc > 0) & (((fc == 0) & (act == 0)) | (act > 0) | (act < 0))
    actual_missing = (fc > 0) & (ppa > 0) & (avc > 0) & (act == 0)
    ppa_avc_missing = (fc > 0) & (act > 0) & ((ppa == 0) | (avc == 0))
    
    available = available_base & ~(actual_missing | ppa_avc_missing)
    available_count = int(available.sum())
    
    return round((available_count / total_ts * 100.0), 2) if total_ts else 0.0

# =========================
# ---------- THEME --------
# =========================
APP_TITLE = "Zelestra: DSM Analytics"
DASH_PORT = 8050  # Fixed local port (explicit in app.run; overrides PORT env when passed)
THEME = dbc.themes.BOOTSTRAP
BUILD_TAG = "assignment-ui-v2 (2026-01-12)"

app: Dash = Dash(
    __name__,
    external_stylesheets=[THEME, *dmc.styles.ALL],
    title=APP_TITLE,
    suppress_callback_exceptions=True,
)
server = app.server

# ==============================
# ---- DEFAULT PRESET BANDS ----
# ==============================
DEFAULT_BANDS = [
    {"direction": "UI", "lower_pct": 0.0,  "upper_pct": 15.0, "tolerance_cut_pct": 15.0,
     "rate_type": "flat_per_kwh", "rate_value": 0.0, "excess_slope_per_pct": 0.0, "loss_zone": False, "label": "UI ≤15% (no penalty)"},
    {"direction": "UI", "lower_pct": 15.0, "upper_pct": 20.0, "tolerance_cut_pct": 15.0,
     "rate_type": "ppa_fraction", "rate_value": 0.10, "excess_slope_per_pct": 0.0, "loss_zone": False, "label": "UI 15–20% (10% of PPA)"},
    {"direction": "UI", "lower_pct": 20.0, "upper_pct": 1_000.0, "tolerance_cut_pct": 20.0,
     "rate_type": "scaled_excess", "rate_value": 3.36, "excess_slope_per_pct": 0.08, "loss_zone": False, "label": "UI >20% (scaled)"},
    {"direction": "OI", "lower_pct": 0.0,  "upper_pct": 15.0, "tolerance_cut_pct": 15.0,
     "rate_type": "flat_per_kwh", "rate_value": 0.0, "excess_slope_per_pct": 0.0, "loss_zone": False, "label": "OI ≤15% (no penalty)"},
    {"direction": "OI", "lower_pct": 15.0, "upper_pct": 20.0, "tolerance_cut_pct": 15.0,
     "rate_type": "ppa_fraction", "rate_value": 0.10, "excess_slope_per_pct": 0.0, "loss_zone": False, "label": "OI 15–20% (10% of PPA)"},
    {"direction": "OI", "lower_pct": 20.0, "upper_pct": 1_000.0, "tolerance_cut_pct": 20.0,
     "rate_type": "scaled_excess", "rate_value": 3.36, "excess_slope_per_pct": 0.08, "loss_zone": True, "label": "OI >20% (scaled)"},
]


BANDS_COLUMNS = [
    {"name": "Direction", "id": "direction", "presentation": "dropdown", "editable": True},
    {"name": "Lower %", "id": "lower_pct", "type": "numeric", "format": {"specifier": ".1f"}},
    {"name": "Upper %", "id": "upper_pct", "type": "numeric", "format": {"specifier": ".1f"}},
    # Tolerance Cut % and Deviated On are now hidden - handled by regulation mode
    {"name": "Rate Type", "id": "rate_type", "presentation": "dropdown", "editable": True},
    {"name": "Rate Value", "id": "rate_value", "type": "numeric", "format": {"specifier": ".2f"}},
    {"name": "Excess Slope %", "id": "excess_slope_per_pct", "type": "numeric", "format": {"specifier": ".2f"}},
    {"name": "Loss Zone", "id": "loss_zone", "presentation": "dropdown", "editable": True},
    {"name": "Label", "id": "label", "editable": False},
]

BANDS_DROPDOWNS = {
    "direction": {"options": [{"label": "UI (Under Injection)", "value": "UI"}, {"label": "OI (Over Injection)", "value": "OI"}]},
    "rate_type": {"options": [
                {"label": "Flat per kWh", "value": "flat_per_kwh"},
                {"label": "PPA Fraction", "value": "ppa_fraction"},
                {"label": "PPA Multiple", "value": "ppa_multiple"},
                {"label": "Flat per MWh", "value": "flat_per_mwh"},
                {"label": "Scaled Excess", "value": "scaled_excess"},
    ]},
    "loss_zone": {"options": [
                {"label": "No", "value": False},
                {"label": "Yes", "value": True},
    ]},
    # Apply To and Deviated On removed; denominator/basis is driven by Error% Mode
}

# =====================================
# ---------- DATA LOADING FUNCTIONS ---
# =====================================
import os
import glob
from functools import lru_cache
from pathlib import Path
import re
from datetime import datetime, timedelta

_DUCKDB_DIR = Path(__file__).resolve().parent

# ---------------------------------------------------------------------------
# AWS S3 helpers — download regional DuckDB files for Render deployment.
# Secrets are read from environment variables (set on Render dashboard or
# loaded from a local .env file via python-dotenv).  Never printed.
# ---------------------------------------------------------------------------

_DSM_S3_KEY_MAP: dict[str, str] = {
    "nrpc": "S3_KEY_NRPC",
    "srpc": "S3_KEY_SRPC",
    "wrpc": "S3_KEY_WRPC",
}


def _get_secret(name: str) -> str | None:
    """Read a secret from environment variables. Never prints the value."""
    v = os.environ.get(name)
    return v if v and v.strip() else None


def _get_s3_key_for_region(region_lower: str) -> str:
    """
    Return the S3 object key for a DSM region.
    Reads from S3_KEY_NRPC / S3_KEY_SRPC / S3_KEY_WRPC env vars.
    Raises RuntimeError if the secret is missing.
    """
    secret_name = _DSM_S3_KEY_MAP.get(region_lower.strip().lower())
    if not secret_name:
        raise ValueError(f"Unknown DSM region '{region_lower}'. Allowed: {list(_DSM_S3_KEY_MAP)}")
    s3_key = _get_secret(secret_name)
    if not s3_key:
        raise RuntimeError(
            f"Missing env var '{secret_name}' required to download '{region_lower}.duckdb' from S3. "
            "Set it in the Render environment variables (or .env for local runs)."
        )
    return s3_key


def _ensure_region_db_local(region_lower: str) -> Path:
    """
    Ensure the regional DuckDB file exists locally.
    Downloads from S3 if the file is absent (e.g. fresh Render container).
    Returns the local Path.  Read credentials come from env vars.
    """
    local_path = _DUCKDB_DIR / f"{region_lower.strip().lower()}.duckdb"

    if local_path.exists() and local_path.stat().st_size > 0:
        return local_path  # already present — fast path

    aws_access_key_id     = _get_secret("AWS_ACCESS_KEY_ID")
    aws_secret_access_key = _get_secret("AWS_SECRET_ACCESS_KEY")
    aws_region            = _get_secret("AWS_REGION")
    s3_bucket             = _get_secret("S3_BUCKET")

    missing = [k for k, v in {
        "AWS_ACCESS_KEY_ID":     aws_access_key_id,
        "AWS_SECRET_ACCESS_KEY": aws_secret_access_key,
        "AWS_REGION":            aws_region,
        "S3_BUCKET":             s3_bucket,
    }.items() if not v]
    if missing:
        raise RuntimeError(
            f"Cannot download '{region_lower}.duckdb' — missing env vars: {', '.join(missing)}. "
            "Set them in the Render environment variables (or .env for local runs)."
        )

    s3_key = _get_s3_key_for_region(region_lower)

    import boto3  # type: ignore

    s3 = boto3.client(
        "s3",
        aws_access_key_id=aws_access_key_id,
        aws_secret_access_key=aws_secret_access_key,
        region_name=aws_region,
    )

    tmp_path = local_path.with_suffix(".duckdb.tmp")
    try:
        print(f"[dsm_dashboard] Downloading {region_lower}.duckdb from S3 ...", flush=True)
        s3.download_file(s3_bucket, s3_key, str(tmp_path))
        os.replace(str(tmp_path), str(local_path))
        print(f"[dsm_dashboard] {region_lower}.duckdb ready.", flush=True)
    except Exception:
        try:
            tmp_path.unlink(missing_ok=True)
        except Exception:
            pass
        raise

    return local_path


def _duckdb_path(region_lower: str) -> Path:
    """
    Resolve DuckDB file path for a region, downloading from S3 if not present locally.
    All existing duckdb.connect() calls go through this — no other changes needed.
    """
    return _ensure_region_db_local(str(region_lower).lower())

def _norm_plant_name(name: str) -> str:
    try:
        s = str(name)
    except Exception:
        s = str(name)
    # Normalize: remove odd chars, collapse any whitespace, normalize dashes, uppercase.
    s = s.replace("\u00A0", " ")  # non-breaking space
    # Normalize common unicode dashes to '-'
    s = s.replace("\u2013", "-").replace("\u2014", "-").replace("\u2212", "-")  # en/em/minus
    s = s.strip().upper()
    # Replace non [A-Z0-9 -] characters with space (handles invisible/garbled chars like �)
    # NOTE: keep '-' literal by placing it at the end of the character class (avoids regex escape edge-cases).
    s = re.sub(r"[^A-Z0-9 -]+", " ", s)
    # Collapse repeated whitespace
    s = re.sub(r"\s+", " ", s).strip()
    return s

def get_nrpc_plants():
    """Get unique plant names from NRPC Excel files with caching for speed."""
    # Check cache first
    if "NRPC" in PLANTS_CACHE:
        return PLANTS_CACHE["NRPC"]
    
    try:
        nrpc_path = "Raw Data for Dashboard/NRPC"
        if not os.path.exists(nrpc_path):
            PLANTS_CACHE["NRPC"] = ["Plant-A", "Plant-B", "Plant-C"]
            return PLANTS_CACHE["NRPC"]

        excel_files = glob.glob(os.path.join(nrpc_path, "*.xlsx"))
        # Filter out temporary Excel lock files
        excel_files = [f for f in excel_files if not os.path.basename(f).startswith("~$")]

        if not excel_files:
            PLANTS_CACHE["NRPC"] = ["Plant-A", "Plant-B", "Plant-C"]
            return PLANTS_CACHE["NRPC"]

        # Helper to normalize names for deduplication
        def _norm(name: str) -> str:
            s = str(name).strip()
            s = " ".join(s.split())
            return s.upper()

        normalized_to_original: dict[str, str] = {}

        # Read only first 10 files for speed - this should be enough to get all unique plants
        for file_path in excel_files[:10]:
            try:
                # Read header first to locate the plant column by name
                header_df = pd.read_excel(file_path, nrows=0)
                columns = list(header_df.columns)
                lower_map = {str(c).strip().lower(): c for c in columns}

                plant_col = None
                if "plant_name" in lower_map:
                    plant_col = lower_map["plant_name"]
                else:
                    # Fallback heuristic: any column containing both tokens 'plant' and 'name'
                    for lc, orig in lower_map.items():
                        if "plant" in lc and "name" in lc:
                            plant_col = orig
                            break

                if plant_col is None:
                    # As a last resort, try any column that includes 'plant'
                    for lc, orig in lower_map.items():
                        if "plant" in lc:
                            plant_col = orig
                            break

                if plant_col is None:
                    continue

                # Read ONLY the identified plant column for the entire file
                series = pd.read_excel(file_path, usecols=[plant_col])[plant_col]
                unique_plants = series.dropna().astype(str).unique()
                
                for raw_name in unique_plants:
                    key = _norm(raw_name)
                    if key and key not in normalized_to_original:
                        normalized_to_original[key] = raw_name.strip()
            except Exception:
                continue

        if not normalized_to_original:
            PLANTS_CACHE["NRPC"] = ["Plant-A", "Plant-B", "Plant-C"]
            return PLANTS_CACHE["NRPC"]

        plants = sorted(normalized_to_original.values(), key=lambda x: x.upper())
        PLANTS_CACHE["NRPC"] = plants  # Cache the result
        return plants
    except Exception:
        PLANTS_CACHE["NRPC"] = ["Plant-A", "Plant-B", "Plant-C"]
        return PLANTS_CACHE["NRPC"]

def get_srpc_plants():
    """Get unique plant names from SRPC CSV files with dev pattern and caching."""
    # Check cache first
    if "SRPC" in PLANTS_CACHE:
        return PLANTS_CACHE["SRPC"]
    
    try:
        srpc_path = "Downloader Setup/Raw_SRPC_Downloaded"
        if not os.path.exists(srpc_path):
            PLANTS_CACHE["SRPC"] = ["Plant-F", "Plant-G", "Plant-H"]
            return PLANTS_CACHE["SRPC"]
        
        plants = set()
        
        # Look in year folders - limit to first 2 for speed
        year_folders = [f for f in os.listdir(srpc_path) if os.path.isdir(os.path.join(srpc_path, f))]
        
        for year_folder in year_folders[:2]:  # Check first 2 year folders
            year_path = os.path.join(srpc_path, year_folder)
            date_folders = [f for f in os.listdir(year_path) if os.path.isdir(os.path.join(year_path, f))]
            
            for date_folder in date_folders[:3]:  # Check first 3 date folders
                date_path = os.path.join(year_path, date_folder)
                zip_data_path = os.path.join(date_path, "Zip_Data")
                
                if os.path.exists(zip_data_path):
                    csv_files = glob.glob(os.path.join(zip_data_path, "*dev*.csv"))
                    
                    for csv_file in csv_files:
                        filename = os.path.basename(csv_file)
                        # Extract plant name from commercial_dev2022_plant_name pattern
                        match = re.search(r'commercial_dev\d+_(.+)\.csv', filename)
                        if match:
                            plant_name = match.group(1)
                            plants.add(plant_name)
        
        result = sorted(list(plants)) if plants else ["Plant-F", "Plant-G", "Plant-H"]
        PLANTS_CACHE["SRPC"] = result  # Cache the result
        return result
    except Exception:
        PLANTS_CACHE["SRPC"] = ["Plant-F", "Plant-G", "Plant-H"]
        return PLANTS_CACHE["SRPC"]

def get_wrpc_plants():
    """Get WRPC plants - placeholder for now"""
    return ["Plant-D", "Plant-E"]

REGIONS = ["NRPC", "WRPC", "SRPC"]

# Cache for plant discovery - populated once and reused
PLANTS_CACHE = {}

# Use fallback plants at import time, will be updated dynamically
PLANTS = {
    "NRPC": ["Plant-A", "Plant-B", "Plant-C"],
    "WRPC": ["Plant-D", "Plant-E"], 
    "SRPC": ["Plant-F", "Plant-G", "Plant-H"],
}

# ------------------------------------------
# Plant classification helpers (Solar vs Wind)
# ------------------------------------------
def classify_plants_by_type(regions: list[str], start_date: str, end_date: str) -> dict:
    """Return mapping {plant_name: 'SOLAR'|'WIND'|'UNKNOWN'} using schedule window.
    Rule: 
    - WIND: Has scheduled power in early morning blocks (blocks 1-20, before 05:00)
    - SOLAR: Scheduled power only from block 21 onwards (05:00 or later)
    - UNKNOWN: No schedule data
    """
    mapping: dict[str, str] = {}
    for region in regions or []:
        region_lower = region.lower()
        db_file = _duckdb_path(region_lower)
        if not db_file.exists():
            continue
        try:
            conn = duckdb.connect(str(db_file), read_only=True)
            q = f"""
                SELECT plant_name,
                       MIN(time_block) AS min_block,
                       MAX(time_block) AS max_block
                FROM {region_lower}
                WHERE date >= '{start_date}'
                  AND date <= '{end_date}'
                  AND forecasted_power > 0
                GROUP BY plant_name
            """
            agg = conn.execute(q).fetchdf()
            conn.close()
            for _, row in agg.iterrows():
                plant = row["plant_name"]
                try:
                    min_b = int(row["min_block"]) if pd.notna(row["min_block"]) else None
                    max_b = int(row["max_block"]) if pd.notna(row["max_block"]) else None
                except Exception:
                    min_b, max_b = None, None
                if min_b is None or max_b is None:
                    mapping[_norm_plant_name(plant)] = "UNKNOWN"
                    continue
                # If minimum block is before block 21 (before 05:00), it's WIND
                # Block 21 = 05:00 AM (each block is 15 mins, so block 1 = 00:00-00:15)
                if min_b < 21:
                    plant_type = "WIND"
                    mapping[_norm_plant_name(plant)] = plant_type
                    print(f"DEBUG: {plant} classified as {plant_type} (min_block={min_b}, max_block={max_b})")
                elif min_b >= 21:
                    plant_type = "SOLAR"
                    mapping[_norm_plant_name(plant)] = plant_type
                    print(f"DEBUG: {plant} classified as {plant_type} (min_block={min_b}, max_block={max_b})")
                else:
                    mapping[_norm_plant_name(plant)] = "UNKNOWN"
        except Exception as e:
            print(f"Error classifying plants in {db_file}: {e}")
            continue
    return mapping

@lru_cache(maxsize=1)
def load_plant_renewable_mapping() -> dict[str, str]:
    """Load plant-to-renewable mapping from consolidated_plant_list.xlsx.
    Returns dict mapping normalized plant name to renewable type (Solar/Wind/Thermal).
    """
    mapping: dict[str, str] = {}

    # Prefer resolving relative to THIS file to avoid issues when cwd changes (Dash reloaders/callbacks).
    script_path = Path(__file__).resolve()
    script_dir = script_path.parent

    candidates: list[Path] = [
        script_dir / "consolidated_plant_list.xlsx",          # same folder as this script (DSM MASTER)
        script_dir.parent / "consolidated_plant_list.xlsx",   # project root (if someone moved the file)
        script_dir.parent / "DSM MASTER" / "consolidated_plant_list.xlsx",  # explicit DSM MASTER from root
        Path.cwd() / "consolidated_plant_list.xlsx",          # cwd fallback
        Path.cwd() / "DSM MASTER" / "consolidated_plant_list.xlsx",  # cwd + DSM MASTER fallback
    ]

    excel_path: Path | None = next((p for p in candidates if p.exists()), None)

    # Last-resort: cheap bounded search (only within the script dir and its parent).
    if excel_path is None:
        for base in (script_dir, script_dir.parent):
            try:
                found = next(base.rglob("consolidated_plant_list.xlsx"), None)
            except Exception:
                found = None
            if found is not None and found.exists():
                excel_path = found
                break

    if excel_path is None:
        attempted = ", ".join(str(p) for p in candidates)
        print(
            "[WARN] consolidated_plant_list.xlsx not found. "
            "Plant filtering by resource will not work. "
            f"(cwd={os.getcwd()}, __file__={script_path}, attempted={attempted})"
        )
        return mapping
    
    try:
        df = pd.read_excel(str(excel_path))
        if "plantname" not in df.columns or "renewable" not in df.columns:
            print(f"[WARN] {excel_path} missing required columns (plantname, renewable).")
            return mapping
        
        for _, row in df.iterrows():
            plant = str(row["plantname"]).strip()
            renewable = str(row["renewable"]).strip()
            if plant and renewable:
                mapping[_norm_plant_name(plant)] = renewable
    except Exception as e:
        print(f"[WARN] Error loading {excel_path}: {e}")
    
    return mapping

def get_filtered_plants_by_type(regions: list[str], type_value: str, start_date: str, end_date: str) -> list[str]:
    """Return plants filtered by type selection (ALL/SOLAR/WIND/THERMAL).
    Uses consolidated_plant_list.xlsx if available, otherwise falls back to classification logic.
    """
    all_plants = get_plants_from_duckdb(regions)
    t = (type_value or "ALL").upper()
    if t == "ALL":
        return all_plants
    
    # Try to use Excel mapping first
    excel_mapping = load_plant_renewable_mapping()
    if excel_mapping:
        # Map Excel renewable values to filter values (Solar->SOLAR, Wind->WIND, Thermal->THERMAL)
        renewable_map = {"Solar": "SOLAR", "Wind": "WIND", "Thermal": "THERMAL"}
        filtered = []
        for p in all_plants:
            plant_norm = _norm_plant_name(p)
            excel_type = excel_mapping.get(plant_norm, "").strip()
            mapped_type = renewable_map.get(excel_type, "").upper()
            if mapped_type == t:
                filtered.append(p)
        
        if filtered:
            return filtered
        # If no matches found with Excel, fall back to classification logic
    
    # Fallback to old classification logic (only handles SOLAR/WIND, not THERMAL)
    mapping = classify_plants_by_type(regions, start_date, end_date)
    filtered = [p for p in all_plants if mapping.get(_norm_plant_name(p), "UNKNOWN") == t]
    return filtered if filtered else all_plants

# ==========================================
# -------- DUCKDB HELPER FUNCTIONS ---------
# ==========================================
import duckdb

def get_regions_from_duckdb() -> list[str]:
    """Get unique regions from all DuckDB databases"""
    regions = []
    try:
        # Check if nrpc.duckdb exists
        nrpc_db = _duckdb_path("nrpc")
        if nrpc_db.exists():
            conn = duckdb.connect(str(nrpc_db), read_only=True)
            result = conn.execute("SELECT DISTINCT region FROM nrpc").fetchall()
            regions.extend([r[0].upper() for r in result])
            conn.close()
    except Exception as e:
        print(f"Error reading nrpc.duckdb: {e}")
    
    try:
        # Check if srpc.duckdb exists
        srpc_db = _duckdb_path("srpc")
        if srpc_db.exists():
            conn = duckdb.connect(str(srpc_db), read_only=True)
            result = conn.execute("SELECT DISTINCT region FROM srpc").fetchall()
            regions.extend([r[0].upper() for r in result])
            conn.close()
    except Exception as e:
        print(f"Error reading srpc.duckdb: {e}")
    
    try:
        # Check if wrpc.duckdb exists
        wrpc_db = _duckdb_path("wrpc")
        if wrpc_db.exists():
            conn = duckdb.connect(str(wrpc_db), read_only=True)
            result = conn.execute("SELECT DISTINCT region FROM wrpc").fetchall()
            regions.extend([r[0].upper() for r in result])
            conn.close()
    except Exception as e:
        print(f"Error reading wrpc.duckdb: {e}")
    
    # Return unique regions, fallback to default if none found
    return sorted(list(set(regions))) if regions else ["NRPC", "SRPC", "WRPC"]

def get_plants_from_duckdb(regions: list[str]) -> list[str]:
    """Get unique plant names from DuckDB for selected regions"""
    plants = []
    
    for region in regions:
        region_lower = region.lower()
        db_file = _duckdb_path(region_lower)
        
        if not db_file.exists():
            pass  # Skip this region
        else:
            try:
                conn = duckdb.connect(str(db_file), read_only=True)
                result = conn.execute(f"SELECT DISTINCT plant_name FROM {region_lower} ORDER BY plant_name").fetchall()
                # Normalize for stability (fixes double-spaces / odd chars so UI selection matches DB)
                plants.extend([_norm_plant_name(r[0]) for r in result if r[0] and r[0] != 'UNKNOWN'])
                conn.close()
            except Exception as e:
                print(f"Error reading {db_file}: {e}")
    
    return sorted(list(set(plants)))

def get_plant_full_active_date_range(plant_name: str, region: str = None) -> tuple[str, str]:
    """
    Query the FULL database for a plant to find min/max dates where Actual_MW > 0.
    This is independent of any user-selected date range.
    
    Returns: (first_active_date, last_active_date) as strings in 'DD-MMM-YYYY' format,
             or ("No Active Data", "No Active Data") if no active power found.
    """
    # If region is provided, ensure it's uppercase and valid
    if region:
        region = str(region).strip().upper()
        if region not in ["NRPC", "SRPC", "WRPC"]:
            region = None
    
    regions_to_check = [region] if region else ["NRPC", "SRPC", "WRPC"]
    
    for reg in regions_to_check:
        try:
            db_file = _duckdb_path(reg.lower())
            if not db_file.exists():
                print(f"DEBUG - DB file not found for {reg}: {db_file}")
                continue
            
            conn = duckdb.connect(str(db_file), read_only=True)
            
            # Normalize plant name for matching (same as load_data_from_duckdb)
            plant_norm = _norm_plant_name(plant_name).replace("'", "''")
            
            # SQL to normalize plant_name in DB (match load_data_from_duckdb; avoid '\u' escapes in DuckDB)
            _plant_key_expr = (
                "trim("
                "regexp_replace("
                "regexp_replace(upper(replace(plant_name, chr(160), ' ')), '[^A-Z0-9\\- ]+', ' ', 'g'), "
                "'\\s+', ' ', 'g')"
                ")"
            )

            # Robust date normalization (mirrors load_data_from_duckdb) so MIN/MAX work even if date is VARCHAR
            _date_expr = (
                "coalesce("
                "  try_cast(date as DATE),"
                "  cast(try_strptime(cast(date as VARCHAR), '%d-%m-%Y') as DATE),"
                "  cast(try_strptime(cast(date as VARCHAR), '%Y-%m-%d') as DATE)"
                ")"
            )

            # Robust actual_power numeric parsing (handles numeric/varchar/commas)
            _power_expr = (
                "coalesce("
                "  try_cast(actual_power as DOUBLE),"
                "  try_cast(replace(cast(actual_power as VARCHAR), ',', '') as DOUBLE),"
                "  0.0"
                ")"
            )
            
            # First, check what columns exist in the table
            try:
                table_cols = conn.execute(f"DESCRIBE {reg.lower()}").fetchdf()
                print(f"DEBUG - Columns in {reg} table: {table_cols['column_name'].tolist()}")
            except Exception:
                pass
            
            # Debug: Check what data exists for this plant
            try:
                debug_query = f"""
                    SELECT 
                        COUNT(*) as total_rows,
                        COUNT(CASE WHEN {_power_expr} > 0 THEN 1 END) as rows_with_power,
                        MIN({_power_expr}) as min_power,
                        MAX({_power_expr}) as max_power,
                        MIN({_date_expr}) as first_date,
                        MAX({_date_expr}) as last_date
                    FROM {reg.lower()}
                    WHERE {_plant_key_expr} = '{plant_norm}'
                """
                debug_result = conn.execute(debug_query).fetchone()
                if debug_result:
                    print(f"DEBUG - Plant '{plant_norm}' in {reg}: {debug_result[0]} total rows, {debug_result[1]} with actual_power>0, power range: {debug_result[2]} to {debug_result[3]}, dates: {debug_result[4]} to {debug_result[5]}")
            except Exception as debug_err:
                print(f"DEBUG - Could not get debug info: {debug_err}")
            
            # Now query for date range where actual_power > 0
            try:
                query = f"""
                    SELECT 
                        MIN({_date_expr}) as first_date,
                        MAX({_date_expr}) as last_date,
                        COUNT(*) as row_count
                    FROM {reg.lower()}
                    WHERE {_plant_key_expr} = '{plant_norm}'
                      AND {_power_expr} > 0
                """
                
                result = conn.execute(query).fetchone()
                
                if result and result[0] is not None and result[1] is not None and result[2] > 0:
                    # Found active data for this plant in this region
                    first_date = pd.to_datetime(result[0]).strftime('%d-%b-%Y')
                    last_date = pd.to_datetime(result[1]).strftime('%d-%b-%Y')
                    print(f"DEBUG - Found date range for '{plant_name}' in {reg}: {first_date} to {last_date} ({result[2]} rows with actual_power>0)")
                    conn.close()
                    return (first_date, last_date)
                else:
                    print(f"DEBUG - No rows with actual_power > 0 for '{plant_norm}' in {reg} (query returned: {result})")
            except Exception as query_err:
                print(f"DEBUG - Query error: {query_err}")
                import traceback
                traceback.print_exc()
            
            conn.close()
            print(f"DEBUG - No active data found for '{plant_name}' (normalized: '{plant_norm}') in {reg}")
            
        except Exception as e:
            print(f"DEBUG - Error querying full date range for '{plant_name}' in {reg}: {e}")
            import traceback
            traceback.print_exc()
            continue
    
    # No active data found in any region
    print(f"DEBUG - No active data found for '{plant_name}' in any region")
    return ("No Active Data", "No Active Data")

def load_data_from_duckdb(region: str, start_date: str, end_date: str, plants: list[str]) -> pd.DataFrame:
    """Load data from DuckDB for specified region, date range, and plants"""
    region_lower = region.lower()
    db_file = _duckdb_path(region_lower)
    
    if not db_file.exists():
        print(f"DEBUG - Database file {db_file} not found")
        return pd.DataFrame()
    
    try:
        conn = duckdb.connect(str(db_file), read_only=True)

        # Normalize plant names for robust matching (fixes double spaces / hidden chars)
        norm_plants = [_norm_plant_name(p).replace("'", "''") for p in (plants or [])]
        plant_filter = ", ".join([f"'{p}'" for p in norm_plants]) if norm_plants else None

        # SQL expression to normalize plant_name inside DuckDB:
        # - uppercase
        # - replace NBSP with space
        # - treat dash-like separators as whitespace (fixes "BHADLA-II" vs "BHADLA II")
        # - replace odd characters (like �) with space
        # - collapse whitespace (handles double spaces)
        # DuckDB regexp_replace defaults to FIRST match only; use the 'g' flag for global replacement.
        # Keep '-' significant (as per master/DB names), but normalize NBSP + collapse whitespace.
        _plant_key_expr = (
            "trim("
            "regexp_replace("
            "regexp_replace(upper(replace(plant_name, chr(160), ' ')), '[^A-Z0-9\\- ]+', ' ', 'g'), "
            "'\\s+', ' ', 'g')"
            ")"
        )

        # Robust date normalization (handles DATE/TIMESTAMP and string dates like 'DD-MM-YYYY' or 'YYYY-MM-DD')
        _date_expr = (
            "coalesce("
            "  try_cast(date as DATE),"
            "  cast(try_strptime(cast(date as VARCHAR), '%d-%m-%Y') as DATE),"
            "  cast(try_strptime(cast(date as VARCHAR), '%Y-%m-%d') as DATE)"
            ")"
        )

        # Debug: Check what plant names actually exist in the database
        debug_query = f"""
            SELECT DISTINCT
              {_plant_key_expr} as plant_key,
              plant_name as plant_orig
            FROM {region_lower}
            LIMIT 50
        """
        try:
            debug_df = conn.execute(debug_query).fetchdf()
            print(f"DEBUG - Sample plant names in {region_lower} database:")
            for _, row in debug_df.iterrows():
                print(f"  Key: '{row['plant_key']}' | Original: '{row['plant_orig']}'")
        except Exception as debug_err:
            print(f"DEBUG - Could not fetch sample plants: {debug_err}")

        # Debug: Check date range in database
        date_range_query = f"SELECT MIN({_date_expr}) as min_date, MAX({_date_expr}) as max_date FROM {region_lower}"
        try:
            date_range = conn.execute(date_range_query).fetchone()
            print(f"DEBUG - Date range in {region_lower} database: {date_range[0]} to {date_range[1]}")
            print(f"DEBUG - Querying for dates: {start_date} to {end_date}")
        except Exception as date_err:
            print(f"DEBUG - Could not fetch date range: {date_err}")

        # Normalize dates to ensure proper format (YYYY-MM-DD)
        try:
            # Try to parse and reformat dates
            start_dt = pd.to_datetime(start_date).strftime('%Y-%m-%d')
            end_dt = pd.to_datetime(end_date).strftime('%Y-%m-%d')
            print(f"DEBUG - Normalized dates: {start_dt} to {end_dt}")
        except Exception as date_parse_err:
            print(f"DEBUG - Date parsing error, using original: {date_parse_err}")
            start_dt = start_date
            end_dt = end_date

        # Query data with normalized comparison to capture whitespace/case differences
        if plant_filter:
            print(f"DEBUG - Querying for plants: {norm_plants}")
            query = f"""
                SELECT 
                    region,
                    plant_name,
                    {_date_expr} AS date,
                    time_block,
                    from_time,
                    to_time,
                    avc AS AvC_MW,
                    forecasted_power AS Scheduled_MW,
                    actual_power AS Actual_MW,
                    ppa AS PPA
                FROM {region_lower}
                WHERE {_date_expr} >= DATE '{start_dt}'
                  AND {_date_expr} <= DATE '{end_dt}'
                  AND {_plant_key_expr} IN ({plant_filter})
                ORDER BY plant_name, date, time_block
            """
            print(f"DEBUG - Executing query with {len(norm_plants)} plant(s)")
        else:
            query = f"""
                SELECT 
                    region,
                    plant_name,
                    {_date_expr} AS date,
                    time_block,
                    from_time,
                    to_time,
                    avc AS AvC_MW,
                    forecasted_power AS Scheduled_MW,
                    actual_power AS Actual_MW,
                    ppa AS PPA
                FROM {region_lower}
                WHERE {_date_expr} >= DATE '{start_dt}'
                  AND {_date_expr} <= DATE '{end_dt}'
                ORDER BY plant_name, date, time_block
            """

        df = conn.execute(query).fetchdf()
        print(f"DEBUG - Query returned {len(df)} rows from {region_lower}")
        
        # If no results, try a more lenient query (fuzzy matching)
        if df.empty and plant_filter and len(norm_plants) > 0:
            print(f"DEBUG - No exact matches found. Trying fuzzy matching...")
            # Try LIKE matching for each plant (handles partial matches)
            like_conditions = []
            for plant in norm_plants:
                # Remove quotes for LIKE matching
                plant_clean = plant.replace("''", "'")
                like_conditions.append(f"{_plant_key_expr} LIKE '%{plant_clean}%'")
            
            fuzzy_query = f"""
                SELECT 
                    region,
                    plant_name,
                    date,
                    time_block,
                    from_time,
                    to_time,
                    avc AS AvC_MW,
                    forecasted_power AS Scheduled_MW,
                    actual_power AS Actual_MW,
                    ppa AS PPA
                FROM {region_lower}
                WHERE date >= '{start_dt}'
                  AND date <= '{end_dt}'
                  AND ({' OR '.join(like_conditions)})
                ORDER BY plant_name, date, time_block
            """
            print(f"DEBUG - Trying fuzzy query...")
            df_fuzzy = conn.execute(fuzzy_query).fetchdf()
            if not df_fuzzy.empty:
                print(f"DEBUG - Fuzzy query found {len(df_fuzzy)} rows")
                # Filter to best matches (exact normalized match preferred)
                df_fuzzy['_plant_norm'] = df_fuzzy['plant_name'].str.strip().str.upper()
                best_matches = df_fuzzy[df_fuzzy['_plant_norm'].isin(norm_plants)]
                if not best_matches.empty:
                    df = best_matches.drop(columns=['_plant_norm'])
                    print(f"DEBUG - Found {len(df)} exact normalized matches")
                else:
                    # Use fuzzy matches but warn
                    df = df_fuzzy.drop(columns=['_plant_norm'])
                    print(f"DEBUG - Using {len(df)} fuzzy matches (plant names may not match exactly)")
        
        conn.close()
        
        # Add required columns for analysis
        if not df.empty:
            # Clean up plant_name for downstream grouping/display (aligns with master Excel/UI)
            df["plant_name"] = df["plant_name"].astype(str).map(_norm_plant_name)
            df['Plant'] = df['plant_name']
            df['block'] = df['time_block']
            df['date_time'] = pd.to_datetime(df['date']) + pd.to_timedelta((df['time_block'] - 1) * 15, unit='m')
            # Debug: Show what was actually found
            found_plants = df['plant_name'].unique().tolist()
            print(f"DEBUG - Successfully loaded data for plants: {found_plants}")
        
        return df
        
    except Exception as e:
        import traceback
        print(f"ERROR - Error loading data from {db_file}: {e}")
        traceback.print_exc()
        return pd.DataFrame()

# ==========================================
# -------- REAL DATA LOADING FUNCTIONS --------
# ==========================================

def load_nrpc_data(start_date, end_date, plants):
    """Load NRPC data - Legacy function, now redirects to DuckDB"""
    return load_data_from_duckdb("NRPC", start_date, end_date, plants)

def load_srpc_data(start_date, end_date, plants):
    """Load SRPC data - Legacy function, now redirects to DuckDB"""
    return load_data_from_duckdb("SRPC", start_date, end_date, plants)

def load_wrpc_data(start_date, end_date, plants):
    """Load WRPC data - Legacy function, now redirects to DuckDB"""
    return load_data_from_duckdb("WRPC", start_date, end_date, plants)

def load_region_data(region, start_date, end_date, plants):
    """Load data based on selected region - now uses DuckDB"""
    # Load from DuckDB only; no synthetic fallback so results reflect reality
    df = load_data_from_duckdb(region, start_date, end_date, plants)
    return df

def make_sample_blocks(start_date: str, end_date: str, plant: str) -> pd.DataFrame:
    start = datetime.fromisoformat(start_date)
    end = datetime.fromisoformat(end_date)
    if end < start:
        end = start

    times = pd.date_range(start, end + timedelta(days=1), freq="15min", inclusive="left")
    df = pd.DataFrame({"date_time": times})
    df["Plant"] = plant

    base_avc = 51.3 if "A" in plant or "B" in plant else 75.0
    rng = np.random.default_rng(42)
    df["AvC_MW"] = (base_avc + rng.normal(0, 0.3, len(df))).round(3)

    hours = df["date_time"].dt.hour + df["date_time"].dt.minute / 60.0
    df["Scheduled_MW"] = (
        (0.55 * df["AvC_MW"]) +
        8 * np.sin((hours / 24) * 2 * np.pi) +
        rng.normal(0, 1.2, len(df))
    ).clip(lower=0).round(4)

    bias = np.where((hours >= 13) & (hours <= 16), -6.0, 0.0)
    df["Actual_MW"] = (df["Scheduled_MW"] + bias + rng.normal(0, 2.0, len(df))).clip(lower=0).round(4)

    df["date"] = df["date_time"].dt.date
    daily_ppa = {d: (4.20 if i % 2 == 0 else 3.95) for i, d in enumerate(sorted(df["date"].unique()))}
    df["PPA"] = df["date"].map(daily_ppa)

    df["block"] = df.groupby("date").cumcount() + 1
    return df

# ==========================================
# --------- ENGINE (CALCULATIONS) ----------
# ==========================================
def compute_error_pct(df: pd.DataFrame, mode: str, x_pct: float) -> pd.Series:
    """Compute error % for each row.

    Note: Some call-sites may accidentally pass a Series/dict (single row). We normalize to DataFrame.
    """
    if isinstance(df, pd.Series):
        df = df.to_frame().T
    elif isinstance(df, dict):
        df = pd.DataFrame([df])

    # Robust numeric conversion (always Series)
    actual = pd.to_numeric(df["Actual_MW"] if "Actual_MW" in df else pd.Series(0, index=df.index), errors="coerce").fillna(0.0)
    scheduled = pd.to_numeric(df["Scheduled_MW"] if "Scheduled_MW" in df else pd.Series(0, index=df.index), errors="coerce").fillna(0.0)
    avc = pd.to_numeric(df["AvC_MW"] if "AvC_MW" in df else pd.Series(0, index=df.index), errors="coerce").fillna(0.0)
    avc = avc.replace(0, np.nan)

    num = (actual - scheduled)
    if mode == "dynamic":
        denom = (x_pct / 100.0) * avc + ((100.0 - x_pct) / 100.0) * scheduled
        denom = denom.replace(0, np.nan)
        return (num / denom * 100.0).fillna(0.0)
    denom = avc
    return (num / denom * 100.0).fillna(0.0)

def compute_basis_mw(df: pd.DataFrame, mode: str, x_pct: float) -> pd.Series:
    """Return the MW basis for energy conversion, aligned with regulation mode.
    - default: basis = AvC_MW
    - dynamic: basis = (X%)*AvC + (1-X%)*Scheduled
    """
    if isinstance(df, pd.Series):
        df = df.to_frame().T
    elif isinstance(df, dict):
        df = pd.DataFrame([df])

    avc = pd.to_numeric(df["AvC_MW"] if "AvC_MW" in df else pd.Series(0, index=df.index), errors="coerce").fillna(0.0)
    sched = pd.to_numeric(df["Scheduled_MW"] if "Scheduled_MW" in df else pd.Series(0, index=df.index), errors="coerce").fillna(0.0)
    if mode == "dynamic":
        return (x_pct / 100.0) * avc + ((100.0 - x_pct) / 100.0) * sched
    return avc

# =========================
# New DSM engine helpers (bands parsing, per-slot calc, Excel export)
# =========================
def _normalize_bands_df(bands_df: pd.DataFrame) -> pd.DataFrame:
    """Make bands robust to older saved presets by adding missing columns with defaults."""
    df = bands_df.copy() if isinstance(bands_df, pd.DataFrame) else pd.DataFrame(bands_df or [])
    if "loss_zone" not in df.columns:
        df["loss_zone"] = False
    if "tolerance_cut_pct" not in df.columns:
        df["tolerance_cut_pct"] = 0.0
    if "label" not in df.columns:
        df["label"] = ""
    if "deviated_on" not in df.columns:
        df["deviated_on"] = "AvC"
    # Ensure required core columns exist
    for col in ["direction","lower_pct","upper_pct","rate_type","rate_value","excess_slope_per_pct"]:
        if col not in df.columns:
            df[col] = 0 if col != "direction" and col != "rate_type" else ("UI" if col == "direction" else "flat_per_kwh")
    return df
def parse_bands_from_settings(settings_rows: List[Dict[str, Any]]) -> Tuple[List[Band], pd.DataFrame]:
    """Convert UI bands rows to Band models expected by the new engine.
    - Maps legacy rate_type values to new RATE_* constants
    - Defaults loss_zone to False if not provided
    - Converts flat_per_mwh to per-kWh
    """
    type_map = {
        "flat_per_kwh": RATE_FLAT,
        "ppa_fraction": RATE_FRAC,
        "ppa_multiple": RATE_MULT,
        "flat_per_mwh": RATE_FLAT,
        "scaled_excess": RATE_SCALED,
    }
    out: List[Band] = []
    for r in settings_rows or []:
        legacy_type = str(r.get("rate_type", "")).strip().lower()
        mapped_type = type_map.get(legacy_type, RATE_FLAT)
        raw_rate_value = float(r.get("rate_value", 0) or 0)
        # Convert MWh -> kWh if needed
        rate_value = (raw_rate_value / 1000.0) if legacy_type == "flat_per_mwh" else raw_rate_value
        out.append(Band(
            direction=str(r.get("direction", "")).strip().upper(),
            lower_pct=float(r.get("lower_pct", 0) or 0),
            upper_pct=float(r.get("upper_pct", 0) or 0),
            rate_type=mapped_type,
            rate_value=rate_value,
            rate_slope=float(r.get("excess_slope_per_pct", 0) or 0),
            loss_zone=bool(r.get("loss_zone", False)),
        ))
    # sort by direction then lower_pct
    out.sort(key=lambda b: (b.direction, b.lower_pct, b.upper_pct))
    bands_df = pd.DataFrame([b.__dict__ for b in out])
    return out, bands_df

def compute_slot_row(slot: Dict[str, Any], bands: List[Band], mode: str, dyn_x: float) -> Dict[str, Any]:
    """Return numeric metrics for a single 15-min slot using the new band engine."""
    avc = float(slot["AvC_MW"]) if pd.notna(slot.get("AvC_MW")) else 0.0
    sch = float(slot["Scheduled_MW"]) if pd.notna(slot.get("Scheduled_MW")) else 0.0
    act = float(slot["Actual_MW"]) if pd.notna(slot.get("Actual_MW")) else 0.0
    ppa = float(slot["PPA"]) if pd.notna(slot.get("PPA")) else 0.0

    # HARD SKIP RULE (block-level):
    # If ANY of Scheduled / AvC / PPA is 0, DSM must NOT run for this block.
    if (sch == 0.0) or (avc == 0.0) or (ppa == 0.0):
        return {
            "dsm_skipped_zero_inputs": True,
            "dsm_skip_reason": "ZERO_SCHEDULE_OR_AVC_OR_PPA",
            # numeric outputs as blanks (NaN) so they do not influence totals/averages
            "error_pct": np.nan,
            "direction": "",
            "abs_err": np.nan,
            "band_level": "",
            "UI_Energy_deviation_bands": np.nan,
            "OI_Energy_deviation_bands": np.nan,
            "Revenue_as_per_generation": np.nan,
            "Scheduled_Revenue_as_per_generation": np.nan,
            "UI_DSM": np.nan,
            "OI_DSM": np.nan,
            "OI_Loss": np.nan,
            "Total_DSM": np.nan,
            "Revenue_Loss": np.nan,
        }

    denom = denominator_and_basis(avc, sch, mode, dyn_x)
    err_pct = 0.0 if denom == 0 else (act - sch) / denom * 100.0
    dirn = direction_from(act, sch)
    abs_err = abs(err_pct)

    ui_dev_kwh = 0.0
    oi_dev_kwh = 0.0
    ui_dsm = 0.0
    oi_dsm = 0.0
    oi_loss = 0.0

    for b in bands:
        if b.direction != dirn:
            continue
        sp = slice_pct(abs_err, b.lower_pct, b.upper_pct)
        if sp <= 0:
            continue
        kwh = kwh_from_slice(sp, denom)
        rate = band_rate(ppa, b.rate_type, b.rate_value, b.rate_slope, abs_err)
        amt = kwh * rate
        if dirn == "UI":
            ui_dev_kwh += kwh
            ui_dsm += amt
        elif dirn == "OI":
            oi_dev_kwh += kwh
            if b.loss_zone:
                oi_loss += amt
            else:
                oi_dsm += amt

    rev_act = act * 0.25 * 1000.0 * ppa
    rev_sch = sch * 0.25 * 1000.0 * ppa
    total_dsm = ui_dsm + oi_dsm
    revenue_loss = total_dsm + oi_loss

    reached = [b for b in bands if b.direction == dirn and abs_err >= b.lower_pct]
    band_level = ""
    if reached:
        top = max(reached, key=lambda x: x.upper_pct)
        lo = int(top.lower_pct)
        up = ("" if top.upper_pct >= 999 else int(top.upper_pct))
        band_level = f"{dirn} {lo}–{up}%" if up != "" else f"{dirn} >{lo}%"

    return {
        "dsm_skipped_zero_inputs": False,
        "dsm_skip_reason": "",
        "error_pct": err_pct,
        "direction": dirn,
        "abs_err": abs_err,
        "band_level": band_level,
        "UI_Energy_deviation_bands": ui_dev_kwh,
        "OI_Energy_deviation_bands": oi_dev_kwh,
        "Revenue_as_per_generation": rev_act,
        "Scheduled_Revenue_as_per_generation": rev_sch,
        "UI_DSM": ui_dsm,
        "OI_DSM": oi_dsm,
        "OI_Loss": oi_loss,
        "Total_DSM": total_dsm,
        "Revenue_Loss": revenue_loss,
    }

def xlsx_col(idx: int) -> str:
    s = ""
    n = idx + 1
    while n:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s

def export_with_formulas(detail_rows: pd.DataFrame, bands_df: pd.DataFrame, mode: str, dyn_x: float) -> bytes:
    """Build an Excel file (Bands, Config, Detail) with embedded formulas."""
    output = BytesIO()
    with pd.ExcelWriter(output, engine="xlsxwriter") as xw:
        wb = xw.book
        # Bands
        bands_out = (bands_df or pd.DataFrame(columns=["direction","lower_pct","upper_pct","rate_type","rate_value","rate_slope","loss_zone"]))
        bands_out = bands_out[[
            "direction","lower_pct","upper_pct",
            "rate_type","rate_value","rate_slope","loss_zone"
        ]].reset_index(drop=True)
        bands_out.to_excel(xw, sheet_name="Bands", index=False, startrow=0)
        nrows = len(bands_out)
        def NR(col_letter: str) -> str:
            return f"Bands!${col_letter}$2:${col_letter}${nrows+1}" if nrows > 0 else f"Bands!${col_letter}$2:${col_letter}$2"
        wb.define_name("Bands_Dir",      f"={NR('A')}")
        wb.define_name("Bands_Lower",    f"={NR('B')}")
        wb.define_name("Bands_Upper",    f"={NR('C')}")
        wb.define_name("Bands_RateType", f"={NR('D')}")
        wb.define_name("Bands_RateVal",  f"={NR('E')}")
        wb.define_name("Bands_RateSlope",f"={NR('F')}")
        wb.define_name("Bands_LossZone", f"={NR('G')}")
        # Config
        cfg = pd.DataFrame({"Key": ["MODE","DYN_X"], "Value": [mode, dyn_x]})
        cfg.to_excel(xw, sheet_name="Config", index=False)
        wb.define_name("CFG_MODE", "=Config!$B$2")
        wb.define_name("CFG_DYNX", "=Config!$B$3")
        # Detail
        ws = wb.add_worksheet("Detail")
        headers = [
            "region","plant_name","date","time_block","from_time","to_time",
            "AvC_MW","Scheduled_MW","Actual_MW","PPA",
            "error_pct","direction","abs_err","band_level",
            "UI_Energy_deviation_bands","OI_Energy_deviation_bands",
            "Revenue_as_per_generation","Scheduled_Revenue_as_per_generation",
            "UI_DSM","OI_DSM","OI_Loss","Total_DSM","Revenue_Loss"
        ]
        for c, h in enumerate(headers):
            ws.write(0, c, h)
        COL = {h: i for i, h in enumerate(headers)}
        BASIS_COL = len(headers)
        ws.write(0, BASIS_COL, "_basis_helper")
        start_row = 1
        for r, row in enumerate(detail_rows.itertuples(index=False), start=start_row):
            ws.write(r, COL["region"],       getattr(row, "region"))
            ws.write(r, COL["plant_name"],   getattr(row, "plant_name"))
            ws.write(r, COL["date"],         getattr(row, "date"))
            ws.write(r, COL["time_block"],   getattr(row, "time_block"))
            ws.write(r, COL["from_time"],    getattr(row, "from_time"))
            ws.write(r, COL["to_time"],      getattr(row, "to_time"))
            avc_v = getattr(row, "AvC_MW", None)
            sch_v = getattr(row, "Scheduled_MW", None)
            act_v = getattr(row, "Actual_MW", None)
            ppa_v = getattr(row, "PPA", None)

            # HARD SKIP RULE for Excel output:
            # If schedule/avc/ppa is blank or 0, DO NOT write any formulas or derived values for that row.
            def _is_zero_or_blank(v) -> bool:
                try:
                    if v is None or (isinstance(v, float) and math.isnan(v)):
                        return True
                    return float(v) == 0.0
                except Exception:
                    return True

            is_skipped = _is_zero_or_blank(sch_v) or _is_zero_or_blank(avc_v) or _is_zero_or_blank(ppa_v)

            # Input columns must be blank for skipped rows
            if is_skipped:
                ws.write_blank(r, COL["AvC_MW"], None)
                ws.write_blank(r, COL["Scheduled_MW"], None)
                ws.write_blank(r, COL["Actual_MW"], None)
                ws.write_blank(r, COL["PPA"], None)
                # Blank all derived columns too (no DSM run)
                for h in [
                    "error_pct","direction","abs_err","band_level",
                    "UI_Energy_deviation_bands","OI_Energy_deviation_bands",
                    "Revenue_as_per_generation","Scheduled_Revenue_as_per_generation",
                    "UI_DSM","OI_DSM","OI_Loss","Total_DSM","Revenue_Loss"
                ]:
                    ws.write_blank(r, COL[h], None)
                ws.write_blank(r, BASIS_COL, None)
                continue

            # Normal rows: write inputs and formulas
            ws.write_number(r, COL["AvC_MW"],       float(avc_v))
            ws.write_number(r, COL["Scheduled_MW"], float(sch_v))
            ws.write_number(r, COL["Actual_MW"],    float(act_v))
            ws.write_number(r, COL["PPA"],          float(ppa_v))
            row1 = r + 1
            avc_ref = f"{xlsx_col(COL['AvC_MW'])}{row1}"
            sch_ref = f"{xlsx_col(COL['Scheduled_MW'])}{row1}"
            act_ref = f"{xlsx_col(COL['Actual_MW'])}{row1}"
            ppa_ref = f"{xlsx_col(COL['PPA'])}{row1}"
            basis_formula = f'=IF(CFG_MODE="DEFAULT",{avc_ref}, CFG_DYNX*{avc_ref} + (1-CFG_DYNX)*{sch_ref})'
            ws.write_formula(r, BASIS_COL, basis_formula)
            basis_ref = f"{xlsx_col(BASIS_COL)}{row1}"
            err_formula = f"=IF({basis_ref}=0,0, ({act_ref}-{sch_ref})/{basis_ref}*100)"
            ws.write_formula(r, COL["error_pct"], err_formula)
            dir_formula = f'=IF({act_ref}<{sch_ref},"UI",IF({act_ref}>{sch_ref},"OI","FLAT"))'
            ws.write_formula(r, COL["direction"], dir_formula)
            abs_formula = f"=ABS({xlsx_col(COL['error_pct'])}{row1})"
            ws.write_formula(r, COL["abs_err"], abs_formula)
            ws.write(r, COL["band_level"], "")
            ui_dev_formula = f'=IF({xlsx_col(COL["direction"])}{row1}="UI",{xlsx_col(COL["abs_err"])}{row1}/100*{basis_ref}*0.25*1000,0)'
            oi_dev_formula = f'=IF({xlsx_col(COL["direction"])}{row1}="OI",{xlsx_col(COL["abs_err"])}{row1}/100*{basis_ref}*0.25*1000,0)'
            ws.write_formula(r, COL["UI_Energy_deviation_bands"], ui_dev_formula)
            ws.write_formula(r, COL["OI_Energy_deviation_bands"], oi_dev_formula)
            rev_act_formula = f"={act_ref}*0.25*1000*{ppa_ref}"
            rev_sch_formula = f"={sch_ref}*0.25*1000*{ppa_ref}"
            ws.write_formula(r, COL["Revenue_as_per_generation"], rev_act_formula)
            ws.write_formula(r, COL["Scheduled_Revenue_as_per_generation"], rev_sch_formula)
            abs_ref = f"{xlsx_col(COL['abs_err'])}{row1}"
            slice_factor = f"MAX(0, MIN({abs_ref}, Bands_Upper) - Bands_Lower)/100 * {basis_ref} * 0.25*1000"
            rate_expr = (
                'IF(Bands_RateType="PPA_FRAC", ' + ppa_ref + '*Bands_RateVal, '
                'IF(Bands_RateType="PPA_MULT", ' + ppa_ref + '*Bands_RateVal, '
                'IF(Bands_RateType="FLAT", Bands_RateVal, '
                'IF(Bands_RateType="SCALED", Bands_RateVal + Bands_RateSlope*' + abs_ref + ', 0))))'
            )
            ui_sp = '=SUMPRODUCT(--(Bands_Dir="UI"),' + slice_factor + ',' + rate_expr + ')'
            oi_dsm_sp = '=SUMPRODUCT(--(Bands_Dir="OI"),--(Bands_LossZone=FALSE),' + slice_factor + ',' + rate_expr + ')'
            oi_loss_sp = '=SUMPRODUCT(--(Bands_Dir="OI"),--(Bands_LossZone=TRUE),' + slice_factor + ',' + rate_expr + ')'
            ws.write_formula(r, COL["UI_DSM"], ui_sp)
            ws.write_formula(r, COL["OI_DSM"], oi_dsm_sp)
            ws.write_formula(r, COL["OI_Loss"], oi_loss_sp)
            ws.write_formula(r, COL["Total_DSM"], f"={xlsx_col(COL['UI_DSM'])}{row1}+{xlsx_col(COL['OI_DSM'])}{row1}")
            ws.write_formula(r, COL["Revenue_Loss"], f"={xlsx_col(COL['Total_DSM'])}{row1}+{xlsx_col(COL['OI_Loss'])}{row1}")
        ws.set_column(BASIS_COL, BASIS_COL, None, None, {'hidden': True})
    output.seek(0)
    return output.getvalue()

def export_with_formulas_openpyxl(detail_rows: pd.DataFrame, bands_df: pd.DataFrame, mode: str, dyn_x: float) -> bytes:
    """OpenPyXL implementation of the same formula-driven workbook with named ranges."""
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    # Try to import DefinedName (optional). If unavailable, formulas will use absolute refs instead of names.
    try:
        from openpyxl.workbook.defined_name import DefinedName  # type: ignore
    except Exception:
        DefinedName = None  # type: ignore

    wb = Workbook()
    # Clear default sheet
    ws_default = wb.active
    wb.remove(ws_default)

    # Bands sheet
    ws_b = wb.create_sheet("Bands")
    bands_out = bands_df.copy().reset_index(drop=True)
    cols = ["direction","lower_pct","upper_pct","rate_type","rate_value","rate_slope","loss_zone"]
    for c_idx, h in enumerate(cols, start=1):
        ws_b.cell(row=1, column=c_idx, value=h)
    for r_idx, row in enumerate(bands_out.itertuples(index=False), start=2):
        ws_b.cell(row=r_idx, column=1, value=getattr(row, "direction"))
        ws_b.cell(row=r_idx, column=2, value=float(getattr(row, "lower_pct")))
        ws_b.cell(row=r_idx, column=3, value=float(getattr(row, "upper_pct")))
        ws_b.cell(row=r_idx, column=4, value=str(getattr(row, "rate_type")))
        ws_b.cell(row=r_idx, column=5, value=float(getattr(row, "rate_value")))
        ws_b.cell(row=r_idx, column=6, value=float(getattr(row, "rate_slope")))
        ws_b.cell(row=r_idx, column=7, value=bool(getattr(row, "loss_zone")))
    nrows = len(bands_out)
    last_row = 1 + (nrows if nrows > 0 else 1)
    def ref(col):
        return f"Bands!${col}$2:${col}${last_row}"
    # Named ranges (best-effort)
    if DefinedName is not None:
        wb.defined_names.append(DefinedName(name="Bands_Dir", attr_text=ref('A')))
        wb.defined_names.append(DefinedName(name="Bands_Lower", attr_text=ref('B')))
        wb.defined_names.append(DefinedName(name="Bands_Upper", attr_text=ref('C')))
        wb.defined_names.append(DefinedName(name="Bands_RateType", attr_text=ref('D')))
        wb.defined_names.append(DefinedName(name="Bands_RateVal", attr_text=ref('E')))
        wb.defined_names.append(DefinedName(name="Bands_RateSlope", attr_text=ref('F')))
        wb.defined_names.append(DefinedName(name="Bands_LossZone", attr_text=ref('G')))
    # Absolute ranges (always available)
    BANDS_DIR = ref('A')
    BANDS_LOWER = ref('B')
    BANDS_UPPER = ref('C')
    BANDS_RATETYPE = ref('D')
    BANDS_RATEVAL = ref('E')
    BANDS_RATESLOPE = ref('F')
    BANDS_LOSSZONE = ref('G')

    # Config
    ws_c = wb.create_sheet("Config")
    ws_c.cell(row=1, column=1, value="Key")
    ws_c.cell(row=1, column=2, value="Value")
    ws_c.cell(row=2, column=1, value="MODE")
    ws_c.cell(row=2, column=2, value=mode)
    ws_c.cell(row=3, column=1, value="DYN_X")
    ws_c.cell(row=3, column=2, value=float(dyn_x))
    # Named cells (best-effort)
    if DefinedName is not None:
        wb.defined_names.append(DefinedName(name="CFG_MODE", attr_text="Config!$B$2"))
        wb.defined_names.append(DefinedName(name="CFG_DYNX", attr_text="Config!$B$3"))
    CFG_MODE_REF = "Config!$B$2"
    CFG_DYNX_REF = "Config!$B$3"

    # Detail with formulas
    ws = wb.create_sheet("Detail")
    headers = [
        "region","plant_name","date","time_block","from_time","to_time",
        "AvC_MW","Scheduled_MW","Actual_MW","PPA",
        "error_pct","direction","abs_err","band_level",
        "UI_Energy_deviation_bands","OI_Energy_deviation_bands",
        "Revenue_as_per_generation","Scheduled_Revenue_as_per_generation",
        "UI_DSM","OI_DSM","OI_Loss","Total_DSM","Revenue_Loss"
    ]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    col_index = {h: i+1 for i, h in enumerate(headers)}
    BASIS_COL = len(headers) + 1
    ws.cell(row=1, column=BASIS_COL, value="_basis_helper")

    for r_idx, row in enumerate(detail_rows.itertuples(index=False), start=2):
        ws.cell(row=r_idx, column=col_index["region"], value=getattr(row, "region"))
        ws.cell(row=r_idx, column=col_index["plant_name"], value=getattr(row, "plant_name"))
        ws.cell(row=r_idx, column=col_index["date"], value=getattr(row, "date"))
        ws.cell(row=r_idx, column=col_index["time_block"], value=getattr(row, "time_block"))
        ws.cell(row=r_idx, column=col_index["from_time"], value=getattr(row, "from_time"))
        ws.cell(row=r_idx, column=col_index["to_time"], value=getattr(row, "to_time"))
        avc_v = getattr(row, "AvC_MW", None)
        sch_v = getattr(row, "Scheduled_MW", None)
        act_v = getattr(row, "Actual_MW", None)
        ppa_v = getattr(row, "PPA", None)

        def _is_zero_or_blank(v) -> bool:
            try:
                if v is None or (isinstance(v, float) and math.isnan(v)):
                    return True
                return float(v) == 0.0
            except Exception:
                return True

        is_skipped = _is_zero_or_blank(sch_v) or _is_zero_or_blank(avc_v) or _is_zero_or_blank(ppa_v)

        if is_skipped:
            # Leave input and derived cells blank (DSM did not run for this block)
            # (openpyxl uses None for blank cell)
            ws.cell(row=r_idx, column=col_index["AvC_MW"], value=None)
            ws.cell(row=r_idx, column=col_index["Scheduled_MW"], value=None)
            ws.cell(row=r_idx, column=col_index["Actual_MW"], value=None)
            ws.cell(row=r_idx, column=col_index["PPA"], value=None)
            for h in [
                "error_pct","direction","abs_err","band_level",
                "UI_Energy_deviation_bands","OI_Energy_deviation_bands",
                "Revenue_as_per_generation","Scheduled_Revenue_as_per_generation",
                "UI_DSM","OI_DSM","OI_Loss","Total_DSM","Revenue_Loss"
            ]:
                ws.cell(row=r_idx, column=col_index[h], value=None)
            ws.cell(row=r_idx, column=BASIS_COL, value=None)
            continue

        ws.cell(row=r_idx, column=col_index["AvC_MW"], value=float(avc_v))
        ws.cell(row=r_idx, column=col_index["Scheduled_MW"], value=float(sch_v))
        ws.cell(row=r_idx, column=col_index["Actual_MW"], value=float(act_v))
        ws.cell(row=r_idx, column=col_index["PPA"], value=float(ppa_v))

        row1 = r_idx
        def A1(col_name):
            return f"{get_column_letter(col_index[col_name])}{row1}"
        avc_ref = A1("AvC_MW")
        sch_ref = A1("Scheduled_MW")
        act_ref = A1("Actual_MW")
        ppa_ref = A1("PPA")
        basis_ref = f"{get_column_letter(BASIS_COL)}{row1}"

        ws.cell(row=row1, column=BASIS_COL, value=f'=IF({CFG_MODE_REF}="DEFAULT",{avc_ref}, {CFG_DYNX_REF}*{avc_ref} + (1-{CFG_DYNX_REF})*{sch_ref})')
        ws.cell(row=row1, column=col_index["error_pct"], value=f"=IF({basis_ref}=0,0, ({act_ref}-{sch_ref})/{basis_ref}*100)")
        ws.cell(row=row1, column=col_index["direction"], value=f'=IF({act_ref}<{sch_ref},"UI",IF({act_ref}>{sch_ref},"OI","FLAT"))')
        ws.cell(row=row1, column=col_index["abs_err"], value=f"=ABS({A1('error_pct')})")
        # band_level left blank
        ws.cell(row=row1, column=col_index["UI_Energy_deviation_bands"], value=f'=IF({A1("direction")}="UI",{A1("abs_err")}/100*{basis_ref}*0.25*1000,0)')
        ws.cell(row=row1, column=col_index["OI_Energy_deviation_bands"], value=f'=IF({A1("direction")}="OI",{A1("abs_err")}/100*{basis_ref}*0.25*1000,0)')
        ws.cell(row=row1, column=col_index["Revenue_as_per_generation"], value=f"={act_ref}*0.25*1000*{ppa_ref}")
        ws.cell(row=row1, column=col_index["Scheduled_Revenue_as_per_generation"], value=f"={sch_ref}*0.25*1000*{ppa_ref}")

        abs_ref = A1("abs_err")
        slice_factor = (
            f"MAX(0, MIN({abs_ref}, {BANDS_UPPER}) - {BANDS_LOWER})/100 * {basis_ref} * 0.25*1000"
        )
        rate_expr = (
            'IF(' + BANDS_RATETYPE + '="PPA_FRAC", ' + ppa_ref + '*' + BANDS_RATEVAL + ', '
            'IF(' + BANDS_RATETYPE + '="PPA_MULT", ' + ppa_ref + '*' + BANDS_RATEVAL + ', '
            'IF(' + BANDS_RATETYPE + '="FLAT", ' + BANDS_RATEVAL + ', '
            'IF(' + BANDS_RATETYPE + '="SCALED", ' + BANDS_RATEVAL + ' + ' + BANDS_RATESLOPE + '*' + abs_ref + ', 0))))'
        )
        ui_sp = '=SUMPRODUCT(--(' + BANDS_DIR + '="UI"),' + slice_factor + ',' + rate_expr + ')'
        oi_dsm_sp = '=SUMPRODUCT(--(' + BANDS_DIR + '="OI"),--(' + BANDS_LOSSZONE + '=FALSE),' + slice_factor + ',' + rate_expr + ')'
        oi_loss_sp = '=SUMPRODUCT(--(' + BANDS_DIR + '="OI"),--(' + BANDS_LOSSZONE + '=TRUE),' + slice_factor + ',' + rate_expr + ')'
        ws.cell(row=row1, column=col_index["UI_DSM"], value=ui_sp)
        ws.cell(row=row1, column=col_index["OI_DSM"], value=oi_dsm_sp)
        ws.cell(row=row1, column=col_index["OI_Loss"], value=oi_loss_sp)
        ws.cell(row=row1, column=col_index["Total_DSM"], value=f"={get_column_letter(col_index['UI_DSM'])}{row1}+{get_column_letter(col_index['OI_DSM'])}{row1}")
        ws.cell(row=row1, column=col_index["Revenue_Loss"], value=f"={get_column_letter(col_index['Total_DSM'])}{row1}+{get_column_letter(col_index['OI_Loss'])}{row1}")

    # Hide basis column
    ws.column_dimensions[get_column_letter(BASIS_COL)].hidden = True

    from io import BytesIO
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()

def export_with_formulas_opc(detail_rows: pd.DataFrame, bands_df: pd.DataFrame, mode: str, dyn_x: float) -> bytes:
    """Create Excel workbook (Detail/Bands/Config) with formulas using only stdlib (zip + XML)."""
    from zipfile import ZipFile, ZIP_DEFLATED
    from datetime import datetime
    import xml.sax.saxutils as saxutils

    headers = [
        "region","plant_name","date","time_block","from_time","to_time",
        "AvC_MW","Scheduled_MW","Actual_MW","PPA",
        "error_pct","direction","abs_err","band_level",
        "UI_Energy_deviation_bands","OI_Energy_deviation_bands",
        "Revenue_as_per_generation","Scheduled_Revenue_as_per_generation",
        "UI_DSM","OI_DSM","OI_Loss","Total_DSM","Revenue_Loss"
    ]
    basis_col_num = len(headers) + 1
    basis_col_letter = xlsx_col(basis_col_num - 1)

    def coord(col_num: int, row_num: int) -> str:
        return f"{xlsx_col(col_num - 1)}{row_num}"

    def cell_inline(col_num: int, row_num: int, text: str) -> str:
        txt = saxutils.escape(text or "")
        return f'<c r="{coord(col_num, row_num)}" t="inlineStr"><is><t>{txt}</t></is></c>'

    def cell_number(col_num: int, row_num: int, value) -> str:
        val = "" if value is None else ("0" if value == 0 else f"{value}")
        return f'<c r="{coord(col_num, row_num)}"><v>{val}</v></c>'

    def cell_formula(col_num: int, row_num: int, formula: str) -> str:
        frm = saxutils.escape(formula)
        return f'<c r="{coord(col_num, row_num)}"><f>{frm}</f></c>'

    # Bands ranges for formulas
    bands_len = len(bands_df)
    bands_last = 1 + (bands_len if bands_len > 0 else 1)
    band_range = lambda col_letter: f'Bands!${col_letter}$2:${col_letter}${bands_last}'
    bands_dir_ref = band_range('A')
    bands_lower_ref = band_range('B')
    bands_upper_ref = band_range('C')
    bands_ratetype_ref = band_range('D')
    bands_rateval_ref = band_range('E')
    bands_rateslope_ref = band_range('F')
    bands_losszone_ref = band_range('G')

    cfg_mode_ref = 'Config!$B$2'
    cfg_dnx_ref = 'Config!$B$3'

    rows_xml = []
    header_cells = [cell_inline(idx + 1, 1, hdr) for idx, hdr in enumerate(headers)]
    header_cells.append(cell_inline(basis_col_num, 1, '_basis_helper'))
    rows_xml.append(f'<row r="1">{"".join(header_cells)}</row>')

    for idx, row in enumerate(detail_rows.itertuples(index=False), start=2):
        cells = []
        cells.append(cell_inline(1, idx, str(getattr(row, 'region', ''))))
        cells.append(cell_inline(2, idx, str(getattr(row, 'plant_name', ''))))
        cells.append(cell_inline(3, idx, str(getattr(row, 'date', ''))))
        cells.append(cell_number(4, idx, getattr(row, 'time_block', '')))
        cells.append(cell_inline(5, idx, str(getattr(row, 'from_time', ''))))
        cells.append(cell_inline(6, idx, str(getattr(row, 'to_time', ''))))
        cells.append(cell_number(7, idx, float(getattr(row, 'AvC_MW'))))
        cells.append(cell_number(8, idx, float(getattr(row, 'Scheduled_MW'))))
        cells.append(cell_number(9, idx, float(getattr(row, 'Actual_MW'))))
        cells.append(cell_number(10, idx, float(getattr(row, 'PPA'))))

        avc_ref = coord(7, idx)
        sch_ref = coord(8, idx)
        act_ref = coord(9, idx)
        ppa_ref = coord(10, idx)
        basis_ref = f'${basis_col_letter}${idx}'
        error_ref = coord(11, idx)
        direction_ref = coord(12, idx)
        abs_ref = coord(13, idx)

        basis_formula = f'IF({cfg_mode_ref}="DEFAULT",{avc_ref}, {cfg_dnx_ref}*{avc_ref} + (1-{cfg_dnx_ref})*{sch_ref})'
        cells.append(cell_formula(basis_col_num, idx, basis_formula))
        cells.append(cell_formula(11, idx, f'IF({basis_ref}=0,0, ({act_ref}-{sch_ref})/{basis_ref}*100)'))
        cells.append(cell_formula(12, idx, f'IF({act_ref}<{sch_ref},"UI",IF({act_ref}>{sch_ref},"OI","FLAT"))'))
        cells.append(cell_formula(13, idx, f'ABS({error_ref})'))
        cells.append(cell_inline(14, idx, ""))  # band_level blank
        cells.append(cell_formula(15, idx, f'IF({direction_ref}="UI",{abs_ref}/100*{basis_ref}*0.25*1000,0)'))
        cells.append(cell_formula(16, idx, f'IF({direction_ref}="OI",{abs_ref}/100*{basis_ref}*0.25*1000,0)'))
        cells.append(cell_formula(17, idx, f'{act_ref}*0.25*1000*{ppa_ref}'))
        cells.append(cell_formula(18, idx, f'{sch_ref}*0.25*1000*{ppa_ref}'))

        slice_factor = f'MAX(0, MIN({abs_ref}, {bands_upper_ref}) - {bands_lower_ref})/100 * {basis_ref} * 0.25*1000'
        rate_expr = (
            f'IF({bands_ratetype_ref}="PPA_FRAC", {ppa_ref}*{bands_rateval_ref}, '
            f'IF({bands_ratetype_ref}="PPA_MULT", {ppa_ref}*{bands_rateval_ref}, '
            f'IF({bands_ratetype_ref}="FLAT", {bands_rateval_ref}, '
            f'IF({bands_ratetype_ref}="SCALED", {bands_rateval_ref} + {bands_rateslope_ref}*{abs_ref}, 0))))'
        )
        cells.append(cell_formula(19, idx, f'SUMPRODUCT(--({bands_dir_ref}="UI"),{slice_factor},{rate_expr})'))
        cells.append(cell_formula(20, idx, f'SUMPRODUCT(--({bands_dir_ref}="OI"),--({bands_losszone_ref}=FALSE),{slice_factor},{rate_expr})'))
        cells.append(cell_formula(21, idx, f'SUMPRODUCT(--({bands_dir_ref}="OI"),--({bands_losszone_ref}=TRUE),{slice_factor},{rate_expr})'))
        ui_ref = coord(19, idx)
        oi_ref = coord(20, idx)
        oi_loss_ref = coord(21, idx)
        cells.append(cell_formula(22, idx, f'{ui_ref}+{oi_ref}'))
        cells.append(cell_formula(23, idx, f'{coord(22, idx)}+{oi_loss_ref}'))

        rows_xml.append(f'<row r="{idx}">{"".join(cells)}</row>')

    cols_xml = f'<cols><col min="{basis_col_num}" max="{basis_col_num}" hidden="1" width="0"/></cols>'
    sheet_data_xml = ''.join(rows_xml)
    detail_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        f'{cols_xml}<sheetData>{sheet_data_xml}</sheetData></worksheet>'
    )

    # Bands sheet
    bands_rows = []
    band_headers = ["direction","lower_pct","upper_pct","rate_type","rate_value","rate_slope","loss_zone"]
    bands_header_cells = [cell_inline(i+1, 1, h) for i, h in enumerate(band_headers)]
    bands_rows.append(f'<row r="1">{"".join(bands_header_cells)}</row>')
    for idx, row in enumerate(bands_df.reset_index(drop=True).itertuples(index=False), start=2):
        cells = [
            cell_inline(1, idx, str(getattr(row, 'direction'))),
            cell_number(2, idx, float(getattr(row, 'lower_pct'))),
            cell_number(3, idx, float(getattr(row, 'upper_pct'))),
            cell_inline(4, idx, str(getattr(row, 'rate_type'))),
            cell_number(5, idx, float(getattr(row, 'rate_value'))),
            cell_number(6, idx, float(getattr(row, 'rate_slope'))),
            cell_inline(7, idx, 'TRUE' if bool(getattr(row, 'loss_zone')) else 'FALSE'),
        ]
        bands_rows.append(f'<row r="{idx}">{"".join(cells)}</row>')
    bands_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        f'<sheetData>{"".join(bands_rows)}</sheetData></worksheet>'
    )

    # Config sheet
    config_rows = [
        '<row r="1">' + cell_inline(1,1,"Key") + cell_inline(2,1,"Value") + '</row>',
        '<row r="2">' + cell_inline(1,2,"MODE") + cell_inline(2,2,str(mode)) + '</row>',
        '<row r="3">' + cell_inline(1,3,"DYN_X") + cell_number(2,3,float(dyn_x)) + '</row>',
    ]
    config_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        f'<sheetData>{"".join(config_rows)}</sheetData></worksheet>'
    )

    # Styles (minimal)
    styles_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        '<fonts count="1"><font><sz val="11"/><color theme="1"/><name val="Calibri"/><family val="2"/></font></fonts>'
        '<fills count="1"><fill><patternFill patternType="none"/></fill></fills>'
        '<borders count="1"><border><left/><right/><top/><bottom/><diagonal/></border></borders>'
        '<cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>'
        '<cellXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/></cellXfs>'
        '</styleSheet>'
    )

    # Workbook + rels
    workbook_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        '<sheets>'
        '<sheet name="Detail" sheetId="1" r:id="rId1"/>'
        '<sheet name="Bands" sheetId="2" r:id="rId2"/>'
        '<sheet name="Config" sheetId="3" r:id="rId3"/>'
        '</sheets>'
        '</workbook>'
    )

    workbook_rels_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>'
        '<Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet2.xml"/>'
        '<Relationship Id="rId3" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet3.xml"/>'
        '<Relationship Id="rId4" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/>'
        '</Relationships>'
    )

    content_types_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
        '<Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
        '<Override PartName="/xl/worksheets/sheet2.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
        '<Override PartName="/xl/worksheets/sheet3.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
        '<Override PartName="/xl/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/>'
        '<Override PartName="/docProps/core.xml" ContentType="application/vnd.openxmlformats-package.core-properties+xml"/>'
        '<Override PartName="/docProps/app.xml" ContentType="application/vnd.openxmlformats-officedocument.extended-properties+xml"/>'
        '</Types>'
    )

    root_rels_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>'
        '<Relationship Id="rId2" Type="http://schemas.openxmlformats.org/package/2006/relationships/metadata/core-properties" Target="docProps/core.xml"/>'
        '<Relationship Id="rId3" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/extended-properties" Target="docProps/app.xml"/>'
        '</Relationships>'
    )

    # docProps
    created_ts = datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%SZ')
    core_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<cp:coreProperties xmlns:cp="http://schemas.openxmlformats.org/package/2006/metadata/core-properties" '
        'xmlns:dc="http://purl.org/dc/elements/1.1/" xmlns:dcterms="http://purl.org/dc/terms/" '
        'xmlns:dcmitype="http://purl.org/dc/dcmitype/" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">'
        f'<dcterms:created xsi:type="dcterms:W3CDTF">{created_ts}</dcterms:created>'
        '<dc:creator>DSM Dashboard</dc:creator>'
        '<cp:lastModifiedBy>DSM Dashboard</cp:lastModifiedBy>'
        f'<dcterms:modified xsi:type="dcterms:W3CDTF">{created_ts}</dcterms:modified>'
        '</cp:coreProperties>'
    )

    app_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Properties xmlns="http://schemas.openxmlformats.org/officeDocument/2006/extended-properties" '
        'xmlns:vt="http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes">'
        '<Application>DSM Dashboard</Application>'
        '<DocSecurity>0</DocSecurity>'
        '<ScaleCrop>0</ScaleCrop>'
        '<HeadingPairs><vt:vector size="2" baseType="variant">'
        '<vt:variant><vt:lpstr>Worksheets</vt:lpstr></vt:variant>'
        '<vt:variant><vt:i4>3</vt:i4></vt:variant>'
        '</vt:vector></HeadingPairs>'
        '<TitlesOfParts><vt:vector size="3" baseType="lpstr">'
        '<vt:lpstr>Detail</vt:lpstr><vt:lpstr>Bands</vt:lpstr><vt:lpstr>Config</vt:lpstr>'
        '</vt:vector></TitlesOfParts>'
        '</Properties>'
    )

    out = BytesIO()
    with ZipFile(out, 'w', ZIP_DEFLATED) as zf:
        zf.writestr('[Content_Types].xml', content_types_xml)
        zf.writestr('_rels/.rels', root_rels_xml)
        zf.writestr('docProps/core.xml', core_xml)
        zf.writestr('docProps/app.xml', app_xml)
        zf.writestr('xl/workbook.xml', workbook_xml)
        zf.writestr('xl/_rels/workbook.xml.rels', workbook_rels_xml)
        zf.writestr('xl/styles.xml', styles_xml)
        zf.writestr('xl/worksheets/sheet1.xml', detail_xml)
        zf.writestr('xl/worksheets/sheet2.xml', bands_xml)
        zf.writestr('xl/worksheets/sheet3.xml', config_xml)
    out.seek(0)
    return out.getvalue()

def build_summary_for_screen(detail_numeric_df: pd.DataFrame) -> Dict[str, Any]:
    avc_mode = safe_mode(detail_numeric_df["AvC_MW"].tolist()) if "AvC_MW" in detail_numeric_df.columns else 0.0
    ppa_mode = safe_mode(detail_numeric_df["PPA"].tolist()) if "PPA" in detail_numeric_df.columns else 0.0
    rev_loss_sum = float(detail_numeric_df.get("Revenue_Loss", pd.Series(dtype=float)).sum())
    rev_act_sum = float(detail_numeric_df.get("Revenue_as_per_generation", pd.Series(dtype=float)).sum())
    dsm_loss_sum = float(detail_numeric_df.get("Total_DSM", pd.Series(dtype=float)).sum())
    rev_loss_pct = 0.0 if rev_act_sum == 0 else (rev_loss_sum / rev_act_sum) * 100.0
    return {
        "plant_capacity_mode_AvC": avc_mode,
        "ppa_mode": ppa_mode,
        "revenue_loss_pct": rev_loss_pct,
        "dsm_loss": dsm_loss_sum,
    }


def apply_bands(df: pd.DataFrame, bands: list[dict], unpaid_oi_threshold: float = 15.0) -> pd.DataFrame:
    out = df.copy()

    # HARD SKIP RULE (block-level):
    # If ANY of Scheduled / AvC / PPA is 0, DSM must NOT run for that block.
    sch0 = pd.to_numeric(out.get("Scheduled_MW"), errors="coerce").fillna(0.0) == 0.0
    avc0 = pd.to_numeric(out.get("AvC_MW"), errors="coerce").fillna(0.0) == 0.0
    ppa0 = pd.to_numeric(out.get("PPA"), errors="coerce").fillna(0.0) == 0.0
    skip_mask = sch0 | avc0 | ppa0
    out["dsm_skipped_zero_inputs"] = skip_mask
    out["dsm_skip_reason"] = np.where(skip_mask, "ZERO_SCHEDULE_OR_AVC_OR_PPA", "")

    # NOTE: We do NOT blank input columns (Scheduled_MW, AvC_MW, Actual_MW, PPA) here
    # because Plant Summary needs them for aggregation (median). Excel export will blank them.
    # We only blank derived/computed fields for skipped blocks:
    if skip_mask.any():
        for c in ("error_pct", "basis_MW"):
            if c in out.columns:
                out.loc[skip_mask, c] = np.nan

    out["direction"] = np.where(out["error_pct"] < 0, "UI", "OI")
    out.loc[skip_mask, "direction"] = ""
    out["abs_err"] = out["error_pct"].abs()
    out.loc[skip_mask, "abs_err"] = np.nan

    for col in ["penalty", "deviation_payable", "receivable", "drawl"]:
        out[col] = 0.0
        out.loc[skip_mask, col] = np.nan

    out["band_label"] = ""
    out.loc[skip_mask, "band_label"] = ""
    out["dev_pct"] = 0.0
    out.loc[skip_mask, "dev_pct"] = np.nan
    out["deviated_kWh"] = 0.0
    out.loc[skip_mask, "deviated_kWh"] = np.nan
    out["rate_applied"] = 0.0
    out.loc[skip_mask, "rate_applied"] = np.nan

    # Determine tolerance end (Band-1 upper) per direction; charging starts above it
    ui_bands = [b for b in bands if b.get("direction") == "UI"]
    oi_bands = [b for b in bands if b.get("direction") == "OI"]
    def tolerance_end(dir_bands: list[dict]) -> float:
        if not dir_bands:
            return 0.0
        # Prefer the band that starts at or near 0; fallback to minimum upper
        zero_start = [float(b.get("upper_pct", 0)) for b in dir_bands if float(b.get("lower_pct", 0)) <= 0.0]
        if zero_start:
            return min(zero_start)
        return min(float(b.get("upper_pct", 0)) for b in dir_bands)
    tol_ui = tolerance_end(ui_bands)
    tol_oi = tolerance_end(oi_bands)

    # Prepare per-band slice columns (kWh) aggregation
    ordered = sorted(bands, key=lambda b: (b.get("direction", ""), float(b.get("lower_pct", 0.0))))
    slice_cols_meta = []  # (colname, label, direction, lower_pct)
    for b in ordered:
        label = b.get("label", f"{b.get('direction','')}{b.get('lower_pct','')}-{b.get('upper_pct','')}")
        safe = "slice_kWh_" + re.sub(r"[^0-9A-Za-z_]+", "_", label.replace(" ", "_"))
        if safe not in out.columns:
            out[safe] = 0.0
        slice_cols_meta.append((safe, label, b.get("direction",""), float(b.get("lower_pct", 0.0))))

    # Assign band_label as the bracket where D falls
    for b in ordered:
        dirn = b["direction"]
        lower = float(b["lower_pct"])
        upper = float(b["upper_pct"])
        mask_bin = (
            (out["direction"] == dirn) &
            (out["abs_err"] >= lower) &
            (out["abs_err"] < upper)
        )
        out.loc[mask_bin & (out["band_label"] == ""), "band_label"] = b.get("label", "")

    # Cumulative slicing above tolerance band, basis from regulation mode when available
    for b in ordered:
        dirn = b["direction"]
        lower = float(b["lower_pct"])
        upper = float(b["upper_pct"])
        tol_end = tol_ui if dirn == "UI" else tol_oi
        # Compute effective lower bound clipped by tolerance end
        eff_lower = np.maximum(lower, tol_end)
        # Mask rows for this direction where deviation exceeds eff_lower
        mask_dir = (out["direction"] == dirn)
        if not mask_dir.any():
            continue
        D = out.loc[mask_dir, "abs_err"]
        slice_pct = np.clip(np.minimum(D, upper) - eff_lower, a_min=0.0, a_max=None)
        if not (slice_pct > 0).any():
            continue

        deviated_on = b.get("deviated_on", "AvC")
        rate_type = b["rate_type"]
        rate_value = float(b["rate_value"])
        slope = float(b.get("excess_slope_per_pct", 0.0))
        apply_to = b.get("apply_to", "penalty")
        label = b.get("label", "")

        # Prefer regulation-aligned basis if precomputed; fallback to per-band basis switch
        if "basis_MW" in out.columns:
            basis_series = out.loc[mask_dir, "basis_MW"]
        else:
            basis_series = out.loc[mask_dir, "Scheduled_MW"] if deviated_on == "Scheduled" else out.loc[mask_dir, "AvC_MW"]
        slice_kwh = (slice_pct / 100.0) * basis_series * 0.25 * 1000.0

        if rate_type == "flat_per_kwh":
            rate_series = rate_value
        elif rate_type == "ppa_fraction":
            rate_series = rate_value * out.loc[mask_dir, "PPA"]
        elif rate_type == "ppa_multiple":
            rate_series = rate_value * out.loc[mask_dir, "PPA"]
        elif rate_type == "flat_per_mwh":
            rate_series = rate_value / 1000.0
        elif rate_type == "scaled_excess":
            # Use actual D (absolute error %) as the scaler
            rate_series = rate_value + slope * D
        else:
            rate_series = 0.0

        amount_series = slice_kwh * rate_series

        if apply_to in ["penalty", "deviation_payable", "receivable", "drawl"]:
            out.loc[mask_dir, apply_to] = out.loc[mask_dir, apply_to] + amount_series

        # Accumulate total deviated energy and dev percent
        out.loc[mask_dir, "deviated_kWh"] = out.loc[mask_dir, "deviated_kWh"] + slice_kwh
        out.loc[mask_dir, "dev_pct"] = out.loc[mask_dir, "dev_pct"] + slice_pct

        # Store per-band slice energy
        safe = "slice_kWh_" + re.sub(r"[^0-9A-Za-z_]+", "_", label.replace(" ", "_"))
        out.loc[mask_dir, safe] = out.loc[mask_dir, safe] + slice_kwh

    # Revenue loss from OI above threshold
    # For OI where abs_err >= unpaid_oi_threshold, calculate unpaid energy * PPA
    oi_mask = (out["direction"] == "OI") & (out["abs_err"] >= unpaid_oi_threshold)
    # Calculate actual_kWh for these rows and multiply by PPA
    out["actual_kWh_pre"] = out["Actual_MW"] * 0.25 * 1000.0  # temporary for calculation
    out["rev_loss_oi_gt_thresh"] = 0.0
    out.loc[oi_mask, "rev_loss_oi_gt_thresh"] = out.loc[oi_mask, "actual_kWh_pre"] * pd.to_numeric(out.loc[oi_mask, "PPA"], errors="coerce").fillna(0)
    # Skipped rows should remain blank (not 0)
    out.loc[out.get("dsm_skipped_zero_inputs", False) == True, "rev_loss_oi_gt_thresh"] = np.nan  # noqa: E712
    # Drop temporary column
    out = out.drop(columns=["actual_kWh_pre"], errors="ignore")

    return out


# =========================
# Band-wise DSM (dashboard)
# =========================
def _band_rows_sorted(bands_rows: list[dict] | None) -> list[dict]:
    """Return normalized + sorted bands rows (UI then OI, by lower_pct)."""
    bands_df = _normalize_bands_df(pd.DataFrame(bands_rows or []))
    rows = bands_df.to_dict("records")

    def _key(r: dict):
        d = str(r.get("direction", "")).upper()
        d_key = 0 if d == "UI" else (1 if d == "OI" else 2)
        try:
            lo = float(r.get("lower_pct", 0) or 0.0)
        except Exception:
            lo = 0.0
        try:
            up = float(r.get("upper_pct", 0) or 0.0)
        except Exception:
            up = 0.0
        return (d_key, lo, up)

    return sorted(rows, key=_key)


def compute_bandwise_for_dashboard(
    detail_df: pd.DataFrame,
    bands_rows: list[dict] | None,
    err_mode: str | None = None,
    x_pct: float | None = None,
) -> tuple[pd.DataFrame, list[str]]:
    """Compute per-plant, per-band aggregation using the same band engine as exports.

    Returns:
      - DataFrame with columns: plant_name, band, direction, energy_kwh, revenue_loss
      - band_order: list of band labels in desired X-axis order
    """
    if detail_df is None or detail_df.empty:
        return pd.DataFrame(columns=["plant_name", "band", "direction", "energy_kwh", "revenue_loss"]), []

    df = detail_df.copy()

    # Plant column normalization
    if "plant_name" not in df.columns and "Plant" in df.columns:
        df["plant_name"] = df["Plant"]
    if "plant_name" not in df.columns:
        df["plant_name"] = "UNKNOWN"

    # Ensure numeric
    for c in ["AvC_MW", "Scheduled_MW", "Actual_MW", "PPA"]:
        if c not in df.columns:
            df[c] = 0.0
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0.0)

    # Ensure error_pct/abs_err/direction exist; prefer existing columns (already computed in pipelines)
    if "error_pct" not in df.columns:
        em = str(err_mode or "default").lower()
        xp = float(x_pct if x_pct is not None else 50.0)
        df["error_pct"] = compute_error_pct(df, em, xp)
    if "abs_err" not in df.columns:
        df["abs_err"] = pd.to_numeric(df["error_pct"], errors="coerce").fillna(0.0).abs()
    if "direction" not in df.columns:
        df["direction"] = np.where(pd.to_numeric(df["error_pct"], errors="coerce").fillna(0.0) < 0, "UI", "OI")

    # Ensure basis_MW exists; prefer existing (analysis/aggregation pipelines already compute it)
    if "basis_MW" not in df.columns:
        em = str(err_mode or "default").lower()
        xp = float(x_pct if x_pct is not None else 50.0)
        df["basis_MW"] = compute_basis_mw(df, em, xp)

    bands_sorted = _band_rows_sorted(bands_rows)
    if not bands_sorted:
        return pd.DataFrame(columns=["plant_name", "band", "direction", "energy_kwh", "revenue_loss"]), []

    # Use labels from the Bands table (single source of truth); fallback to generated label if missing.
    band_order: list[str] = []
    for r in bands_sorted:
        lbl = (r.get("label") or "").strip()
        if not lbl:
            try:
                lbl = generate_label(r)
            except Exception:
                d = str(r.get("direction", "")).upper()
                lo = float(r.get("lower_pct", 0) or 0.0)
                up = float(r.get("upper_pct", 0) or 0.0)
                lbl = f"{d} {lo}-{up}%"
        band_order.append(lbl)

    # Vectorize core inputs
    plant = df["plant_name"].astype(str)
    direction = df["direction"].astype(str).str.upper()
    abs_err = pd.to_numeric(df["abs_err"], errors="coerce").fillna(0.0).to_numpy()
    basis_mw = pd.to_numeric(df["basis_MW"], errors="coerce").fillna(0.0).to_numpy()
    ppa = pd.to_numeric(df["PPA"], errors="coerce").fillna(0.0).to_numpy()

    out_chunks: list[pd.DataFrame] = []
    for r, band_label in zip(bands_sorted, band_order):
        dirn = str(r.get("direction", "")).strip().upper()
        try:
            lower = float(r.get("lower_pct", 0) or 0.0)
        except Exception:
            lower = 0.0
        try:
            upper = float(r.get("upper_pct", 0) or 0.0)
        except Exception:
            upper = 0.0

        mask_dir = (direction.values == dirn)
        if not mask_dir.any():
            continue

        # Slice percent and convert to kWh for this band
        sp = np.maximum(0.0, np.minimum(abs_err, upper) - lower)
        sp = np.where(mask_dir, sp, 0.0)
        kwh = (sp / 100.0) * basis_mw * 0.25 * 1000.0

        # Rate rule mirrors apply_bands (rate_type strings come from UI table)
        rate_type = str(r.get("rate_type", "")).strip().lower()
        rate_value = float(r.get("rate_value", 0) or 0.0)
        slope = float(r.get("excess_slope_per_pct", 0) or 0.0)

        if rate_type == "flat_per_kwh":
            rate = rate_value
        elif rate_type in ("ppa_fraction", "ppa_multiple"):
            rate = rate_value * ppa
        elif rate_type == "flat_per_mwh":
            rate = rate_value / 1000.0
        elif rate_type == "scaled_excess":
            rate = rate_value + slope * abs_err
        else:
            rate = 0.0

        amount = kwh * rate

        tmp = pd.DataFrame({
            "plant_name": plant,
            "band": band_label,
            "direction": dirn,
            "energy_kwh": kwh,
            "revenue_loss": amount,
        })
        tmp = tmp[tmp["energy_kwh"] > 0]  # keep compact
        if tmp.empty:
            continue
        tmp = tmp.groupby(["plant_name", "band", "direction"], as_index=False)[["energy_kwh", "revenue_loss"]].sum()
        out_chunks.append(tmp)

    if not out_chunks:
        # return explicit zeros for all plants? caller can handle; keep empty for now
        return pd.DataFrame(columns=["plant_name", "band", "direction", "energy_kwh", "revenue_loss"]), band_order

    out_df = pd.concat(out_chunks, ignore_index=True)
    return out_df, band_order


def make_bandwise_bar_chart(df_band: pd.DataFrame, band_order: list[str], metric: str, plant_name: str):
    """Create a single bar chart across all bands; toggles Energy vs Revenue Loss."""
    metric_norm = str(metric or "ENERGY").upper()
    y_col = "energy_kwh" if metric_norm == "ENERGY" else "revenue_loss"
    y_title = "Energy Deviated (kWh)" if metric_norm == "ENERGY" else "Revenue Loss (₹)"

    if df_band is None or df_band.empty:
        fig = px.bar(pd.DataFrame({"band": [], y_col: []}), x="band", y=y_col)
        fig.update_layout(
            template="plotly_white",
            title=f"Band-wise DSM Analysis - {plant_name}",
            xaxis_title="Band",
            yaxis_title=y_title,
        )
        return fig

    fig = px.bar(
        df_band,
        x="band",
        y=y_col,
        color="direction",
        category_orders={"band": band_order} if band_order else None,
        labels={"band": "Band", y_col: y_title, "direction": "Direction"},
        hover_data={
            "band": True,
            "direction": True,
            "energy_kwh": ":,.0f",
            "revenue_loss": ":,.0f",
        },
    )
    fig.update_layout(
        template="plotly_white",
        title=f"Band-wise DSM Analysis - {plant_name}",
        xaxis_title="Bands",
        yaxis_title=y_title,
        bargap=0.25,
    )
    return fig


def bandwise_section_layout(prefix: str):
    """Reusable Band-wise DSM Analysis card."""
    return dbc.Card(
        [
            dbc.CardHeader("Band-wise DSM Analysis"),
            dbc.CardBody(
                [
                    dbc.Row(
                        [
                            dbc.Col(
                                [
                                    html.Label("Plant", style={"fontWeight": 600}),
                                    dcc.Dropdown(
                                        id=f"{prefix}-band-plant-dd",
                                        placeholder="Select Plant",
                                        clearable=False,
                                    ),
                                ],
                                md=4,
                            ),
                            dbc.Col(
                                [
                                    html.Label("Metric", style={"fontWeight": 600}),
                                    dcc.RadioItems(
                                        id=f"{prefix}-band-metric",
                                        options=[
                                            {"label": "Energy Deviated (kWh)", "value": "ENERGY"},
                                            {"label": "Revenue Loss (₹)", "value": "REVENUE"},
                                        ],
                                        value="ENERGY",
                                        inline=True,
                                    ),
                                ],
                                md=8,
                            ),
                        ],
                        className="mb-2",
                    ),
                    dcc.Graph(id=f"{prefix}-band-chart", config={"displayModeBar": True}),
                    dash_table.DataTable(
                        id=f"{prefix}-band-table",
                        columns=[
                            {"name": "Band", "id": "band"},
                            {"name": "Direction", "id": "direction"},
                            {"name": "Energy Deviated (kWh)", "id": "energy_kwh", "type": "numeric", "format": {"specifier": ",.0f"}},
                            {"name": "Revenue Loss (₹)", "id": "revenue_loss", "type": "numeric", "format": {"specifier": ",.0f"}},
                        ],
                        data=[],
                        sort_action="native",
                        style_table={"overflowX": "auto"},
                        style_cell={"padding": "6px", "textAlign": "left"},
                        style_header={"fontWeight": "600", "backgroundColor": "#f8f9fa"},
                        page_size=20,
                    ),
                ]
            ),
        ],
        className="mb-3",
    )

def summarize(df: pd.DataFrame, selected_plants: list[str] | None = None, bands_rows: List[Dict[str, Any]] | None = None, err_mode: str = "default", x_pct: float = 50.0, start_date: str = "", end_date: str = "") -> dict:
    # Ensure numeric and non-negative
    df["Scheduled_MW"] = pd.to_numeric(df["Scheduled_MW"], errors="coerce").fillna(0).clip(lower=0)
    df["Actual_MW"] = pd.to_numeric(df["Actual_MW"], errors="coerce").fillna(0).clip(lower=0)
    df["AvC_MW"] = pd.to_numeric(df["AvC_MW"], errors="coerce").fillna(0).clip(lower=0)
    df["PPA"] = pd.to_numeric(df["PPA"], errors="coerce").fillna(0).clip(lower=0)

    df["sched_kWh"] = df["Scheduled_MW"] * 0.25 * 1000
    df["actual_kWh"] = df["Actual_MW"] * 0.25 * 1000
    df["rev_sched"] = df["sched_kWh"] * df["PPA"]
    df["rev_actual"] = df["actual_kWh"] * df["PPA"]

    # Blockwise bins retained for UI compatibility
    df["bin"] = df.get("band_label", "").where(df.get("band_label", "") != "", "Unlabeled")
    if not df.empty:
        blockwise = (
            df.groupby("bin")
              .agg(**{
                  "No. of Blocks": ("bin", "size"),
                  "Deviated Energy (kWh)": ("deviated_kWh", "sum"),
                  "Penalty (₹)": ("penalty", "sum")
              })
              .reset_index()
        )
    else:
        blockwise = pd.DataFrame({"bin": [], "No. of Blocks": [], "Deviated Energy (kWh)": [], "Penalty (₹)": []})

    # New per-slot engine for summary metrics
    bands_list, _ = parse_bands_from_settings(bands_rows or [])
    mode_upper = MODE_DEFAULT if str(err_mode).lower() == "default" else MODE_DYNAMIC
    dyn_x = (float(x_pct) / 100.0) if mode_upper == MODE_DYNAMIC else 0.0

    # Build numeric detail for grouping
    detail_numeric_rows: List[Dict[str, Any]] = []
    for r in df.itertuples(index=False):
        slot = {
            "region": getattr(r, "region", None),
            "plant_name": getattr(r, "plant_name", getattr(r, "Plant", "")),
            "date": getattr(r, "date", None),
            "time_block": getattr(r, "time_block", getattr(r, "block", None)),
            "from_time": getattr(r, "from_time", None),
            "to_time": getattr(r, "to_time", None),
            "AvC_MW": float(getattr(r, "AvC_MW", 0.0) or 0.0),
            "Scheduled_MW": float(getattr(r, "Scheduled_MW", 0.0) or 0.0),
            "Actual_MW": float(getattr(r, "Actual_MW", 0.0) or 0.0),
            "PPA": float(getattr(r, "PPA", 0.0) or 0.0),
        }
        calc = compute_slot_row(slot, bands_list, mode_upper, dyn_x)
        calc.update({
            "Plant": getattr(r, "Plant", getattr(r, "plant_name", "")),
            "AvC_MW": slot["AvC_MW"],
            "PPA": slot["PPA"],
        })
        detail_numeric_rows.append(calc)
    detail_numeric_df = pd.DataFrame(detail_numeric_rows)

    # Normalize Plant column to avoid duplicates caused by whitespace/NBSP variations
    if "Plant" in df.columns:
        df["Plant"] = df["Plant"].apply(lambda x: _norm_plant_name(str(x)) if pd.notna(x) else x)
    if "Plant" in detail_numeric_df.columns:
        detail_numeric_df["Plant"] = detail_numeric_df["Plant"].apply(lambda x: _norm_plant_name(str(x)) if pd.notna(x) else x)

    # Per-plant summary aligned with new definitions
    if not detail_numeric_df.empty and "Plant" in df.columns:
        summaries = []
        for plant, idxs in df.groupby("Plant").groups.items():
            # Select same indices in detail_numeric_df
            dn = detail_numeric_df.iloc[list(idxs)] if len(detail_numeric_df) == len(df) else detail_numeric_df
            dn_p = dn  # assume aligned
            df_p = df.loc[list(idxs)]
            rev_loss_sum = float(dn_p.get("Revenue_Loss", pd.Series(dtype=float)).sum())
            rev_act_sum = float(dn_p.get("Revenue_as_per_generation", pd.Series(dtype=float)).sum())
            actual_kwh_sum = float((df_p["Actual_MW"] * 0.25 * 1000).sum())
            dsm_loss_sum = round(float(dn_p.get("Total_DSM", pd.Series(dtype=float)).sum()), 2)
            loss_pct = round((rev_loss_sum / rev_act_sum * 100.0) if rev_act_sum > 0 else 0.0, 2)
            paise_per_k = round((rev_loss_sum / actual_kwh_sum * 100.0) if actual_kwh_sum > 0 else 0.0, 2)
            
            # AvC median: only from blocks where Schedule > 0
            sch = pd.to_numeric(df_p.get("Scheduled_MW", 0), errors="coerce").fillna(0)
            active_avc = df_p.loc[sch > 0, "AvC_MW"] if "AvC_MW" in df_p.columns else pd.Series(dtype=float)
            cap_mode = round(float(active_avc.median()), 2) if len(active_avc) > 0 else 0.0
            
            # PPA median: only from blocks where Schedule > 0 AND PPA > 0
            ppa_series = pd.to_numeric(df_p.get("PPA", 0), errors="coerce").fillna(0)
            active_ppa = ppa_series[(sch > 0) & (ppa_series > 0)]
            ppa_mode = round(float(active_ppa.median()), 2) if len(active_ppa) > 0 else 0.0
            # Compute Data Availability % for this plant
            data_avail_pct = compute_plant_data_availability_pct(df_p)
            region_val = None
            try:
                region_val = str(df_p.get("region").iloc[0]) if "region" in df_p.columns and not df_p.empty else ""
            except Exception:
                region_val = ""
            
            # Calculate Run Date Range and Data Date Range
            def format_date_range(start_str, end_str):
                try:
                    start_dt = pd.to_datetime(start_str)
                    end_dt = pd.to_datetime(end_str)
                    return f"{start_dt.strftime('%d-%b-%Y')} to {end_dt.strftime('%d-%b-%Y')}"
                except Exception:
                    return f"{start_str} to {end_str}"
            
            run_date_range_str = format_date_range(start_date, end_date) if start_date and end_date else ""
            
            # Data Date Range: query FULL database for this plant (not filtered by user date range)
            first_date, last_date = get_plant_full_active_date_range(plant, region_val)
            if first_date == "No Active Data":
                data_date_range_str = "No Active Data"
            else:
                data_date_range_str = f"{first_date} to {last_date}"
            
            summaries.append({
                "Region": region_val,
                "Plant name": plant,
                "Plant Capacity": cap_mode,
                "PPA": ppa_mode,
                "Data Availability %": data_avail_pct,
                "Revenue Loss (%)": loss_pct,
                "DSM Loss": dsm_loss_sum,
                "Revenue Loss (p/k)": paise_per_k,
                "Run Date Range": run_date_range_str,
                "Data Date Range": data_date_range_str,
            })
        plant_summary = pd.DataFrame(summaries)
    else:
        plant_summary = pd.DataFrame(columns=["Region", "Plant name", "Plant Capacity", "PPA", "Data Availability %", "Revenue Loss (%)", "DSM Loss", "Revenue Loss (p/k)", "Run Date Range", "Data Date Range"])

    # Ensure selected plants are present and ordered; add zero rows for missing
    if selected_plants and len(selected_plants) > 0:
        desired_order = list(dict.fromkeys(selected_plants))
        existing = set(plant_summary["Plant name"]) if not plant_summary.empty else set()
        missing = [p for p in desired_order if p not in existing]
        if missing:
            zeros = pd.DataFrame({
                "Region": [str(df.get("region").iloc[0]) if "region" in df.columns and not df.empty else ""] * len(missing),
                "Plant name": missing,
                "Plant Capacity": [0.0] * len(missing),
                "PPA": [0.0] * len(missing),
                "Data Availability %": [0.0] * len(missing),
                "Revenue Loss (%)": [0.0] * len(missing),
                "DSM Loss": [0.0] * len(missing),
                "Revenue Loss (p/k)": [0.0] * len(missing),
            })
            plant_summary = pd.concat([plant_summary, zeros], ignore_index=True)
        cat = pd.Categorical(plant_summary["Plant name"], categories=desired_order, ordered=True)
        plant_summary = plant_summary.sort_values(by="Plant name", key=lambda s: cat).reset_index(drop=True)

    return {
        "df": df,
        "kpis": [],
        "blockwise": blockwise,
        "plant_summary": plant_summary,
    }

# ==========================================
# ---------------- LAYOUT ------------------
# ==========================================
def kpi_card(title: str, value: str) -> dbc.Card:
    return dbc.Card(
            dbc.CardBody([
            html.Div(title, className="text-muted", style={"fontSize": "0.9rem"}),
            html.H4(value, className="mt-1")
        ]),
        className="shadow-sm rounded-4"
    )

def sidebar():
    return html.Div([
        # Branding
        html.Div([
            html.Span("⚡", style={"fontSize": "1.5rem", "marginRight": "8px"}),
            html.Span(
                [
                    "Zelestra: DSM Analytics",
                    html.Span(
                        f"  •  {BUILD_TAG}",
                        style={"fontSize": "0.8rem", "fontWeight": "400", "opacity": "0.85", "marginLeft": "6px"},
                    ),
                ],
                style={"fontSize": "1.2rem", "fontWeight": "600", "color": "white"},
            ),
        ], style={"padding": "1.5rem 1rem", "borderBottom": "1px solid #444"}),
        
        # Navigation Menu
        html.Div([
            html.Div([
                html.Span("🏠", style={"marginRight": "12px"}),
                html.Span("Welcome", style={"color": "#ff6b35", "fontWeight": "500"})
            ], id="nav-welcome", className="nav-item active", style={
                "padding": "12px 20px", "cursor": "pointer", "borderLeft": "3px solid #ff6b35",
                "backgroundColor": "rgba(255, 107, 53, 0.1)", "marginBottom": "4px"
            }),
            html.Div([
                html.Span("⚙️", style={"marginRight": "12px"}),
                html.Span("Custom Settings", style={"color": "white", "fontWeight": "400"})
            ], id="nav-settings", className="nav-item", style={
                "padding": "12px 20px", "cursor": "pointer", "marginBottom": "4px"
            }),
            html.Div([
                html.Span("📊", style={"marginRight": "12px"}),
                html.Span("Analysis", style={"color": "white", "fontWeight": "400"})
            ], id="nav-analysis", className="nav-item", style={
                "padding": "12px 20px", "cursor": "pointer", "marginBottom": "4px"
            }),
            html.Div([
                html.Span("📊", style={"marginRight": "12px"}),
                html.Span("Aggregation Analysis", style={"color": "white", "fontWeight": "400"})
            ], id="nav-aggregation-analysis", className="nav-item", style={
                "padding": "12px 20px", "cursor": "pointer", "marginBottom": "4px"
            }),
            html.Div([
                html.Span("🔋", style={"marginRight": "12px"}),
                html.Span("Battery Investment", style={"color": "white", "fontWeight": "400"})
            ], id="nav-battery-investment", className="nav-item", style={
                "padding": "12px 20px", "cursor": "pointer", "marginBottom": "4px"
            }),
            html.Div([
                html.Span("📤", style={"marginRight": "12px"}),
                html.Span("Custom Upload", style={"color": "white", "fontWeight": "400"})
            ], id="nav-custom-upload", className="nav-item", style={
                "padding": "12px 20px", "cursor": "pointer", "marginBottom": "4px"
            }),
            html.Div([
                html.Span("📈", style={"marginRight": "12px"}),
                html.Span("Data Statistics", style={"color": "white", "fontWeight": "400"})
            ], id="nav-stats", className="nav-item", style={
                "padding": "12px 20px", "cursor": "pointer", "marginBottom": "4px"
            }),
        ], style={"padding": "1rem 0"}),
        
        # Logout
        html.Div([
            html.Div([
                html.Span("🌙", style={"marginRight": "12px"}),
                html.Span("Logout", style={"color": "#ccc"})
            ], style={
                "padding": "12px 20px", "cursor": "pointer", "borderTop": "1px solid #444",
                "marginTop": "auto"
            })
        ], style={"position": "absolute", "bottom": "0", "left": "0", "right": "0"})
    ], style={
        "backgroundColor": "#1a1a1a", "height": "100vh", "width": "280px", "position": "fixed", 
        "left": "0", "top": "0", "zIndex": "1000", "display": "flex", "flexDirection": "column"
    })

def main_content():
    return html.Div([
        # Header (dynamic per selected sidebar section)
        html.Div([
            html.H2(id="page-title", style={"margin": "0", "fontWeight": "600", "color": "#333"}),
            html.P(id="page-subtitle", style={"margin": "0", "color": "#666", "fontSize": "0.9rem"})
        ], style={"padding": "2rem 2rem 1rem", "borderBottom": "1px solid #eee"}),
        
        # Content
        html.Div([
            # Welcome Dashboard Content
            html.Div([
                html.Div([
                    html.Div([
                        html.H1("Welcome to DSM Analytics Dashboard", style={
                            "fontSize": "2.5rem", "fontWeight": "700", "color": "#1a1a1a", 
                            "marginBottom": "1rem", "textAlign": "center"
                        }),
                        html.P("Real-time Deviation Settlement Mechanism Analytics", style={
                            "fontSize": "1.2rem", "color": "#666", "textAlign": "center", "marginBottom": "3rem"
                        }),
                    ], style={"marginBottom": "3rem"}),
                    
                    dbc.Row([
                        dbc.Col([
                    dbc.Card([
                        dbc.CardBody([
                                    html.Div("⚙️", style={"fontSize": "3rem", "marginBottom": "1rem"}),
                                    html.H4("Step 1: Configure Settings", style={"color": "#ff6b35", "fontWeight": "600"}),
                                    html.P("Go to Custom Settings to configure your analysis parameters, bands, and rates.", 
                                          style={"color": "#666", "lineHeight": "1.8"}),
                                    html.Ul([
                                        html.Li("Set Error% calculation mode"),
                                        html.Li("Define deviation bands and penalties"),
                                        html.Li("Save your custom settings for reuse")
                                    ], style={"textAlign": "left", "color": "#666"}),
                                ])
                            ], className="shadow-sm rounded-4", style={"height": "100%"})
                        ], md=4),
                        
                            dbc.Col([
                                dbc.Card([
                                    dbc.CardBody([
                                    html.Div("📊", style={"fontSize": "3rem", "marginBottom": "1rem"}),
                                    html.H4("Step 2: Run Analysis", style={"color": "#28a745", "fontWeight": "600"}),
                                    html.P("Navigate to Analysis to select region, plants, and date range for your analysis.", 
                                          style={"color": "#666", "lineHeight": "1.8"}),
                                    html.Ul([
                                        html.Li("Select region(s) from DuckDB"),
                                        html.Li("Choose specific plants or select all"),
                                        html.Li("Pick date range and plot results")
                                    ], style={"textAlign": "left", "color": "#666"}),
                                ])
                            ], className="shadow-sm rounded-4", style={"height": "100%"})
                            ], md=4),
                        
                            dbc.Col([
                                dbc.Card([
                                    dbc.CardBody([
                                    html.Div("📈", style={"fontSize": "3rem", "marginBottom": "1rem"}),
                                    html.H4("Step 3: View Results", style={"color": "#17a2b8", "fontWeight": "600"}),
                                    html.P("Analyze comprehensive DSM calculations, KPIs, and blockwise breakdown.", 
                                          style={"color": "#666", "lineHeight": "1.8"}),
                                    html.Ul([
                                        html.Li("View detailed KPIs and penalties"),
                                        html.Li("Explore blockwise analysis"),
                                        html.Li("Download full calculation in Excel")
                                    ], style={"textAlign": "left", "color": "#666"}),
                                ])
                            ], className="shadow-sm rounded-4", style={"height": "100%"})
                            ], md=4),
                    ], className="mb-4"),
                    
                    html.Div([
                                dbc.Card([
                                    dbc.CardBody([
                                html.H5("🚀 Quick Start", style={"color": "#333", "marginBottom": "1rem"}),
                                html.P([
                                    "1. Click on ",
                                    html.Strong("Custom Settings", style={"color": "#ff6b35"}),
                                    " to configure your parameters and save settings"
                                ], style={"marginBottom": "0.5rem"}),
                                html.P([
                                    "2. Navigate to ",
                                    html.Strong("Analysis", style={"color": "#28a745"}),
                                    " to select data and run calculations"
                                ], style={"marginBottom": "0.5rem"}),
                                html.P([
                                    "3. Review results and download Excel report when needed"
                                ], style={"marginBottom": "0"}),
                            ])
                        ], className="shadow-sm rounded-4", style={"backgroundColor": "#f8f9fa"})
                    ], style={"marginTop": "3rem"}),
                    
                    html.Div([
                        dbc.Button([
                            html.Span("Get Started ", style={"marginRight": "8px"}),
                            html.Span("→")
                        ], id="btn-get-started", color="primary", size="lg", 
                                 style={"backgroundColor": "#ff6b35", "border": "none", "fontWeight": "600", "padding": "12px 40px"})
                    ], style={"textAlign": "center", "marginTop": "3rem"}),
                    
                ], style={"maxWidth": "1200px", "margin": "0 auto"})
            ], id="welcome-content", style={"display": "block"}),
            
            # Custom Settings Content
            html.Div([

                # ══════════════════════════════════════════
                # PRESET MANAGER  — top-of-page control bar
                # ══════════════════════════════════════════
                html.Div([
                    # Header strip
                    html.Div([
                        dbc.Row([
                            dbc.Col(
                                html.Div([
                                    html.Span("Saved Presets", style={
                                        "fontWeight": "700", "fontSize": "1rem", "color": "#1a1a2e"
                                    }),
                                    html.Span(" — select to view, load or delete", style={
                                        "fontSize": "0.82rem", "color": "#888", "marginLeft": "6px"
                                    }),
                                ]),
                                md=8,
                            ),
                            dbc.Col(
                                html.Div(
                                    id="preset-mode-badge",
                                    style={"textAlign": "right"},
                                ),
                                md=4,
                            ),
                        ], align="center"),
                    ], style={
                        "backgroundColor": "#f0f4ff",
                        "borderBottom": "1px solid #d0d8f0",
                        "padding": "12px 20px",
                        "borderRadius": "12px 12px 0 0",
                    }),

                    # Body
                    html.Div([
                        # Row: Dropdown + Buttons
                        dbc.Row([
                            dbc.Col([
                                dcc.Dropdown(
                                    id="preset-manage-select",
                                    placeholder="\u2014  Choose a saved preset to inspect  \u2014",
                                    clearable=True,
                                    style={"fontSize": "0.92rem"},
                                ),
                            ], md=7),
                            dbc.Col([
                                html.Div([
                                    dbc.Button(
                                        "\u271a  Create New",
                                        id="btn-preset-new",
                                        color="secondary",
                                        outline=True,
                                        size="sm",
                                        className="me-2",
                                        style={"fontWeight": "600"},
                                    ),
                                    dbc.Button(
                                        "\u2b07\ufe0f  Load & Edit",
                                        id="btn-load-preset",
                                        color="primary",
                                        size="sm",
                                        disabled=True,
                                        className="me-2",
                                        style={"fontWeight": "600"},
                                    ),
                                    dbc.Button(
                                        "\U0001f5d1\ufe0f  Delete",
                                        id="btn-delete-preset-manage",
                                        color="danger",
                                        outline=True,
                                        size="sm",
                                        disabled=True,
                                        style={"fontWeight": "600"},
                                    ),
                                ], style={"display": "flex", "gap": "6px", "alignItems": "center"}),
                            ], md=5),
                        ]),

                        # Preset detail (shown after selection)
                        html.Div(id="preset-detail-view"),
                        html.Div(id="preset-manage-message", style={"marginTop": "8px"}),
                    ], style={"padding": "18px 20px"}),
                ], style={
                    "borderRadius": "12px",
                    "border": "2px solid #d0d8f0",
                    "backgroundColor": "#ffffff",
                    "boxShadow": "0 4px 20px rgba(60,80,180,0.10)",
                    "marginBottom": "24px",
                }),

                # Global Regulation Controls Section (Zero Basis Guard only)
                html.Div([
                    html.H5("Global Regulation Controls", style={"marginBottom": "1rem", "color": "#555"}),
                    dbc.Row([
                        dbc.Col([
                            html.Label("Zero Basis Guard", style={"fontWeight": "500", "marginBottom": "8px", "fontSize": "0.95rem", "color": "#555"}),
                            dbc.Checklist(
                                id="zero-basis-guard",
                                options=[{"label": "Skip night blocks (AvC≈0 & Schedule≈0)", "value": "on"}],
                                value=["on"],
                                inline=False,
                                style={"fontSize": "0.9rem"}
                            ),
                        ], md=12),
                    ], className="mb-3"),
                ], style={"marginBottom": "2rem", "padding": "1.5rem", "backgroundColor": "#f8f9fa", "borderRadius": "8px"}),
                
                # Error% Mode Section
                html.Div([
                    html.H5("Error% Mode", style={"marginBottom": "1rem", "color": "#555"}),
            dcc.RadioItems(
                        id="err-mode",
                options=[
                            {"label": "Default: (Actual - Scheduled) / AvC × 100", "value": "default"},
                            {"label": "Dynamic: 100×(Actual - Scheduled) / (X%·AvC + (100−X)%·Scheduled)", "value": "dynamic"},
                ],
                value="default",
                        inputStyle={"marginRight": "8px", "accentColor": "#ff6b35"},
                        labelStyle={"display": "block", "marginBottom": "12px", "color": "#666", "fontSize": "1rem"},
                    ),
                    html.Div([
                        html.Label(id="x-pct-label", children="X% (for Dynamic Error%)", 
                                  style={"color": "#555", "fontSize": "1rem", "marginBottom": "8px", "fontWeight": "500"}),
                        dcc.Slider(
                            id="x-pct", 
                        min=0,
                        max=100,
                            value=50, 
                        step=1,
                            marks={i: str(i) for i in range(0, 101, 10)},
                            tooltip={"placement": "bottom", "always_visible": True}
                        ),
                        html.Div(id="x-pct-readout", style={
                            "textAlign": "center", "color": "#ff6b35", "fontWeight": "bold", 
                            "fontSize": "1.2rem", "marginTop": "8px"
                        }),
                    ], id="x-pct-container", style={"display": "none", "marginTop": "1rem"}),
                ], style={"marginBottom": "2rem", "padding": "1.5rem", "backgroundColor": "#f8f9fa", "borderRadius": "8px"}),
                
                # User-defined Bands & Rates Section
                html.Div([
                    html.H5("User-defined Bands & Rates", style={"marginBottom": "1rem", "color": "#555"}),
                    
                    
                    # Add New Row Form
                    dbc.Card([
                        dbc.CardHeader([
                            html.H6("Add New Band", style={"margin": "0", "color": "#555", "fontWeight": "600"})
                        ]),
                        dbc.CardBody([
                            dbc.Row([
                                # Direction
                                dbc.Col([
                                    html.Label("Direction", style={"fontWeight": "500", "marginBottom": "4px", "fontSize": "0.9rem"}),
            dcc.Dropdown(
                                        id="form-direction",
                                        options=BANDS_DROPDOWNS["direction"]["options"],
                                        value="UI",
                clearable=False,
                                        style={"fontSize": "0.9rem"}
                                    ),
                                ], md=2),
                                
                                # Lower %
                                dbc.Col([
                                    html.Label("Lower %", style={"fontWeight": "500", "marginBottom": "4px", "fontSize": "0.9rem"}),
                                    dbc.Input(
                                        id="form-lower-pct",
                                        type="number",
                                        value=0.0,
                                        step=0.1,
                                        min=0,
                                        style={"fontSize": "0.9rem"}
                                    ),
                                ], md=1),
                                
                                # Upper %
                                dbc.Col([
                                    html.Label("Upper %", style={"fontWeight": "500", "marginBottom": "4px", "fontSize": "0.9rem"}),
                                    dbc.Input(
                                        id="form-upper-pct",
                                        type="number",
                                        value=10.0,
                                        step=0.1,
                                        min=0,
                                        style={"fontSize": "0.9rem"}
                                    ),
                                ], md=1),
                # Tolerance Cut % (hidden)
                dbc.Col([
                    html.Label("Tolerance Cut %", style={"fontWeight": "500", "marginBottom": "4px", "fontSize": "0.9rem"}),
                    dbc.Input(
                        id="form-tolerance-cut-pct",
                        type="number",
                        value=10.0,
                        step=0.1,
                        min=0,
                        style={"fontSize": "0.9rem"}
                    ),
                ], md=1, style={"display": "none"}),
                
                # Loss Zone
                dbc.Col([
                    html.Label("Loss Zone", style={"fontWeight": "500", "marginBottom": "4px", "fontSize": "0.9rem"}),
                    dcc.Dropdown(
                        id="form-loss-zone",
                        options=BANDS_DROPDOWNS["loss_zone"]["options"],
                        value=False,
                        clearable=False,
                        style={"fontSize": "0.9rem"}
                    ),
                ], md=2),
                            ], className="mb-2"),
                            
                            dbc.Row([
                                # Rate Type
                                dbc.Col([
                                    html.Label("Rate Type", style={"fontWeight": "500", "marginBottom": "4px", "fontSize": "0.9rem"}),
                                    dcc.Dropdown(
                                        id="form-rate-type",
                                        options=BANDS_DROPDOWNS["rate_type"]["options"],
                                        value="flat_per_kwh",
                                        clearable=False,
                                        style={"fontSize": "0.9rem"}
                                    ),
                                ], md=2),
                                
                                # Rate Value
                                dbc.Col([
                                    html.Label("Rate Value", style={"fontWeight": "500", "marginBottom": "4px", "fontSize": "0.9rem"}),
                                    dbc.Input(
                                        id="form-rate-value",
                                        type="number",
                                        value=0.0,
                                        step=0.01,
                                        min=0,
                                        style={"fontSize": "0.9rem"}
                                    ),
                                ], md=1),
                                
                                # Excess Slope %
                                dbc.Col([
                                    html.Label("Excess Slope %", style={"fontWeight": "500", "marginBottom": "4px", "fontSize": "0.9rem"}),
                                    dbc.Input(
                                        id="form-excess-slope-per-pct",
                                        type="number",
                                        value=0.0,
                                        step=0.01,
                                        min=0,
                                        style={"fontSize": "0.9rem"}
                                    ),
                                ], md=1),
                                
                                # Apply To removed (routing is via Loss Zone + direction)
                                
                                # Add Button
                                dbc.Col([
                                    html.Label("&nbsp;", style={"marginBottom": "4px"}),  # Spacer
        dbc.Button(
                                        "Add to Table", 
                                        id="add-from-form", 
                                        color="success", 
                                        size="sm", 
                                        className="w-100",
                                        style={"fontSize": "0.9rem"}
                                    ),
                                ], md=2),
                            ], className="mb-2"),
                            
                            # Auto-generated Label Preview
                dbc.Row([
                                dbc.Col([
                                    html.Label("Preview Label:", style={"fontWeight": "500", "marginBottom": "4px", "fontSize": "0.9rem"}),
                                    html.Div(
                                        id="form-label-preview",
                                        style={
                                            "padding": "8px", 
                                            "backgroundColor": "#f8f9fa", 
                                            "border": "1px solid #dee2e6", 
                                            "borderRadius": "4px",
                                            "fontSize": "0.9rem",
                                            "color": "#495057"
                                        }
                                    ),
                                ], md=12),
                            ]),
                        ]),
                    ], className="mb-3"),
                    
        dash_table.DataTable(
                        id="bands-table",
                        columns=BANDS_COLUMNS,
                        data=DEFAULT_BANDS.copy(),
                        editable=True,
                        row_deletable=True,
                        sort_action="native",
                        page_size=8,
                        style_table={"overflow": "visible", "overflowX": "auto", "overflowY": "visible"},
                        style_cell={"padding": "12px", "fontSize": "0.9rem"},
                        style_header={"backgroundColor": "#f8f9fa", "fontWeight": "600"},
                        dropdown={
                            "direction": {**BANDS_DROPDOWNS["direction"], "clearable": False},
                            "rate_type": {**BANDS_DROPDOWNS["rate_type"], "clearable": False},
                            "loss_zone": {**BANDS_DROPDOWNS["loss_zone"], "clearable": False},
                        },
                    ),
                    html.Div([
                        dbc.Button("Reset Table", id="reset-bands", size="sm", className="mt-3 me-2",
                                 style={"backgroundColor": "#6c757d", "border": "none"}),
                        dcc.Upload(
                            id="upload-bands",
                            children=dbc.Button("Load from File", size="sm", className="mt-3 me-2",
                                             style={"backgroundColor": "#17a2b8", "border": "none"}),
                            multiple=False
                        ),
                        dbc.Button("Save to File", id="save-bands", size="sm", className="mt-3",
                                 style={"backgroundColor": "#28a745", "border": "none"}),
                        dcc.Download(id="download-bands-json"),
                    ]),
                ], style={"padding": "1.5rem", "backgroundColor": "#f8f9fa", "borderRadius": "8px", "marginBottom": "2rem"}),

                # ══════════════════════════════════════════
                # SAVE / UPDATE  — bottom action panel
                # ══════════════════════════════════════════
                html.Div([
                    # Header strip
                    html.Div([
                        dbc.Row([
                            dbc.Col(
                                html.Span("\U0001f4be  Save / Update Preset", style={
                                    "fontWeight": "700", "fontSize": "1rem", "color": "#1a1a2e"
                                }),
                                md=7,
                            ),
                            dbc.Col(
                                html.Div(id="save-mode-hint", style={
                                    "textAlign": "right", "fontSize": "0.82rem", "color": "#666"
                                }),
                                md=5,
                            ),
                        ], align="center"),
                    ], style={
                        "backgroundColor": "#f0fff4",
                        "borderBottom": "1px solid #b2dfc4",
                        "padding": "12px 20px",
                        "borderRadius": "12px 12px 0 0",
                    }),

                    # Body
                    html.Div([
                        dbc.Row([
                            dbc.Col([
                                html.Label("Preset Name", style={
                                    "fontSize": "0.78rem", "fontWeight": "700", "color": "#555",
                                    "textTransform": "uppercase", "letterSpacing": "0.4px", "marginBottom": "6px"
                                }),
                                dbc.Input(
                                    id="preset-name",
                                    placeholder="e.g., Solar A \u2013 Default, Wind B \u2013 Dynamic X=50",
                                    type="text",
                                    style={"fontSize": "0.92rem"},
                                ),
                            ], md=6),
                            dbc.Col([
                                html.Label("\u00a0", style={"display": "block", "marginBottom": "6px"}),
                                dbc.Button(
                                    "\U0001f4be  Save as New Preset",
                                    id="btn-save-preset",
                                    color="success",
                                    className="w-100",
                                    style={"fontWeight": "600"},
                                ),
                            ], md=3),
                            dbc.Col([
                                html.Label("\u00a0", style={"display": "block", "marginBottom": "6px"}),
                                dbc.Button(
                                    "\U0001f504  Update Selected",
                                    id="btn-save-settings",
                                    color="primary",
                                    outline=True,
                                    className="w-100",
                                    style={"fontWeight": "600"},
                                ),
                            ], md=3),
                        ], className="mb-2"),
                        html.Div(id="preset-save-message", style={"marginTop": "6px"}),
                        html.Div(id="settings-save-message", style={"marginTop": "4px"}),
                        # Hidden button retained for legacy delete callback compatibility
                        html.Div(
                            dbc.Button("", id="btn-delete-preset", style={"display": "none"}),
                            style={"display": "none"},
                        ),
                    ], style={"padding": "18px 20px"}),
                ], style={
                    "borderRadius": "12px",
                    "border": "2px solid #b2dfc4",
                    "backgroundColor": "#ffffff",
                    "boxShadow": "0 4px 20px rgba(30,150,80,0.08)",
                    "marginBottom": "12px",
                }),

            ], id="custom-settings-content", style={"display": "none"}),
            
            # Analysis Content
            html.Div([
                # Analysis intro (header handled globally)
                html.Div([]),
                
                # Global Filters Section
                html.Div([
                    html.Div([
                        html.Span("🔍", style={"fontSize": "1.5rem", "marginRight": "10px"}),
                        html.Span("Data Selection", style={"fontSize": "1.1rem", "fontWeight": "600", "color": "#333"})
                    ], style={"marginBottom": "1.5rem"}),
                    
                        dbc.Row([
                            dbc.Col([
                                html.Label([
                                html.Span("🌐 ", style={"marginRight": "6px"}),
                                "Region"
                            ], style={"fontWeight": "600", "marginBottom": "10px", "color": "#333", "fontSize": "0.95rem"}),
                            dcc.Loading(
                                id="region-loading",
                                type="circle",
                                children=dmc.MultiSelect(
                                    id="region-dd",
                                    value=[],
                                    data=[],
                                    searchable=True,
                                    clearable=True,
                                    placeholder="Select Region(s)...",
                                    comboboxProps={"keepOpened": True},
                                    maxDropdownHeight=300,
                                ),
                                ),
                            ], md=4),
                            dbc.Col([
                                html.Label([
                                html.Span("🏭 ", style={"marginRight": "6px"}),
                                "State"
                            ], style={"fontWeight": "600", "marginBottom": "10px", "color": "#333", "fontSize": "0.95rem"}),
                                dmc.MultiSelect(
                                    id="state-dd",
                                    value=[],
                                    data=[],
                                    searchable=True,
                                    clearable=True,
                                    placeholder="Select State(s)...",
                                    comboboxProps={"keepOpened": True},
                                    maxDropdownHeight=300,
                                ),
                            ], md=2),
                            dbc.Col([
                                html.Label([
                                html.Span("🏭 ", style={"marginRight": "6px"}),
                                "Resource"
                            ], style={"fontWeight": "600", "marginBottom": "10px", "color": "#333", "fontSize": "0.95rem"}),
                                dmc.MultiSelect(
                                    id="resource-type-dd",
                                        value=[],
                                    data=[],
                                        searchable=True,
                                    clearable=True,
                                    placeholder="Select Resource(s)...",
                                    comboboxProps={"keepOpened": True},
                                    maxDropdownHeight=300,
                                ),
                            ], md=2),
                            dbc.Col([
                                html.Label([
                                html.Span("📅 ", style={"marginRight": "6px"}),
                                "Date Range"
                            ], style={"fontWeight": "600", "marginBottom": "10px", "color": "#333", "fontSize": "0.95rem"}),
                                dcc.DatePickerRange(
                                    id="date-range",
                                    start_date=(datetime.now() - timedelta(days=6)).date(),
                                    end_date=datetime.now().date(),
                                    display_format="DD-MMM-YYYY",
                                    minimum_nights=0,
                                style={"fontSize": "0.95rem"}
                                ),
                            ], md=4),
                        ], className="mb-4"),

                        dbc.Row([
                            dbc.Col([
                                html.Label([
                                    html.Span("🏢 ", style={"marginRight": "6px"}),
                                    "QCA (Agency)"
                                ], style={"fontWeight": "600", "marginBottom": "10px", "color": "#333", "fontSize": "0.95rem"}),
                                dmc.MultiSelect(
                                    id="qca-dd",
                                    value=[],
                                    data=[],
                                    searchable=True,
                                    clearable=True,
                                    placeholder="Select QCA(s)...",
                                    comboboxProps={"keepOpened": True},
                                    maxDropdownHeight=300,
                                ),
                            ], md=4),
                            dbc.Col([
                                html.Label([
                                    html.Span("🔌 ", style={"marginRight": "6px"}),
                                    "Pooling Station"
                                ], style={"fontWeight": "600", "marginBottom": "10px", "color": "#333", "fontSize": "0.95rem"}),
                                dmc.MultiSelect(
                                    id="pooling-dd",
                                    value=[],
                                    data=[],
                                    searchable=True,
                                    clearable=True,
                                    placeholder="Select Pooling Station(s)...",
                                    comboboxProps={"keepOpened": True},
                                    maxDropdownHeight=300,
                                ),
                            ], md=4),
                            dbc.Col([
                                html.Label([
                                    html.Span("🏭 ", style={"marginRight": "6px"}),
                                    "Plant"
                                ], style={"fontWeight": "600", "marginBottom": "10px", "color": "#333", "fontSize": "0.95rem"}),
                                dcc.Loading(
                                    id="plant-loading",
                                    type="circle",
                                    children=dmc.MultiSelect(
                                        id="plant-dd",
                                        value=[],
                                        data=[],
                                        searchable=True,
                                        clearable=True,
                                        placeholder="Select Plant(s) or Select All...",
                                        comboboxProps={"keepOpened": True},
                                        maxDropdownHeight=400,
                                    ),
                                ),
                            ], md=4),
                        ], className="mb-4"),
                        # Exclude Plants Section (appears only when SELECT_ALL is active)
                        html.Div([
                            dbc.Row([
                                dbc.Col([
                                    html.Div([
                                        html.Label([
                                            html.Span("❌ ", style={"marginRight": "6px", "color": "#dc3545"}),
                                            "Exclude Plants from Analysis (Optional)",
                                            html.Small(" - Only available when 'Select All' is active", 
                                                      style={"marginLeft": "8px", "color": "#666", "fontWeight": "normal", "fontSize": "0.85rem"})
                                        ], style={"fontWeight": "600", "marginBottom": "10px", "color": "#333", "fontSize": "0.95rem"}),
                                        dcc.Loading(
                                            id="exclude-plant-loading",
                                            type="circle",
                                            children=dmc.MultiSelect(
                                                id="exclude-plant-dd",
                                                value=[],
                                                data=[],
                                                searchable=True,
                                                clearable=True,
                                                disabled=True,
                                                placeholder="Select plants to exclude from analysis...",
                                                comboboxProps={"keepOpened": True},
                                                maxDropdownHeight=300,
                                            ),
                                        ),
                                    ], style={
                                        "padding": "1rem",
                                        "backgroundColor": "#fff3cd",
                                        "border": "1px solid #ffc107",
                                        "borderRadius": "8px",
                                        "marginBottom": "1rem"
                                    })
                                ], md=12),
                            ], className="mb-3"),
                        ], id="exclude-section", style={"display": "none"}),
                        # Application Mode (GLOBAL vs PLANT_WISE)
                        dbc.Row([
                            dbc.Col([
                                html.Label("Application Mode", style={"fontWeight": 600, "marginBottom": "10px", "color": "#333", "fontSize": "0.95rem"}),
                                dcc.RadioItems(
                                    id="custom-setting-mode",
                                    options=[
                                        {"label": "Apply SAME custom setting to all plants", "value": "GLOBAL"},
                                        {"label": "Apply DIFFERENT custom settings per plant", "value": "PLANT_WISE"},
                                    ],
                                    value="GLOBAL",
                                    inline=True,
                                    style={"fontSize": "0.9rem"},
                                )
                            ], md=12),
                        ], className="mb-2"),
                        # Preset selection (optional) - GLOBAL mode only
                        dbc.Row([
                            dbc.Col([
                                html.Label("Custom Setting(s)", style={"fontWeight": 600, "marginBottom": "10px", "color": "#333", "fontSize": "0.95rem"}),
                                dmc.MultiSelect(
                                    id="analysis-preset-select",
                                    data=[],
                                    placeholder="(Optional) Choose one or more saved presets",
                                    value=[],
                                    searchable=True,
                                    clearable=True,
                                    comboboxProps={"keepOpened": True},
                                    maxDropdownHeight=300,
                                )
                            ], md=4),
                        ], className="mb-2", id="global-setting-row"),
                        # Plant-wise mapping table (PLANT_WISE mode only)
                        dbc.Row([
                            dbc.Col([
                                # Store for settings→plants assignment rows (PLANT_WISE mode)
                                dcc.Store(
                                    id="setting-assignments-store",
                                    data={
                                        "next_id": 1,
                                        "rows": [
                                            {"row": 0, "preset": None, "plants": [], "apply_remaining": True}
                                        ],
                                    },
                                ),
                                html.Div(
                                    id="plant-setting-mapping-container",
                                    style={"display": "none"}
                                )
                            ], md=12),
                        ], className="mb-2"),
                        
                        dbc.Row([
                            dbc.Col([
                                dbc.Button([
                                    html.Span("📊 ", style={"marginRight": "8px"}),
                                "Run Analysis"
                                ], id="plot-now", size="lg", className="w-100", disabled=True,
                                     style={"backgroundColor": "#ff6b35", "border": "none", "fontWeight": "600", "padding": "14px",
                                       "boxShadow": "0 4px 6px rgba(255, 107, 53, 0.3)"}),
                            ], md=6, className="offset-md-3"),
                        ]),
                    # Horizontal progress bar while computing
                        dbc.Row([
                            dbc.Col([
                            html.Div(id="progress-container", style={"display": "none", "marginTop": "1.5rem"}, children=[
                                html.Div([
                                    html.Span("⏳ ", style={"marginRight": "8px"}),
                                    html.Span("Processing data...", style={"fontWeight": "500", "color": "#333"})
                                ], style={"marginBottom": "10px", "textAlign": "center"}),
                                dbc.Progress(
                                    id="compute-progress",
                                    value=100,
                                    striped=True,
                                    animated=True,
                                    style={"height": "25px"},
                                    className="mb-2"
                                ),
                            ]),
                        ], md=8, className="offset-md-2")
                    ]),
                ], style={"padding": "2rem", "backgroundColor": "#fff", "borderRadius": "12px", "marginBottom": "2rem",
                         "boxShadow": "0 2px 8px rgba(0,0,0,0.05)"}),
                
                # Results Section
                html.Div(id="results-section", style={"display": "none"}, children=[
                    html.Div(id="tab-content", className="mt-3"),
                ]),
            ], id="analysis-content", style={"display": "none"}),
            # Aggregation Analysis Content
            html.Div([
                # Aggregation Analysis intro (header handled globally)
                html.Div([]),
                # Global Filters Section (mirrors Analysis)
                html.Div([
                    html.Div([
                        html.Span("🔍", style={"fontSize": "1.5rem", "marginRight": "10px"}),
                        html.Span("Data Selection", style={"fontSize": "1.1rem", "fontWeight": "600", "color": "#333"})
                    ], style={"marginBottom": "1.5rem"}),
                    dbc.Row([
                        dbc.Col([
                            html.Label([
                                html.Span("🌐 ", style={"marginRight": "6px"}),
                                "Region"
                            ], style={"fontWeight": "600", "marginBottom": "10px", "color": "#333", "fontSize": "0.95rem"}),
                            dcc.Loading(
                                id="agg-region-loading",
                                type="circle",
                                children=dmc.MultiSelect(
                                    id="agg-region-dd",
                                    value=[],
                                    data=[],
                                    searchable=True,
                                    clearable=True,
                                    placeholder="Select Region(s)...",
                                    comboboxProps={"keepOpened": True},
                                    maxDropdownHeight=300,
                                ),
                            ),
                        ], md=4),
                        dbc.Col([
                            html.Label([
                                html.Span("🏭 ", style={"marginRight": "6px"}),
                                "State"
                            ], style={"fontWeight": "600", "marginBottom": "10px", "color": "#333", "fontSize": "0.95rem"}),
                            dmc.MultiSelect(
                                id="agg-state-dd",
                                data=[],
                                value=[],
                                searchable=True,
                                clearable=True,
                                placeholder="Select State(s)...",
                                comboboxProps={"keepOpened": True},
                                maxDropdownHeight=300,
                            ),
                        ], md=2),
                        dbc.Col([
                            html.Label([
                                html.Span("🏭 ", style={"marginRight": "6px"}),
                                "Resource"
                            ], style={"fontWeight": "600", "marginBottom": "10px", "color": "#333", "fontSize": "0.95rem"}),
                            dmc.MultiSelect(
                                id="agg-resource-type-dd",
                                data=[],
                                    value=[],
                                    searchable=True,
                                clearable=True,
                                placeholder="Select Resource(s)...",
                                comboboxProps={"keepOpened": True},
                                maxDropdownHeight=300,
                            ),
                        ], md=2),
                        dbc.Col([
                            html.Label([
                                html.Span("📅 ", style={"marginRight": "6px"}),
                                "Date Range"
                            ], style={"fontWeight": "600", "marginBottom": "10px", "color": "#333", "fontSize": "0.95rem"}),
                            dcc.DatePickerRange(
                                id="agg-date-range",
                                start_date=(datetime.now() - timedelta(days=6)).date(),
                                end_date=datetime.now().date(),
                                display_format="DD-MMM-YYYY",
                                minimum_nights=0,
                                style={"fontSize": "0.95rem"}
                            ),
                        ], md=4),
                    ], className="mb-4"),

                    dbc.Row([
                        dbc.Col([
                            html.Label([
                                html.Span("🏢 ", style={"marginRight": "6px"}),
                                "QCA (Agency)"
                            ], style={"fontWeight": "600", "marginBottom": "10px", "color": "#333", "fontSize": "0.95rem"}),
                            dmc.MultiSelect(
                                id="agg-qca-dd",
                                data=[],
                                value=[],
                                searchable=True,
                                clearable=True,
                                placeholder="Select QCA(s)...",
                                comboboxProps={"keepOpened": True},
                                maxDropdownHeight=300,
                            ),
                        ], md=4),
                        dbc.Col([
                            html.Label([
                                html.Span("🔌 ", style={"marginRight": "6px"}),
                                "Pooling Station"
                            ], style={"fontWeight": "600", "marginBottom": "10px", "color": "#333", "fontSize": "0.95rem"}),
                            dmc.MultiSelect(
                                id="agg-pooling-dd",
                                data=[],
                                value=[],
                                searchable=True,
                                clearable=True,
                                placeholder="Select Pooling Station(s)...",
                                comboboxProps={"keepOpened": True},
                                maxDropdownHeight=300,
                            ),
                        ], md=4),
                        dbc.Col([
                            html.Label([
                                html.Span("🏭 ", style={"marginRight": "6px"}),
                                "Plant"
                            ], style={"fontWeight": "600", "marginBottom": "10px", "color": "#333", "fontSize": "0.95rem"}),
                            dcc.Loading(
                                id="agg-plant-loading",
                                type="circle",
                                children=dmc.MultiSelect(
                                    id="agg-plant-dd",
                                    data=[],
                                    value=[],
                                    searchable=True,
                                    clearable=True,
                                    placeholder="Select Plant(s) or Select All...",
                                    comboboxProps={"keepOpened": True},
                                    maxDropdownHeight=400,
                                ),
                            ),
                        ], md=4),
                    ], className="mb-4"),
                    # Exclude Plants Section (appears only when SELECT_ALL is active)
                    html.Div([
                        dbc.Row([
                            dbc.Col([
                                html.Div([
                                    html.Label([
                                        html.Span("❌ ", style={"marginRight": "6px", "color": "#dc3545"}),
                                        "Exclude Plants from Aggregation (Optional)",
                                        html.Small(" - Only available when 'Select All' is active", 
                                                  style={"marginLeft": "8px", "color": "#666", "fontWeight": "normal", "fontSize": "0.85rem"})
                                    ], style={"fontWeight": "600", "marginBottom": "10px", "color": "#333", "fontSize": "0.95rem"}),
                                    dcc.Loading(
                                        id="agg-exclude-plant-loading",
                                        type="circle",
                                        children=dmc.MultiSelect(
                                            id="agg-exclude-plant-dd",
                                            data=[],
                                            value=[],
                                            clearable=True,
                                            placeholder="Select plants to exclude from aggregation...",
                                            searchable=True,
                                            comboboxProps={"keepOpened": True},
                                            maxDropdownHeight=300,
                                        ),
                                    ),
                                ], style={
                                    "padding": "1rem",
                                    "backgroundColor": "#fff3cd",
                                    "border": "1px solid #ffc107",
                                    "borderRadius": "8px",
                                    "marginBottom": "1rem"
                                })
                            ], md=12),
                        ], className="mb-3"),
                    ], id="agg-exclude-section", style={"display": "none"}),
                    # Custom Settings (presets) selection
                    dbc.Row([
                        dbc.Col([
                            html.Label("Custom Setting(s)", style={"fontWeight": 600, "marginBottom": "10px", "color": "#333", "fontSize": "0.95rem"}),
                            dmc.MultiSelect(
                                id="agg-analysis-preset-select",
                                data=[],
                                value=[],
                                searchable=True,
                                clearable=True,
                                placeholder="(Optional) Choose one or more saved presets",
                                comboboxProps={"keepOpened": True},
                                maxDropdownHeight=300,
                            )
                        ], md=4),
                    ], className="mb-3"),
                    # PPA Configuration
                    html.Div([
                        html.Div([
                            html.Span("⚡", style={"fontSize": "1.3rem", "marginRight": "8px"}),
                            html.Span("PPA Configuration", style={"fontSize": "1.05rem", "fontWeight": "600", "color": "#333"})
                        ], style={"marginBottom": "0.75rem"}),
                        dbc.Row([
                            dbc.Col([
                                dcc.Dropdown(
                                    id="agg-ppa-mode",
                                    options=[
                                        {"label": "Mean", "value": "mean"},
                                        {"label": "Median", "value": "median"},
                                        {"label": "Mode", "value": "mode"},
                                        {"label": "Weighted Average", "value": "weighted"},
                                        {"label": "Numeric", "value": "numeric"},
                                    ],
                                    value=["mean"],
                                    multi=True,
                                    clearable=False,
                                    placeholder="Select one or more PPA methods...",
                                    style={"fontSize": "0.95rem"},
                                ),
                            ], md=6),
                            dbc.Col([
                                html.Div([
                                    html.Label("Numeric PPA (₹/kWh)", style={"fontWeight": "500", "marginBottom": "6px", "fontSize": "0.9rem", "color": "#555"}),
                                    dbc.Input(
                                        id="agg-ppa-value",
                                        type="number",
                                        value=None,
                                        min=0,
                                        step=0.01,
                                        placeholder="Enter PPA value",
                                        style={"fontSize": "0.9rem"}
                                    ),
                                ], id="agg-ppa-numeric-container", style={"display": "none"}),
                            ], md=6),
                        ]),
                    ], style={"marginBottom": "1.5rem", "padding": "1.25rem", "backgroundColor": "#f8f9fa", "borderRadius": "8px"}),
                    # Action buttons
                    dbc.Row([
                        dbc.Col([
                            dbc.Button([
                                html.Span("📊 ", style={"marginRight": "8px"}),
                                "Plot Now"
                            ], id="agg-plot-now", size="md", className="w-100", disabled=True,
                               style={"backgroundColor": "#ff6b35", "border": "none", "fontWeight": "600", "padding": "10px",
                                      "boxShadow": "0 4px 6px rgba(255, 107, 53, 0.3)"}),
                        ], md=4, className="offset-md-4"),
                    ]),
                    # Progress bar
                    dbc.Row([
                        dbc.Col([
                            html.Div(id="agg-progress-container", style={"display": "none", "marginTop": "1.5rem"}, children=[
                                html.Div([
                                    html.Span("⏳ ", style={"marginRight": "8px"}),
                                    html.Span("Processing aggregated data...", style={"fontWeight": "500", "color": "#333"})
                                ], style={"marginBottom": "10px", "textAlign": "center"}),
                                dbc.Progress(
                                    id="agg-compute-progress",
                                    value=100,
                                    striped=True,
                                    animated=True,
                                    style={"height": "25px"},
                                    className="mb-2"
                                ),
                            ]),
                        ], md=8, className="offset-md-2")
                    ]),
                ], style={"padding": "2rem", "backgroundColor": "#fff", "borderRadius": "12px", "marginBottom": "2rem",
                         "boxShadow": "0 2px 8px rgba(0,0,0,0.05)"}),
                # Results Section
                html.Div(id="agg-results-section", style={"display": "none"}, children=[
                    html.Div(id="agg-tab-content", className="mt-3"),
                ]),
            ], id="aggregation-analysis-content", style={"display": "none"}),

            # Battery Investment Content (modular)
            battery_investment_layout(),

            # Custom Upload Content
            html.Div([
                # Custom Upload intro (header handled globally)
                html.Div([]),
                # Download sample buttons
                html.Div([
                    dbc.Row([
                        dbc.Col([
                            dbc.Button("Download XLSX Sample", id="btn-download-sample-xlsx", color="secondary", className="me-2"),
                            dbc.Button("Download CSV Sample", id="btn-download-sample-csv", color="secondary"),
                        ], md=12),
                    ], className="mb-3"),
                ], style={"padding": "2rem", "backgroundColor": "#fff", "borderRadius": "12px", "marginBottom": "1rem",
                         "boxShadow": "0 2px 8px rgba(0,0,0,0.05)"}),
                dcc.Download(id="dl-sample-xlsx"),
                dcc.Download(id="dl-sample-csv"),

                # Upload control and preset select
                html.Div([
                    dcc.Upload(
                        id="upload-custom-file",
                        children=html.Div(["Drag & drop or ", html.A("choose a CSV/XLSX")]),
                        multiple=False,
                        style={"border":"1px dashed #bbb", "padding":"16px", "borderRadius":"8px"}
                    ),
                    html.Div(id="custom-upload-validate", className="mt-2"),

                    html.Label("Custom Setting(s)", className="mt-3", style={"fontWeight": 600, "marginBottom": "10px", "color": "#333", "fontSize": "0.95rem"}),
                    dcc.Dropdown(id="custom-upload-preset-select", multi=True, placeholder="Choose saved preset(s)"),
                    dbc.Button("Run Analysis", id="btn-run-custom", color="primary", className="mt-3", disabled=True),
                ], style={"padding": "2rem", "backgroundColor": "#fff", "borderRadius": "12px", "marginBottom": "1rem",
                         "boxShadow": "0 2px 8px rgba(0,0,0,0.05)"}),

                # Results and download
                html.Div([
                    html.Div(id="custom-results"),
                    dbc.Button("Download Output (Excel)", id="download-custom-output-btn", className="mt-3", color="success"),
                    dcc.Download(id="download-custom-output"),
                ], style={"padding": "2rem", "backgroundColor": "#fff", "borderRadius": "12px", "marginBottom": "2rem",
                         "boxShadow": "0 2px 8px rgba(0,0,0,0.05)"}),
            ], id="custom-upload-content", style={"display": "none"}),
            # Data Statistics Content
            html.Div([
                # Data Statistics intro (header handled globally)
                html.Div([]),
                html.Div([
                    dbc.Row([
                        dbc.Col([
                            html.Label([html.Span("🌐 ", style={"marginRight": "6px"}), "Region"], 
                                       style={"fontWeight": "600", "marginBottom": "10px", "color": "#333", "fontSize": "0.95rem"}),
                            dcc.Dropdown(id="stats-region-dd", options=[], value=[], multi=True, clearable=False,
                                         placeholder="Select Region(s) from DuckDB...", style={"fontSize": "0.95rem"}),
                        ], md=4),
                        dbc.Col([
                            html.Label([html.Span("🏭 ", style={"marginRight": "6px"}), "Resource"], 
                                       style={"fontWeight": "600", "marginBottom": "10px", "color": "#333", "fontSize": "0.95rem"}),
                            dcc.Dropdown(id="stats-resource-dd", options=[
                                {"label": "All", "value": "ALL"},
                                {"label": "Solar", "value": "SOLAR"},
                                {"label": "Wind", "value": "WIND"},
                            ], value="ALL", clearable=False, style={"fontSize": "0.95rem"}),
                        ], md=2),
                        dbc.Col([
                            html.Label([html.Span("🏭 ", style={"marginRight": "6px"}), "Plant"], 
                                       style={"fontWeight": "600", "marginBottom": "10px", "color": "#333", "fontSize": "0.95rem"}),
                            dcc.Dropdown(id="stats-plant-dd", options=[], value=[], multi=True, clearable=False,
                                         placeholder="Select Plant(s) or Select All...", style={"fontSize": "0.95rem"}),
                        ], md=4),
                        dbc.Col([
                            html.Label([html.Span("📅 ", style={"marginRight": "6px"}), "Date Range"], 
                                       style={"fontWeight": "600", "marginBottom": "10px", "color": "#333", "fontSize": "0.95rem"}),
                            dcc.DatePickerRange(id="stats-date-range", start_date=(datetime.now()-timedelta(days=6)).date(),
                                                end_date=datetime.now().date(), display_format="DD-MMM-YYYY", minimum_nights=0,
                                                style={"fontSize": "0.95rem"}),
                        ], md=2),
                    ], className="mb-3"),
                    dbc.Row([
                        dbc.Col([
                            dbc.Button([html.Span("📈 ", style={"marginRight": "8px"}), "Run"], id="stats-run", size="lg", className="w-100",
                                       style={"backgroundColor": "#ff6b35", "border": "none", "fontWeight": "600", "padding": "14px"}),
                        ], md=6, className="offset-md-3")
                    ]),
                ], style={"padding": "2rem", "backgroundColor": "#fff", "borderRadius": "12px", "marginBottom": "2rem",
                         "boxShadow": "0 2px 8px rgba(0,0,0,0.05)"}),
                html.Div(id="stats-results")
            ], id="stats-content", style={"display": "none"}),
            
        ], style={"padding": "2rem"})
    ], style={"marginLeft": "280px", "minHeight": "100vh", "backgroundColor": "#f8f9fa"})

app.layout = dmc.MantineProvider(
    children=html.Div([
        dcc.Store(id="results-store"),
        dcc.Store(id="agg-results-store"),
        # Persist nav selection across reloads (helps if browser tab reloads after a crash)
        dcc.Store(id="nav-store", data="welcome", storage_type="session"),
        dcc.Store(id="saved-settings-store", storage_type="local"),
        dcc.Store(id="presets-store", storage_type="local"),
        dcc.Store(id="custom-upload-store"),
        dcc.Store(id="custom-results-store"),
        dcc.Interval(id="progress-interval", interval=400, n_intervals=0, disabled=True),
        sidebar(),
        main_content(),
        dcc.Download(id="download-excel"),
        dcc.Download(id="download-plant-summary"),
        # ---- Download progress overlay ----
        dcc.Store(id="dl-progress-store", data={"active": False, "t0": 0, "label": "", "est_sec": 10}),
        dcc.Store(id="dl-done-store", data=0),
        dcc.Interval(id="dl-progress-interval", interval=300, n_intervals=0, disabled=False),
        html.Div(
            id="dl-progress-overlay",
            style={
                "display": "none",
                "position": "fixed",
                "top": 0, "left": 0, "right": 0, "bottom": 0,
                "zIndex": 9999,
                "backgroundColor": "rgba(15, 20, 40, 0.55)",
                "alignItems": "center",
                "justifyContent": "center",
                "backdropFilter": "blur(4px)",
            },
            children=[
                html.Div(
                    style={
                        "backgroundColor": "#ffffff",
                        "borderRadius": "18px",
                        "padding": "36px 44px",
                        "minWidth": "380px",
                        "maxWidth": "500px",
                        "boxShadow": "0 24px 64px rgba(0,0,0,0.25)",
                        "textAlign": "center",
                    },
                    children=[
                        html.Div(
                            "\u2b07\ufe0f",
                            style={"fontSize": "2rem", "marginBottom": "8px"}
                        ),
                        html.Div(
                            "Preparing Download",
                            style={"fontWeight": "700", "fontSize": "1.1rem", "color": "#1a1a2e", "marginBottom": "4px"}
                        ),
                        html.Div(
                            id="dl-progress-label",
                            style={"color": "#666", "fontSize": "0.85rem", "marginBottom": "22px"}
                        ),
                        html.Div([
                            dbc.Progress(
                                id="dl-progress-bar",
                                value=0,
                                striped=True,
                                animated=True,
                                color="primary",
                                style={"height": "26px", "borderRadius": "13px"},
                            ),
                            html.Div(
                                id="dl-progress-pct",
                                style={
                                    "position": "absolute",
                                    "top": 0, "left": 0, "right": 0,
                                    "lineHeight": "26px",
                                    "textAlign": "center",
                                    "fontWeight": "700",
                                    "fontSize": "0.82rem",
                                    "color": "#fff",
                                    "pointerEvents": "none",
                                    "textShadow": "0 1px 3px rgba(0,0,0,0.4)",
                                }
                            ),
                        ], style={"position": "relative", "marginBottom": "14px"}),
                        html.Div(
                            id="dl-progress-time",
                            style={"color": "#888", "fontSize": "0.78rem", "minHeight": "20px"}
                        ),
                    ]
                )
            ]
        ),
    ])
)

# Register Battery Investment callbacks (modular)
register_battery_callbacks(app)

# ==========================================
# ---------------- CALLBACKS ---------------
# ==========================================
@app.callback(
    Output("nav-welcome", "style"),
    Output("nav-settings", "style"),
    Output("nav-analysis", "style"),
    Output("nav-aggregation-analysis", "style"),
    Output("nav-battery-investment", "style"),
    Output("nav-custom-upload", "style"),
    Output("nav-stats", "style"),
    Output("welcome-content", "style"),
    Output("custom-settings-content", "style"),
    Output("analysis-content", "style"),
    Output("aggregation-analysis-content", "style"),
    Output("battery-investment-content", "style"),
    Output("custom-upload-content", "style"),
    Output("stats-content", "style"),
    Output("nav-store", "data"),
    Input("nav-welcome", "n_clicks"),
    Input("nav-settings", "n_clicks"),
    Input("nav-analysis", "n_clicks"),
    Input("nav-aggregation-analysis", "n_clicks"),
    Input("nav-battery-investment", "n_clicks"),
    Input("nav-custom-upload", "n_clicks"),
    Input("nav-stats", "n_clicks"),
    Input("btn-get-started", "n_clicks"),
    State("nav-store", "data"),
)
def switch_nav_tabs(welcome_clicks, settings_clicks, analysis_clicks, agg_analysis_clicks, battery_clicks, custom_upload_clicks, nav_stats, get_started_clicks, current):
    ctx_triggered = ctx.triggered_id
    target = current or "welcome"

    if ctx_triggered == "nav-welcome":
        target = "welcome"
    elif ctx_triggered == "nav-settings":
        target = "settings"
    elif ctx_triggered == "nav-analysis":
        target = "analysis"
    elif ctx_triggered == "nav-aggregation-analysis":
        target = "aggregation_analysis"
    elif ctx_triggered == "nav-battery-investment":
        target = "battery_investment"
    elif ctx_triggered == "nav-custom-upload":
        target = "custom_upload"
    elif ctx_triggered == "nav-stats":
        target = "stats"
    elif ctx_triggered == "btn-get-started":
        target = "settings"

    # Define nav item styles
    active_style = {
        "padding": "12px 20px", "cursor": "pointer", "borderLeft": "3px solid #ff6b35",
        "backgroundColor": "rgba(255, 107, 53, 0.1)", "marginBottom": "4px",
        "color": "#ff6b35", "fontWeight": "500"
    }
    inactive_style = {
        "padding": "12px 20px", "cursor": "pointer", "marginBottom": "4px", "color": "white"
    }

    if target == "welcome":
        return (
            active_style,
            inactive_style,
            inactive_style,
            inactive_style,
            inactive_style,
            inactive_style,
            inactive_style,
            {"display": "block"},
            {"display": "none"},
            {"display": "none"},
            {"display": "none"},
            {"display": "none"},
            {"display": "none"},
            {"display": "none"},
            target,
        )
    elif target == "settings":
        return (
            inactive_style,
            active_style,
            inactive_style,
            inactive_style,
            inactive_style,
            inactive_style,
            inactive_style,
            {"display": "none"},
            {"display": "block"},
            {"display": "none"},
            {"display": "none"},
            {"display": "none"},
            {"display": "none"},
            {"display": "none"},
            target,
        )
    elif target == "analysis":
        return (
            inactive_style,
            inactive_style,
            active_style,
            inactive_style,
            inactive_style,
            inactive_style,
            inactive_style,
            {"display": "none"},
            {"display": "none"},
            {"display": "block"},
            {"display": "none"},
            {"display": "none"},
            {"display": "none"},
            {"display": "none"},
            target,
        )
    elif target == "aggregation_analysis":
        return (
            inactive_style,
            inactive_style,
            inactive_style,
            active_style,
            inactive_style,
            inactive_style,
            inactive_style,
            {"display": "none"},
            {"display": "none"},
            {"display": "none"},
            {"display": "block"},
            {"display": "none"},
            {"display": "none"},
            {"display": "none"},
            target,
        )
    elif target == "battery_investment":
        return (
            inactive_style,
            inactive_style,
            inactive_style,
            inactive_style,
            active_style,
            inactive_style,
            inactive_style,
            {"display": "none"},
            {"display": "none"},
            {"display": "none"},
            {"display": "none"},
            {"display": "block"},
            {"display": "none"},
            {"display": "none"},
            target,
        )
    elif target == "custom_upload":
        return (
            inactive_style,
            inactive_style,
            inactive_style,
            inactive_style,
            inactive_style,
            active_style,
            inactive_style,
            {"display": "none"},
            {"display": "none"},
            {"display": "none"},
            {"display": "none"},
            {"display": "none"},
            {"display": "block"},
            {"display": "none"},
            target,
        )
    else:  # stats
        return (
            inactive_style,
            inactive_style,
            inactive_style,
            inactive_style,
            inactive_style,
            inactive_style,
            active_style,
            {"display": "none"},
            {"display": "none"},
            {"display": "none"},
            {"display": "none"},
            {"display": "none"},
            {"display": "none"},
            {"display": "block"},
            target,
        )


@app.callback(
    Output("page-title", "children"),
    Output("page-subtitle", "children"),
    Input("nav-store", "data"),
)
def update_page_header(nav_key):
    """Set main page header based on current sidebar selection."""
    nav_key = nav_key or "welcome"
    title_map = {
        "welcome": "Welcome",
        "settings": "Custom Settings",
        "analysis": "Analysis",
        "aggregation_analysis": "Aggregation Analysis",
        "battery_investment": "Battery Investment",
        "custom_upload": "Custom Upload",
        "stats": "Data Statistics",
    }
    subtitle_map = {
        "welcome": "Start here to understand how to use the DSM analytics tools.",
        "settings": "Configure error modes and DSM deviation bands.",
        "analysis": f"Run DSM deviation analysis for selected plants.  |  Build: {BUILD_TAG}",
        "aggregation_analysis": "Aggregate multiple plants, then run DSM analysis on the combined profile.",
        "battery_investment": "Plan battery size by visualizing block-level energy deviation (planning only; not settlement).",
        "custom_upload": "Upload your own DSM dataset and apply saved presets.",
        "stats": "Compute availability statistics for selected plants.",
    }
    return title_map.get(nav_key, "Welcome"), subtitle_map.get(nav_key, "")

# Callback to load regions from DuckDB on page load
@app.callback(
    Output("region-dd", "data"),
    Input("nav-store", "data"),
    prevent_initial_call=False
)
def load_regions_from_duckdb(nav_data):
    """Load regions (prefer Excel master for Analysis cascading filters)."""
    try:
        df = _load_plant_master_df()
        regions = sorted({str(r).upper() for r in df.get("region", pd.Series(dtype=str)).dropna().tolist() if str(r).strip()})
        if not regions:
            regions = get_regions_from_duckdb()  # fallback
        return [{"label": r, "value": r} for r in regions]
    except Exception as e:
        print(f"Error loading regions: {e}")
        return [{"label": r, "value": r} for r in ["NRPC", "SRPC", "WRPC"]]


@app.callback(
    Output("agg-region-dd", "data"),
    Input("nav-store", "data"),
    prevent_initial_call=False
)
def load_regions_for_agg(nav_data):
    """Load regions for Aggregation Analysis from DuckDB databases."""
    try:
        df = _load_plant_master_df()
        regions = sorted({str(r).upper() for r in df.get("region", pd.Series(dtype=str)).dropna().tolist() if str(r).strip()})
        if not regions:
            regions = get_regions_from_duckdb()  # fallback
        return [{"label": r, "value": r} for r in regions]
    except Exception as e:
        print(f"Error loading regions (agg): {e}")
        return [{"label": r, "value": r} for r in ["NRPC", "SRPC", "WRPC"]]

# Regions for stats page
@app.callback(
    Output("stats-region-dd", "options"),
    Input("nav-store", "data"),
    prevent_initial_call=False
)
def load_regions_stats(nav_data):
    try:
        regions = get_regions_from_duckdb()
        return [{"label": r, "value": r} for r in regions]
    except Exception as e:
        print(f"Error loading regions: {e}")
        return [{"label": r, "value": r} for r in ["NRPC", "SRPC", "WRPC"]]

@app.callback(
    Output("state-dd", "data"),
    Output("state-dd", "value"),
    Input("region-dd", "value"),
    State("state-dd", "value"),
    prevent_initial_call=False
)
def update_states_from_master(regions, current_state_value):
    if not regions:
        return [], []
    df = _filter_master(regions, None, None, None, None)
    states = sorted({s for s in df.get("state", pd.Series(dtype=str)).dropna().astype(str).tolist() if s.strip()})
    opts = [{"label": "✓ All", "value": ALL_SENTINEL}] + [{"label": s, "value": s} for s in states]

    cur = _strip_all_sentinel(_normalize_multi(current_state_value))
    if not cur:
        return opts, []
    # Keep only valid
    valid = [v for v in cur if v in states]
    return opts, (valid if valid else [])


@app.callback(
    Output("resource-type-dd", "data"),
    Output("resource-type-dd", "value"),
    Input("region-dd", "value"),
    Input("state-dd", "value"),
    State("resource-type-dd", "value"),
    prevent_initial_call=False
)
def update_resources_from_master(regions, states, current_resource_value):
    if not regions:
        return [], []
    df = _filter_master(regions, states, None, None, None)
    res = sorted({r for r in df.get("resource", pd.Series(dtype=str)).dropna().astype(str).tolist() if r.strip()})
    opts = [{"label": "✓ All", "value": ALL_SENTINEL}] + [{"label": r, "value": r} for r in res]

    cur = _strip_all_sentinel(_normalize_multi(current_resource_value))
    if not cur:
        return opts, []
    valid = [v for v in cur if v in res]
    return opts, (valid if valid else [])


@app.callback(
    Output("qca-dd", "data"),
    Output("qca-dd", "value"),
    Input("region-dd", "value"),
    Input("state-dd", "value"),
    Input("resource-type-dd", "value"),
    State("qca-dd", "value"),
    prevent_initial_call=False
)
def update_qca_from_master(regions, states, resources, current_qca_value):
    if not regions:
        return [], []
    df = _filter_master(regions, states, resources, None, None)
    qcas = sorted({q for q in df.get("qca", pd.Series(dtype=str)).dropna().astype(str).tolist() if q.strip()})
    opts = [{"label": "✓ All", "value": ALL_SENTINEL}] + [{"label": q, "value": q} for q in qcas]

    cur = _strip_all_sentinel(_normalize_multi(current_qca_value))
    if not cur:
        return opts, []
    valid = [v for v in cur if v in qcas]
    return opts, (valid if valid else [])


@app.callback(
    Output("pooling-dd", "data"),
    Output("pooling-dd", "value"),
    Input("region-dd", "value"),
    Input("state-dd", "value"),
    Input("resource-type-dd", "value"),
    Input("qca-dd", "value"),
    State("pooling-dd", "value"),
    prevent_initial_call=False
)
def update_pooling_from_master(regions, states, resources, qcas, current_pool_value):
    if not regions:
        return [], []
    df = _filter_master(regions, states, resources, qcas, None)
    pools = sorted({p for p in df.get("pooling_station", pd.Series(dtype=str)).dropna().astype(str).tolist() if p.strip()})
    opts = [{"label": "✓ All", "value": ALL_SENTINEL}] + [{"label": p, "value": p} for p in pools]

    cur = _strip_all_sentinel(_normalize_multi(current_pool_value))
    if not cur:
        return opts, []
    valid = [v for v in cur if v in pools]
    return opts, (valid if valid else [])


@app.callback(
    Output("plant-dd", "data"),
    Output("plant-dd", "value"),
    Input("region-dd", "value"),
    Input("state-dd", "value"),
    Input("resource-type-dd", "value"),
    Input("qca-dd", "value"),
    Input("pooling-dd", "value"),
    State("plant-dd", "value"),
    prevent_initial_call=False
)
def update_plants_from_duckdb(regions, states, resource_type, qcas, pools, current_plant_value):
    """Update plant options using consolidated_plant_list.xlsx (cascading, single source of truth)."""
    if not regions:
        return [], []

    df = _filter_master(regions, states, resource_type, qcas, pools)
    filtered_plants = sorted({p for p in df.get("plant_name", pd.Series(dtype=str)).dropna().astype(str).tolist() if p.strip()})
    if not filtered_plants:
        return [], []

    def truncate_label(name, max_length=50):
        return name if len(name) <= max_length else name[:max_length-3] + "..."

    # Get currently selected plants
    cur = _normalize_multi(current_plant_value)
    selected_plants = set(cur) if cur else set()
    
    # Remove already-selected plants from dropdown options (except SELECT_ALL)
    available_plants = [p for p in filtered_plants if p not in selected_plants]

    options = [{"label": "✓ Select All", "value": "SELECT_ALL"}] + [
        {"label": truncate_label(p), "value": p} for p in available_plants
    ]
    
    # Preserve current selection if still valid; otherwise clear it.
    if cur:
        valid = [p for p in cur if p == "SELECT_ALL" or p in filtered_plants]
        return options, valid
    return options, []


@app.callback(
    Output("agg-state-dd", "data"),
    Output("agg-state-dd", "value"),
    Input("agg-region-dd", "value"),
    State("agg-state-dd", "value"),
    prevent_initial_call=False
)
def update_agg_states_from_master(regions, current_state_value):
    if not regions:
        return [], []
    df = _filter_master(regions, None, None, None, None)
    states = sorted({s for s in df.get("state", pd.Series(dtype=str)).dropna().astype(str).tolist() if s.strip()})
    opts = [{"label": "✓ All", "value": ALL_SENTINEL}] + [{"label": s, "value": s} for s in states]
    cur = _strip_all_sentinel(_normalize_multi(current_state_value))
    if not cur:
        return opts, []
    valid = [v for v in cur if v in states]
    return opts, (valid if valid else [])


@app.callback(
    Output("agg-resource-type-dd", "data"),
    Output("agg-resource-type-dd", "value"),
    Input("agg-region-dd", "value"),
    Input("agg-state-dd", "value"),
    State("agg-resource-type-dd", "value"),
    prevent_initial_call=False
)
def update_agg_resources_from_master(regions, states, current_resource_value):
    if not regions:
        return [], []
    df = _filter_master(regions, states, None, None, None)
    res = sorted({r for r in df.get("resource", pd.Series(dtype=str)).dropna().astype(str).tolist() if r.strip()})
    opts = [{"label": "✓ All", "value": ALL_SENTINEL}] + [{"label": r, "value": r} for r in res]
    cur = _strip_all_sentinel(_normalize_multi(current_resource_value))
    if not cur:
        return opts, []
    valid = [v for v in cur if v in res]
    return opts, (valid if valid else [])


@app.callback(
    Output("agg-qca-dd", "data"),
    Output("agg-qca-dd", "value"),
    Input("agg-region-dd", "value"),
    Input("agg-state-dd", "value"),
    Input("agg-resource-type-dd", "value"),
    State("agg-qca-dd", "value"),
    prevent_initial_call=False
)
def update_agg_qca_from_master(regions, states, resources, current_qca_value):
    if not regions:
        return [], []
    df = _filter_master(regions, states, resources, None, None)
    qcas = sorted({q for q in df.get("qca", pd.Series(dtype=str)).dropna().astype(str).tolist() if q.strip()})
    opts = [{"label": "✓ All", "value": ALL_SENTINEL}] + [{"label": q, "value": q} for q in qcas]
    cur = _strip_all_sentinel(_normalize_multi(current_qca_value))
    if not cur:
        return opts, []
    valid = [v for v in cur if v in qcas]
    return opts, (valid if valid else [])


@app.callback(
    Output("agg-pooling-dd", "data"),
    Output("agg-pooling-dd", "value"),
    Input("agg-region-dd", "value"),
    Input("agg-state-dd", "value"),
    Input("agg-resource-type-dd", "value"),
    Input("agg-qca-dd", "value"),
    State("agg-pooling-dd", "value"),
    prevent_initial_call=False
)
def update_agg_pooling_from_master(regions, states, resources, qcas, current_pool_value):
    if not regions:
        return [], []
    df = _filter_master(regions, states, resources, qcas, None)
    pools = sorted({p for p in df.get("pooling_station", pd.Series(dtype=str)).dropna().astype(str).tolist() if p.strip()})
    opts = [{"label": "✓ All", "value": ALL_SENTINEL}] + [{"label": p, "value": p} for p in pools]
    cur = _strip_all_sentinel(_normalize_multi(current_pool_value))
    if not cur:
        return opts, []
    valid = [v for v in cur if v in pools]
    return opts, (valid if valid else [])


@app.callback(
    Output("agg-plant-dd", "data"),
    Output("agg-plant-dd", "value"),
    Input("agg-region-dd", "value"),
    Input("agg-state-dd", "value"),
    Input("agg-resource-type-dd", "value"),
    Input("agg-qca-dd", "value"),
    Input("agg-pooling-dd", "value"),
    State("agg-plant-dd", "value"),
    prevent_initial_call=False
)
def update_agg_plants_from_master(regions, states, resources, qcas, pools, current_plant_value):
    if not regions:
        return [], []
    df = _filter_master(regions, states, resources, qcas, pools)
    filtered_plants = sorted({p for p in df.get("plant_name", pd.Series(dtype=str)).dropna().astype(str).tolist() if p.strip()})
    if not filtered_plants:
        return [], []

    def truncate_label(name, max_length=50):
        return name if len(name) <= max_length else name[:max_length-3] + "..."

    # Get currently selected plants
    cur = _normalize_multi(current_plant_value)
    selected_plants = set(cur) if cur else set()
    
    # Remove already-selected plants from dropdown options (except SELECT_ALL)
    available_plants = [p for p in filtered_plants if p not in selected_plants]
    
    options = [{"label": "✓ Select All", "value": "SELECT_ALL"}] + [{"label": truncate_label(p), "value": p} for p in available_plants]
    
    if cur:
        valid = [p for p in cur if p == "SELECT_ALL" or p in filtered_plants]
        return options, valid
    return options, []


@app.callback(
    Output("agg-exclude-section", "style"),
    Output("agg-exclude-plant-dd", "data"),
    Output("agg-exclude-plant-dd", "value"),
    Input("agg-plant-dd", "value"),
    Input("agg-region-dd", "value"),
    Input("agg-state-dd", "value"),
    Input("agg-resource-type-dd", "value"),
    Input("agg-qca-dd", "value"),
    Input("agg-pooling-dd", "value"),
    State("agg-exclude-plant-dd", "value"),
    prevent_initial_call=False
)
def toggle_exclude_section(plant_value, regions, states, resources, qcas, pools, current_exclude_value):
    """Show/hide exclude section and populate exclude dropdown when SELECT_ALL is active."""
    # Check if SELECT_ALL is selected
    is_select_all = False
    if isinstance(plant_value, list):
        is_select_all = "SELECT_ALL" in plant_value
    elif plant_value == "SELECT_ALL":
        is_select_all = True
    
    # Hide section if SELECT_ALL is not active
    if not is_select_all:
        return {"display": "none"}, [], []
    
    # Show section and populate options
    if not regions:
        return {"display": "block"}, [], current_exclude_value if current_exclude_value else []
    
    df = _filter_master(regions, states, resources, qcas, pools)
    filtered_plants = sorted({p for p in df.get("plant_name", pd.Series(dtype=str)).dropna().astype(str).tolist() if p.strip()})
    
    if not filtered_plants:
        return {"display": "block"}, [], current_exclude_value if current_exclude_value else []
    
    def truncate_label(name, max_length=50):
        return name if len(name) <= max_length else name[:max_length-3] + "..."
    
    exclude_options = [
        {"label": truncate_label(p), "value": p} for p in filtered_plants
    ]
    
    # Filter out invalid excluded plants if filter changed
    if current_exclude_value:
        if isinstance(current_exclude_value, list):
            valid_excluded = [p for p in current_exclude_value if p in filtered_plants]
            return {"display": "block"}, exclude_options, valid_excluded
        elif current_exclude_value in filtered_plants:
            return {"display": "block"}, exclude_options, current_exclude_value
    
    return {"display": "block"}, exclude_options, current_exclude_value if current_exclude_value else []


@app.callback(
    Output("exclude-section", "style"),
    Output("exclude-plant-dd", "data"),
    Output("exclude-plant-dd", "value"),
    Output("exclude-plant-dd", "disabled"),
    Input("plant-dd", "value"),
    Input("region-dd", "value"),
    Input("state-dd", "value"),
    Input("resource-type-dd", "value"),
    Input("qca-dd", "value"),
    Input("pooling-dd", "value"),
    State("exclude-plant-dd", "value"),
    prevent_initial_call=False
)
def toggle_exclude_section_analysis(plant_value, regions, states, resource_type, qcas, pools, current_exclude_value):
    """Show/hide exclude section and populate exclude dropdown when SELECT_ALL is active in Analysis tab."""
    # Check if SELECT_ALL is selected
    is_select_all = False
    if plant_value:
        if isinstance(plant_value, list):
            is_select_all = "SELECT_ALL" in plant_value
        elif plant_value == "SELECT_ALL":
            is_select_all = True
    
    # Hide section if SELECT_ALL is not active
    if not is_select_all:
        return {"display": "none"}, [], [], True
    
    # Show section and populate options
    if not regions:
        return {"display": "block"}, [], (current_exclude_value if current_exclude_value else []), False
    
    df = _filter_master(regions, states, resource_type, qcas, pools)
    filtered_plants = sorted({p for p in df.get("plant_name", pd.Series(dtype=str)).dropna().astype(str).tolist() if p.strip()})
    
    if not filtered_plants:
        return {"display": "block"}, [], (current_exclude_value if current_exclude_value else []), False
    
    def truncate_label(name, max_length=50):
        return name if len(name) <= max_length else name[:max_length-3] + "..."
    
    exclude_options = [
        {"label": truncate_label(p), "value": p} for p in filtered_plants
    ]
    
    # Filter out invalid excluded plants if filter changed
    if current_exclude_value:
        if isinstance(current_exclude_value, list):
            valid_excluded = [p for p in current_exclude_value if p in filtered_plants]
            return {"display": "block"}, exclude_options, valid_excluded, False
        elif current_exclude_value in filtered_plants:
            return {"display": "block"}, exclude_options, current_exclude_value, False
    
    return {"display": "block"}, exclude_options, (current_exclude_value if current_exclude_value else []), False

# Stats plants based on filters
@app.callback(
    Output("stats-plant-dd", "options"),
    Output("stats-plant-dd", "value"),
    Input("stats-region-dd", "value"),
    Input("stats-resource-dd", "value"),
    Input("stats-date-range", "start_date"),
    Input("stats-date-range", "end_date"),
    prevent_initial_call=False
)
def update_stats_plants(regions, resource_type, start_date, end_date):
    if not regions:
        return [], []
    sd = str(start_date) if start_date else str(datetime.now().date())
    ed = str(end_date) if end_date else sd
    try:
        filtered_plants = get_filtered_plants_by_type(regions, resource_type or "ALL", sd, ed)
    except Exception:
        filtered_plants = get_plants_from_duckdb(regions)
    if not filtered_plants:
        return [], []
    def trunc(n, L=50):
        return n if len(n) <= L else n[:L-3] + "..."
    opts = [{"label": "✓ Select All", "value": "SELECT_ALL"}] + [{"label": trunc(p), "value": p} for p in filtered_plants]
    return opts, []

# Disable Run Analysis until all inputs present
@app.callback(
    Output("plot-now", "disabled"),
    Input("region-dd", "value"),
    Input("resource-type-dd", "value"),
    Input("state-dd", "value"),
    Input("qca-dd", "value"),
    Input("pooling-dd", "value"),
    Input("plant-dd", "value"),
    Input("exclude-plant-dd", "value"),
    Input("date-range", "start_date"),
    Input("date-range", "end_date"),
    Input("custom-setting-mode", "value"),
    Input("setting-assignments-store", "data"),
    prevent_initial_call=False
)
def toggle_plot_now(regions, resource_type, states, qcas, pools, plants, excluded_plants, start_date, end_date, mode, assignments_store):
    has_regions = bool(regions)
    has_plants = bool(plants)
    has_dates = bool(start_date) and bool(end_date)
    
    # Validate that not all plants are excluded when SELECT_ALL is active
    if has_plants:
        is_select_all = False
        if isinstance(plants, list):
            is_select_all = "SELECT_ALL" in plants
        elif plants == "SELECT_ALL":
            is_select_all = True
        
        if is_select_all and excluded_plants:
            # Check if all plants would be excluded
            df = _filter_master(regions, states, resource_type, qcas, pools)
            filtered_plants = sorted({p for p in df.get("plant_name", pd.Series(dtype=str)).dropna().astype(str).tolist() if p.strip()})
            
            if filtered_plants:
                excluded_list = excluded_plants if isinstance(excluded_plants, list) else [excluded_plants]
                remaining = [p for p in filtered_plants if p not in excluded_list]
                if not remaining:
                    # All plants excluded - disable button
                    return True
    
    # Validate plant-wise settings when in PLANT_WISE mode
    application_mode = mode if mode else "GLOBAL"
    if application_mode == "PLANT_WISE":
        store = assignments_store or {}
        rows_state = list(store.get("rows") or [])
        if not rows_state:
            return True
        # Require a preset selected for each row
        if any(not (r.get("preset")) for r in rows_state):
            return True
        # If any row is SELECTED scope, require at least one plant in that row
        for r in rows_state:
            if not bool(r.get("apply_remaining", False)):
                pl = r.get("plants") or []
                if not pl:
                    return True
    
    disabled = not (has_regions and has_plants and has_dates)
    return disabled


@app.callback(
    Output("agg-plot-now", "disabled"),
    Input("agg-region-dd", "value"),
    Input("agg-state-dd", "value"),
    Input("agg-resource-type-dd", "value"),
    Input("agg-qca-dd", "value"),
    Input("agg-pooling-dd", "value"),
    Input("agg-plant-dd", "value"),
    Input("agg-exclude-plant-dd", "value"),
    Input("agg-date-range", "start_date"),
    Input("agg-date-range", "end_date"),
    prevent_initial_call=False
)
def toggle_agg_buttons(regions, states, resources, qcas, pools, plants, excluded_plants, start_date, end_date):
    has_regions = bool(regions)
    has_plants = bool(plants)
    has_dates = bool(start_date) and bool(end_date)
    
    # Validate that not all plants are excluded when SELECT_ALL is active
    if has_plants:
        is_select_all = False
        if isinstance(plants, list):
            is_select_all = "SELECT_ALL" in plants
        elif plants == "SELECT_ALL":
            is_select_all = True
        
        if is_select_all and excluded_plants:
            # Check if all plants would be excluded
            df = _filter_master(regions, states, resources, qcas, pools)
            filtered_plants = sorted({p for p in df.get("plant_name", pd.Series(dtype=str)).dropna().astype(str).tolist() if p.strip()})
            
            if filtered_plants:
                excluded_list = excluded_plants if isinstance(excluded_plants, list) else [excluded_plants]
                remaining = [p for p in filtered_plants if p not in excluded_list]
                if not remaining:
                    # All plants excluded - disable button
                    return True
    
    disabled = not (has_regions and has_plants and has_dates)
    return disabled

# Enable stats Run button
@app.callback(
    Output("stats-run", "disabled"),
    Input("stats-region-dd", "value"),
    Input("stats-plant-dd", "value"),
    Input("stats-date-range", "start_date"),
    Input("stats-date-range", "end_date"),
    prevent_initial_call=False
)
def toggle_stats_run(regions, plants, start_date, end_date):
    has_regions = bool(regions)
    has_plants = bool(plants)
    has_dates = bool(start_date) and bool(end_date)
    return not (has_regions and has_plants and has_dates)

# Callback to load saved settings into form (without bands-table to avoid duplicate)
@app.callback(
    Output("err-mode", "value"),
    Output("x-pct", "value"),
    Input("nav-store", "data"),
    State("saved-settings-store", "data"),
    prevent_initial_call=False
)
def load_saved_settings(nav_data, saved_settings):
    """Load saved settings into form when navigating to settings page"""
    if nav_data == "settings" and saved_settings and isinstance(saved_settings, dict):
        # Load saved settings into form
        return (
            saved_settings.get("err_mode", "default"),
            saved_settings.get("x_pct", 50)
        )
    else:
        # Return defaults
        return "default", 50

# Callback to save settings
@app.callback(
    Output("settings-save-message", "children"),
    Output("saved-settings-store", "data"),
    Input("btn-save-settings", "n_clicks"),
    State("err-mode", "value"),
    State("x-pct", "value"),
    State("bands-table", "data"),
    State("zero-basis-guard", "value"),
    prevent_initial_call=True
)
def save_settings(n_clicks, err_mode, x_pct, bands_data, zero_basis_guard):
    """Save current settings to local storage"""
    if not n_clicks:
        raise PreventUpdate
    
    # Package settings
    settings = {
        "err_mode": err_mode,
        "x_pct": x_pct,
        "bands": bands_data,
        "zero_basis_guard": "on" in (zero_basis_guard or [])
    }
    
    # Success message
    message = dbc.Alert([
        html.Span("✓ ", style={"fontSize": "1.2rem", "marginRight": "8px"}),
        html.Strong("Settings saved successfully!"),
        html.Br(),
        html.Small("Your configuration will be used when running analysis.")
    ], color="success", dismissable=True, duration=4000)
    
    return message, settings

# Single callback to handle bands table data from saved settings
@app.callback(
    Output("bands-table", "data", allow_duplicate=True),
    Input("saved-settings-store", "data"),
    Input("nav-store", "data"),
    prevent_initial_call=True
)
def load_saved_bands(saved_settings, nav_data):
    """Load saved bands data when settings change or navigating to settings page"""
    if saved_settings and isinstance(saved_settings, dict):
        bands_data = saved_settings.get("bands", DEFAULT_BANDS.copy())
        print(f"DEBUG - Loading saved bands: {len(bands_data)} bands")
        return bands_data
    else:
        print(f"DEBUG - Using default bands: {len(DEFAULT_BANDS)} bands")
        return DEFAULT_BANDS.copy()

@app.callback(
    Output("x-pct-container", "style"),
    Output("x-pct-readout", "children"),
    Output("x-pct-label", "children"),
    Input("err-mode", "value"),
    Input("x-pct", "value"),
)
def toggle_xpct(err_mode, x):
    show = {"display": "block"} if err_mode == "dynamic" else {"display": "none"}
    readout = f"X = {x:.0f}%" if err_mode == "dynamic" else ""
    label = f"X% = {x:.0f}% (for Dynamic Error%)" if err_mode == "dynamic" else "X% (for Dynamic Error%)"
    return show, readout, label


@app.callback(
    Output("agg-ppa-numeric-container", "style"),
    Input("agg-ppa-mode", "value"),
)
def toggle_agg_numeric_ppa(mode):
    # agg-ppa-mode is multi-select; show numeric input if "numeric" is among selected modes
    if isinstance(mode, list):
        show_numeric = "numeric" in mode
    else:
        show_numeric = mode == "numeric"
    show = {"display": "block"} if show_numeric else {"display": "none"}
    return show

def generate_label(row):
    """Generate automatic label based on row values"""
    direction = row.get("direction", "")
    lower_pct = row.get("lower_pct", 0)
    upper_pct = row.get("upper_pct", 0)
    rate_type = row.get("rate_type", "")
    rate_value = row.get("rate_value", 0)
    
    # Format percentage ranges
    if upper_pct >= 1000:
        range_str = f"{direction} >{lower_pct}%"
    else:
        range_str = f"{direction} {lower_pct}-{upper_pct}%"
    
    # Generate description based on rate_type only (Apply To removed)
    if float(lower_pct) <= 0.0 and rate_value == 0:
        desc = "(Tolerance)"
    elif rate_type == "flat_per_kwh":
        desc = f"({rate_value} Rs/kWh)"
    elif rate_type == "ppa_fraction":
        desc = f"({rate_value*100:.0f}% of PPA)"
    elif rate_type == "scaled_excess":
        desc = "(scaled)"
    else:
        desc = f"({rate_type})"
    
    return f"{range_str} {desc}"

@app.callback(
    Output("bands-table", "data", allow_duplicate=True),
    Output("download-bands-json", "data"),
    Output("form-direction", "value"),
    Output("form-lower-pct", "value"),
    Output("form-upper-pct", "value"),
    Output("form-tolerance-cut-pct", "value"),
    Output("form-loss-zone", "value"),
    Output("form-rate-type", "value"),
    Output("form-rate-value", "value"),
    Output("form-excess-slope-per-pct", "value"),
    Input("reset-bands", "n_clicks"),
    Input("bands-table", "data_timestamp"),
    Input("upload-bands", "contents"),
    Input("save-bands", "n_clicks"),
    Input("add-from-form", "n_clicks"),
    State("bands-table", "data"),
    State("form-direction", "value"),
    State("form-lower-pct", "value"),
    State("form-upper-pct", "value"),
    State("form-tolerance-cut-pct", "value"),
    State("form-loss-zone", "value"),
    State("form-rate-type", "value"),
    State("form-rate-value", "value"),
    State("form-excess-slope-per-pct", "value"),
    prevent_initial_call=True
)
def manage_bands(reset_clicks, timestamp, uploaded, save_clicks, add_clicks, rows, 
                direction, lower_pct, upper_pct, tolerance_cut_pct, loss_zone, 
                rate_type, rate_value, excess_slope_per_pct):
    trig = ctx.triggered_id
    rows = rows or []
    
    # Default form reset values
    form_reset = ("UI", 0.0, 10.0, 10.0, False, "flat_per_kwh", 0.0, 0.0)
    
    # Handle reset
    if trig == "reset-bands":
        return DEFAULT_BANDS.copy(), dash.no_update, *form_reset
    
    # Handle upload
    if trig == "upload-bands" and uploaded:
        header, b64 = uploaded.split(",", 1)
        import base64, io
        payload = base64.b64decode(b64)
        try:
            js = json.loads(payload.decode("utf-8"))
            assert isinstance(js, list)
            return js, dash.no_update, *form_reset
        except Exception:
            return dash.no_update, dash.no_update, *form_reset
    
    # Handle save
    if trig == "save-bands" and save_clicks:
        content = json.dumps(rows or [], indent=2)
        return dash.no_update, dict(content=content, filename="bands_config_preset.json"), *form_reset
    
    # Handle add from form
    if trig == "add-from-form" and add_clicks:
        # Validate required fields
        if all([direction, lower_pct is not None, upper_pct is not None, tolerance_cut_pct is not None, 
                rate_type, rate_value is not None, excess_slope_per_pct is not None]):
            
            # Create new row
            new_row = {
                "direction": direction,
                "lower_pct": float(lower_pct),
                "upper_pct": float(upper_pct),
                "tolerance_cut_pct": float(tolerance_cut_pct),
                "rate_type": rate_type,
                "rate_value": float(rate_value),
                "excess_slope_per_pct": float(excess_slope_per_pct),
                "loss_zone": bool(loss_zone),
                "label": ""
            }
            
            # Generate label for new row
            new_row["label"] = generate_label(new_row)
            
            # Add to existing rows
            updated_rows = (rows or []) + [new_row]
            
            return updated_rows, dash.no_update, *form_reset
    
    # Auto-update labels when data changes
    if trig == "bands-table" and rows:
        for row in rows:
            if row.get("label") == "" or not row.get("label"):
                row["label"] = generate_label(row)
        return rows, dash.no_update, *form_reset
    
    return dash.no_update, dash.no_update, *form_reset


# Form label preview callback
@app.callback(
    Output("form-label-preview", "children"),
    Input("form-direction", "value"),
    Input("form-lower-pct", "value"),
    Input("form-upper-pct", "value"),
    Input("form-rate-type", "value"),
    Input("form-rate-value", "value"),
    prevent_initial_call=True
)
def preview_form_label(direction, lower_pct, upper_pct, rate_type, rate_value):
    if not all([direction, lower_pct is not None, upper_pct is not None, rate_type, rate_value is not None]):
        return "Fill in all fields to see preview..."
    
    # Generate label using the same logic as the table
    row = {
        "direction": direction,
        "lower_pct": float(lower_pct) if lower_pct else 0,
        "upper_pct": float(upper_pct) if upper_pct else 0,
        "rate_type": rate_type,
        "rate_value": float(rate_value) if rate_value else 0
    }
    return generate_label(row)


def _compute_pipeline(regions, plants, start_date, end_date, err_mode, x_pct, bands_rows, unpaid_oi_threshold=15.0, excluded_plants=None):
    if not regions or not plants:
        return {"error": "Please select at least one region and plant"}
    
    # Handle "Select All" case - resolve from DuckDB only for consistency
    if "SELECT_ALL" in plants:
        plants = get_plants_from_duckdb(regions)
        # Apply exclusion logic: Final Plants = Selected Plants - Excluded Plants
        if excluded_plants:
            excluded_list = excluded_plants if isinstance(excluded_plants, list) else [excluded_plants]
            plants = [p for p in plants if p not in excluded_list]
            if not plants:
                return {"error": "All plants have been excluded. Please select at least one plant."}
    
    # Look up each plant's region from the master Excel file
    # This ensures we query the correct region database for each plant
    master_df = _load_plant_master_df()
    plant_to_region = {}
    if not master_df.empty:
        # Create a mapping: plant_name (normalized) -> region
        for _, row in master_df.iterrows():
            plant_name = str(row.get("plant_name", "")).strip()
            region = str(row.get("region", "")).strip().upper()
            if plant_name and region:
                # Normalize plant name for matching (case-insensitive)
                plant_key = _norm_plant_name(plant_name)
                # Store both the normalized key and allow multiple regions per plant
                if plant_key not in plant_to_region:
                    plant_to_region[plant_key] = region
                # Also create reverse lookup for debugging
        print(f"DEBUG - Loaded {len(plant_to_region)} plant-to-region mappings from master Excel")
    else:
        print(f"DEBUG - Master Excel is empty, will search all regions")
    
    # Group plants by their actual region (from master Excel)
    plants_by_region = {}
    unmatched_plants = []
    
    for plant in plants:
        plant_norm = _norm_plant_name(plant)
        region = plant_to_region.get(plant_norm)
        
        if region:
            # Verify region is in the selected regions list (case-insensitive)
            region_upper = region.upper()
            selected_regions_upper = [r.upper() for r in regions]
            
            if region_upper in selected_regions_upper:
                if region_upper not in plants_by_region:
                    plants_by_region[region_upper] = []
                plants_by_region[region_upper].append(plant)
            else:
                # Plant's region doesn't match selected regions - try all selected regions
                unmatched_plants.append(plant)
        else:
            # Plant not found in master Excel - try all selected regions
            unmatched_plants.append(plant)
    
    # For unmatched plants, try querying ALL available regions (not just selected ones)
    # This handles cases where:
    # 1. Master Excel might be incomplete
    # 2. Plant might be in a different region than selected
    # 3. User might have selected wrong region
    if unmatched_plants:
        # Get all available regions from database files
        all_available_regions = []
        for region_name in ["NRPC", "SRPC", "WRPC"]:
            db_file = _duckdb_path(region_name.lower())
            if db_file.exists():
                all_available_regions.append(region_name)
                # Try to find the plant in this database
                try:
                    conn = duckdb.connect(str(db_file), read_only=True)
                    table_name = region_name.lower()
                    # Check if any of the unmatched plants exist in this database
                    for plant in unmatched_plants:
                        plant_key = _norm_plant_name(plant).replace("'", "''")
                        check_query = f"""
                            SELECT DISTINCT plant_name 
                            FROM {table_name} 
                            WHERE trim(regexp_replace(regexp_replace(upper(plant_name), '[^A-Z0-9\\- ]+', ' '), '\\s+', ' ')) = '{plant_key}'
                            LIMIT 1
                        """
                        result = conn.execute(check_query).fetchone()
                        if result:
                            print(f"DEBUG - Found '{plant}' in {region_name} database (original name: '{result[0]}')")
                            if region_name.upper() not in plants_by_region:
                                plants_by_region[region_name.upper()] = []
                            if plant not in plants_by_region[region_name.upper()]:
                                plants_by_region[region_name.upper()].append(plant)
                    conn.close()
                except Exception as check_err:
                    print(f"DEBUG - Error checking {db_file} for unmatched plants: {check_err}")
        
        # If no regions found in files, fall back to selected regions
        search_regions = all_available_regions if all_available_regions else regions
        
        # Add remaining unmatched plants to all search regions (fallback)
        still_unmatched = [p for p in unmatched_plants if not any(p in plants_by_region.get(r.upper(), []) for r in search_regions)]
        if still_unmatched:
            print(f"DEBUG - {len(still_unmatched)} plants still unmatched after database check. Adding to all search regions: {still_unmatched}")
            for region in search_regions:
                region_upper = region.upper()
                if region_upper not in plants_by_region:
                    plants_by_region[region_upper] = []
                # Only add if not already added
                for plant in still_unmatched:
                    if plant not in plants_by_region[region_upper]:
                        plants_by_region[region_upper].append(plant)
    
    # Load data from each region for its respective plants
    all_dfs = []
    found_plants = set()
    for region_upper, region_plants in plants_by_region.items():
        if region_plants:
            # Remove duplicates while preserving order
            unique_plants = list(dict.fromkeys(region_plants))
            print(f"DEBUG - Querying {region_upper} for plants: {unique_plants[:3]}... ({len(unique_plants)} total)")
            region_df = load_region_data(region_upper, str(start_date), str(end_date), unique_plants)
            if not region_df.empty:
                all_dfs.append(region_df)
                # Track which plants were found
                if "plant_name" in region_df.columns:
                    found_plants.update(region_df["plant_name"].str.strip().str.upper().unique())
                elif "Plant" in region_df.columns:
                    found_plants.update(region_df["Plant"].str.strip().str.upper().unique())
                print(f"DEBUG - Found {len(region_df)} rows from {region_upper}")
            else:
                print(f"DEBUG - No data returned from {region_upper} for {len(unique_plants)} plants")
    
    if not all_dfs:
        # Comprehensive diagnostic: Check all regions for the requested plants
        diagnostic_info = []
        requested_plants_str = ", ".join(plants[:5])  # Show first 5 plants
        if len(plants) > 5:
            requested_plants_str += f", ... ({len(plants)} total)"
        
        # Check each requested plant in all available databases
        for plant in plants[:10]:  # Limit to first 10 for performance
            plant_norm = _norm_plant_name(plant)
            plant_found_in = []
            
            for region_name in ["NRPC", "SRPC", "WRPC"]:
                db_file = _duckdb_path(region_name.lower())
                if db_file.exists():
                    try:
                        conn = duckdb.connect(str(db_file), read_only=True)
                        table_name = region_name.lower()
                        
                        # Check if plant exists
                        plant_norm_escaped = plant_norm.replace("'", "''")
                        check_plant = (
                            f"SELECT DISTINCT plant_name FROM {table_name} "
                            f"WHERE trim(regexp_replace(regexp_replace(upper(plant_name), '[^A-Z0-9\\- ]+', ' '), '\\s+', ' ')) = '{plant_norm_escaped}' "
                            f"LIMIT 1"
                        )
                        plant_result = conn.execute(check_plant).fetchone()
                        
                        if plant_result:
                            # Check if data exists for the date range
                            check_date = f"""
                                SELECT COUNT(*) as cnt, MIN(date) as min_d, MAX(date) as max_d 
                                FROM {table_name} 
                                WHERE trim(regexp_replace(regexp_replace(upper(plant_name), '[^A-Z0-9\\- ]+', ' '), '\\s+', ' ')) = '{plant_norm_escaped}'
                                  AND date >= '{start_date}' AND date <= '{end_date}'
                            """
                            date_result = conn.execute(check_date).fetchone()
                            if date_result and date_result[0] > 0:
                                plant_found_in.append(f"{region_name} (✓ has data for {start_date} to {end_date})")
                            else:
                                # Check what date range is available
                                date_range_query = f"""
                                    SELECT MIN(date) as min_d, MAX(date) as max_d 
                                    FROM {table_name} 
                                    WHERE trim(regexp_replace(regexp_replace(upper(plant_name), '[^A-Z0-9\\- ]+', ' '), '\\s+', ' ')) = '{plant_norm_escaped}'
                                """
                                date_range = conn.execute(date_range_query).fetchone()
                                if date_range and date_range[0]:
                                    plant_found_in.append(f"{region_name} (✗ no data for {start_date}-{end_date}, but has data from {date_range[0]} to {date_range[1]})")
                                else:
                                    plant_found_in.append(f"{region_name} (✗ plant exists but no data)")
                        conn.close()
                    except Exception as diag_err:
                        print(f"DEBUG - Diagnostic error for {region_name}: {diag_err}")
            
            if plant_found_in:
                diagnostic_info.append(f"  '{plant}': Found in {', '.join(plant_found_in)}")
            else:
                diagnostic_info.append(f"  '{plant}': NOT FOUND in any database")
        
        # Get sample plant names from databases
        available_plants_info = []
        for region in regions:
            region_lower = region.lower()
            db_file = _duckdb_path(region_lower)
            if db_file.exists():
                try:
                    conn = duckdb.connect(str(db_file), read_only=True)
                    sample_query = f"SELECT DISTINCT plant_name FROM {region_lower} LIMIT 10"
                    sample_plants = conn.execute(sample_query).fetchall()
                    sample_list = [p[0] for p in sample_plants if p[0]]
                    conn.close()
                    if sample_list:
                        available_plants_info.append(f"{region}: {', '.join(sample_list[:5])}...")
                except Exception as e:
                    print(f"DEBUG - Error checking {db_file}: {e}")
        
        available_info = "\n".join(available_plants_info) if available_plants_info else "Unable to query database"
        diagnostic_text = "\n".join(diagnostic_info) if diagnostic_info else "No diagnostic information available"
        
        error_msg = (
            f"No data found for the selected plants in date range {start_date} to {end_date}.\n\n"
            f"Requested plants: {requested_plants_str}\n"
            f"Selected regions: {', '.join(regions)}\n\n"
            f"Diagnostic Information:\n{diagnostic_text}\n\n"
            f"Sample plants in selected region(s):\n{available_info}\n\n"
            f"Please verify:\n"
            f"1. Plant names match exactly (case-insensitive) with database\n"
            f"2. Date range contains data for these plants\n"
            f"3. Plants belong to the selected region(s)\n"
            f"4. Database files (nrpc.duckdb, srpc.duckdb, wrpc.duckdb) are accessible"
        )
        return {"error": error_msg}
    
    # Combine data from all regions
    df = pd.concat(all_dfs, ignore_index=True) if len(all_dfs) > 1 else all_dfs[0]
    
    # The SQL query in load_data_from_duckdb already filters by plant names,
    # so we don't need to filter again here. However, if unmatched plants were
    # added to multiple regions, we might have duplicates. Let's ensure we only
    # keep data for the requested plants (case-insensitive).
    if not df.empty and plants:
        # Normalize plant names for comparison
        requested_plants_norm = {_norm_plant_name(p) for p in plants}
        # Check both plant_name and Plant columns (case-insensitive)
        if "plant_name" in df.columns:
            df["_plant_norm"] = df["plant_name"].astype(str).str.strip().str.upper()
            df = df[df["_plant_norm"].isin(requested_plants_norm)]
            df = df.drop(columns=["_plant_norm"], errors="ignore")
        elif "Plant" in df.columns:
            df["_plant_norm"] = df["Plant"].astype(str).str.strip().str.upper()
            df = df[df["_plant_norm"].isin(requested_plants_norm)]
            df = df.drop(columns=["_plant_norm"], errors="ignore")
    
    # Debug: Log what was found vs requested
    if not df.empty:
        found_plant_names = set()
        if "plant_name" in df.columns:
            found_plant_names = set(df["plant_name"].str.strip().str.upper().unique())
        elif "Plant" in df.columns:
            found_plant_names = set(df["Plant"].str.strip().str.upper().unique())
        requested_upper = {_norm_plant_name(p) for p in plants}
        print(f"DEBUG - Requested plants: {requested_upper}")
        print(f"DEBUG - Found plants: {found_plant_names}")
        print(f"DEBUG - Total rows loaded: {len(df)}")

    # Validate required columns exist before processing
    required_cols = ["AvC_MW", "Scheduled_MW", "Actual_MW", "PPA"]
    missing_cols = [col for col in required_cols if col not in df.columns]
    if missing_cols:
        error_msg = f"Missing required columns in data: {', '.join(missing_cols)}. Available columns: {', '.join(df.columns.tolist())}"
        print(f"ERROR - {error_msg}")
        return {"error": error_msg}
    
    # Data quality computed on the raw slot series (before any DSM band calcs)
    try:
        data_quality = compute_data_quality_stats(df)
    except Exception as e:
        print(f"ERROR - Failed to compute data quality stats: {e}")
        import traceback
        traceback.print_exc()
        return {"error": f"Failed to compute data quality: {str(e)}"}

    try:
        df["error_pct"] = compute_error_pct(df, err_mode, float(x_pct))
        df["basis_MW"] = compute_basis_mw(df, err_mode, float(x_pct))
    except Exception as e:
        print(f"ERROR - Failed to compute error metrics: {e}")
        import traceback
        traceback.print_exc()
        return {"error": f"Failed to compute error metrics: {str(e)}"}

    bands_df = _normalize_bands_df(pd.DataFrame(bands_rows or []))
    # Only core fields are truly required after normalization
    core_required = {"direction","lower_pct","upper_pct","rate_type","rate_value","excess_slope_per_pct"}
    if not core_required.issubset(set(bands_df.columns)):
        missing = core_required - set(bands_df.columns)
        raise ValueError(f"Bands config missing required columns: {', '.join(sorted(missing))}")
    bands_rows = bands_df.to_dict("records")

    df = apply_bands(df, bands_rows, unpaid_oi_threshold)
    summary = summarize(df, selected_plants=plants, bands_rows=bands_rows, err_mode=err_mode, x_pct=float(x_pct), start_date=start_date, end_date=end_date)
    summary["data_quality"] = data_quality
    return summary


def _compute_pipeline_aggregated(
    regions,
    plants,
    start_date,
    end_date,
    err_mode,
    x_pct,
    bands_rows,
    ppa_mode: str,
    ppa_value: float | None,
    unpaid_oi_threshold: float = 15.0,
):
    """Run the same analysis engine on an aggregated multi-plant time series.

    This function only changes the input dataset by first aggregating AvC/Scheduled/Actual
    across the selected plants per time block, and then computing PPA as per the
    requested aggregation mode. All downstream calculations reuse the existing
    analysis logic unchanged.
    """
    if not regions or not plants:
        return {"error": "Please select at least one region and plant"}

    # Resolve "Select All" - but plants should already be resolved by caller (with exclusions applied)
    # Handle both cases: list of plants or SELECT_ALL string (for backward compatibility)
    if isinstance(plants, list):
        if "SELECT_ALL" in plants:
            plants = get_plants_from_duckdb(regions)
    elif plants == "SELECT_ALL":
        plants = get_plants_from_duckdb(regions)
    elif not isinstance(plants, list):
        plants = [plants] if plants else []

    if not plants:
        return {"error": "No plants resolved for aggregation"}

    # Load data from ALL selected regions and combine
    all_dfs = []
    for region in regions:
        region_df = load_region_data(region, str(start_date), str(end_date), plants)
        if not region_df.empty:
            all_dfs.append(region_df)
    
    if not all_dfs:
        return {"error": "No data found for the selected filters"}
    
    # Combine data from all regions
    df_raw = pd.concat(all_dfs, ignore_index=True) if len(all_dfs) > 1 else all_dfs[0]
    
    # Create comma-separated region string for display
    regions_str = ",".join(sorted(regions))
    
    # Ensure core numeric fields are usable before aggregation
    # Clip negative values to 0 (negative power is invalid)
    df_raw["AvC_MW"] = pd.to_numeric(df_raw.get("AvC_MW"), errors="coerce").fillna(0.0).clip(lower=0)
    df_raw["Scheduled_MW"] = pd.to_numeric(df_raw.get("Scheduled_MW"), errors="coerce").fillna(0.0).clip(lower=0)
    df_raw["Actual_MW"] = pd.to_numeric(df_raw.get("Actual_MW"), errors="coerce").fillna(0.0).clip(lower=0)
    df_raw["PPA"] = pd.to_numeric(df_raw.get("PPA"), errors="coerce").fillna(0.0).clip(lower=0)

    # Data quality computed on the raw (pre-aggregation) slot series
    data_quality = compute_data_quality_stats(df_raw)

    # Group per time block/slot (aggregate across all regions)
    # Note: region column will be updated to show comma-separated regions after grouping
    group_cols = ["date", "time_block", "from_time", "to_time"]
    for col in group_cols:
        if col not in df_raw.columns:
            df_raw[col] = None

    agg_df = (
        df_raw.groupby(group_cols, as_index=False)
        .agg({
            "AvC_MW": "sum",
            "Scheduled_MW": "sum",
            "Actual_MW": "sum",
        })
    )

    # PPA aggregation per slot (across all regions)
    mode_norm = (ppa_mode or "mean").strip().lower()
    if mode_norm == "numeric" and ppa_value is not None:
        agg_df["PPA"] = float(ppa_value)
    elif mode_norm == "weighted":
        # Weighted PPA per time-block:
        #   Weighted_PPA[t] = sum(PPA * AvC_MW) / sum(AvC_MW)
        # Uses the exact tags and does not change any downstream formulas.
        df_raw["_ppa_x_avc"] = df_raw["PPA"] * df_raw["AvC_MW"]
        num = df_raw.groupby(group_cols, as_index=False)["_ppa_x_avc"].sum().rename(columns={"_ppa_x_avc": "_ppa_num"})
        agg_df = agg_df.merge(num, on=group_cols, how="left")
        agg_df["_ppa_num"] = pd.to_numeric(agg_df.get("_ppa_num"), errors="coerce").fillna(0.0)
        denom = pd.to_numeric(agg_df.get("AvC_MW"), errors="coerce").fillna(0.0)
        agg_df["PPA"] = np.where(denom > 0, agg_df["_ppa_num"] / denom, 0.0)
        agg_df = agg_df.drop(columns=["_ppa_num"], errors="ignore")
        df_raw = df_raw.drop(columns=["_ppa_x_avc"], errors="ignore")
    else:
        def _agg_ppa(series: pd.Series) -> float:
            vals = pd.to_numeric(series, errors="coerce").dropna()
            if vals.empty:
                return 0.0
            if mode_norm == "median":
                return float(vals.median())
            if mode_norm == "mode":
                # Use existing safe_mode helper to respect mode semantics
                return float(safe_mode(vals.tolist()))
            # Default to mean
            return float(vals.mean())

        ppa_by_slot = (
            df_raw.groupby(group_cols)["PPA"]
            .apply(_agg_ppa)
            .reset_index(name="PPA")
        )
        agg_df = agg_df.merge(ppa_by_slot, on=group_cols, how="left")
        agg_df["PPA"] = pd.to_numeric(agg_df["PPA"], errors="coerce").fillna(0.0)

    # Add region column with comma-separated region names
    agg_df["region"] = regions_str

    # Synthesize a single aggregated plant identity for downstream logic
    agg_df["Plant"] = "AGGREGATED"
    agg_df["plant_name"] = "AGGREGATED"

    # Recreate date_time for completeness (matches load_region_data helper)
    try:
        agg_df["date_time"] = pd.to_datetime(agg_df["date"]) + pd.to_timedelta(
            (agg_df["time_block"] - 1) * 15, unit="m"
        )
    except Exception:
        pass

    # Error% and basis computed on aggregated MW values using existing helpers
    agg_df["error_pct"] = compute_error_pct(agg_df, err_mode, float(x_pct))
    agg_df["basis_MW"] = compute_basis_mw(agg_df, err_mode, float(x_pct))

    # Bands normalization exactly as in the main pipeline
    bands_df = _normalize_bands_df(pd.DataFrame(bands_rows or []))
    core_required = {"direction", "lower_pct", "upper_pct", "rate_type", "rate_value", "excess_slope_per_pct"}
    if not core_required.issubset(set(bands_df.columns)):
        missing = core_required - set(bands_df.columns)
        raise ValueError(f"Bands config missing required columns: {', '.join(sorted(missing))}")
    bands_rows_norm = bands_df.to_dict("records")

    # Apply bands and summarize using the existing engine; do NOT pass per-plant selection
    agg_df = apply_bands(agg_df, bands_rows_norm, unpaid_oi_threshold)
    summary = summarize(agg_df, selected_plants=None, bands_rows=bands_rows_norm, err_mode=err_mode, x_pct=float(x_pct), start_date=start_date, end_date=end_date)
    summary["data_quality"] = data_quality
    summary["aggregated_plants"] = list(plants)
    summary["aggregated_region"] = regions_str  # Use comma-separated regions string
    return summary


# --- Preset management callbacks ---
@app.callback(
    Output("preset-save-message", "children"),
    Output("presets-store", "data"),
    Input("btn-save-preset", "n_clicks"),
    State("preset-name", "value"),
    State("err-mode", "value"),
    State("x-pct", "value"),
    State("bands-table", "data"),
    State("zero-basis-guard", "value"),
    State("presets-store", "data"),
    prevent_initial_call=True
)
def save_preset(n_clicks, name, err_mode, x_pct, bands_data, zero_basis_guard, presets):
    if not n_clicks:
        raise PreventUpdate
    name = (name or "").strip()
    if not name:
        return dbc.Alert("Please enter a preset name.", color="warning"), dash.no_update
    preset = {
        "name": name,
        "settings": {
            "err_mode": err_mode,
            "x_pct": x_pct,
            "bands": bands_data,
            "zero_basis_guard": "on" in (zero_basis_guard or [])
        }
    }
    presets = list(presets or [])
    names = [p.get("name") for p in presets]
    if name in names:
        presets[names.index(name)] = preset
        msg = f"Preset '{name}' updated."
    else:
        presets.append(preset)
        msg = f"Preset '{name}' saved."
    print(f"DEBUG - {msg} (total presets: {len(presets)})")
    return dbc.Alert(msg, color="success"), presets


@app.callback(
    Output("analysis-preset-select", "data"),
    Input("presets-store", "data"),
)
def load_preset_options(presets):
    presets = presets or []
    return [{"label": p.get("name", "Unnamed"), "value": p.get("name", "Unnamed")} for p in presets]


@app.callback(
    Output("global-setting-row", "style"),
    Output("plant-setting-mapping-container", "style"),
    Output("analysis-preset-select", "disabled"),
    Input("custom-setting-mode", "value"),
)
def toggle_mapping_ui(mode):
    """Toggle visibility of global setting dropdown and plant-wise mapping table based on mode."""
    if mode == "PLANT_WISE":
        return {"display": "none"}, {"display": "block"}, True
    return {"display": "block"}, {"display": "none"}, False


@app.callback(
    Output("custom-setting-mode", "value"),
    Output("setting-assignments-store", "data", allow_duplicate=True),
    Input("analysis-preset-select", "value"),
    Input("plant-dd", "value"),
    State("custom-setting-mode", "value"),
    prevent_initial_call="initial_duplicate",
)
def auto_switch_to_assignment_ui(selected_presets, plants_value, mode):
    """If user selects exactly ONE custom setting while multiple plants are selected,
    auto-switch to PLANT_WISE assignment UI and prefill row 1 as 'apply to all remaining plants'.

    This matches the UX: pick a setting -> choose all/specific plants -> add another setting for remaining.
    """
    plant_list = _normalize_multi(plants_value)
    plant_list = [p for p in plant_list if p and p != "SELECT_ALL"]
    has_multi_plants = len(plant_list) > 1

    # Auto-switch when presets are selected and multiple plants are selected
    if not has_multi_plants or not selected_presets:
        # if user reduced to 0/1 plants or cleared preset, keep current mode
        raise PreventUpdate

    presets_list = list(selected_presets) if isinstance(selected_presets, list) else [selected_presets]
    presets_list = [p for p in presets_list if p]
    if not presets_list:
        raise PreventUpdate

    # If multiple presets are selected, create multiple assignment rows so user can allocate plants.
    # (We do NOT auto-apply "all remaining" in this case, since it would consume all plants.)
    rows = []
    if len(presets_list) == 1:
        rows = [{"row": 0, "preset": presets_list[0], "plants": [], "apply_remaining": True}]
    else:
        rows = [{"row": i, "preset": p, "plants": [], "apply_remaining": False} for i, p in enumerate(presets_list)]

    store = {
        "next_id": len(rows),
        "rows": rows,
    }
    return "PLANT_WISE", store


@app.callback(
    Output("setting-assignments-store", "data"),
    Input("btn-add-setting-row", "n_clicks"),
    Input("btn-reset-setting-rows", "n_clicks"),
    Input({"type": "setting-row-remove", "row": ALL}, "n_clicks"),
    Input({"type": "setting-row-preset", "row": ALL}, "value"),
    Input({"type": "setting-row-scope", "row": ALL}, "value"),
    Input({"type": "setting-row-plants", "row": ALL}, "value"),
    State({"type": "setting-row-preset", "row": ALL}, "id"),
    State({"type": "setting-row-scope", "row": ALL}, "id"),
    State({"type": "setting-row-plants", "row": ALL}, "id"),
    State({"type": "setting-row-remove", "row": ALL}, "id"),
    State("setting-assignments-store", "data"),
    prevent_initial_call=True,
)
def update_setting_assignments(
    n_add,
    n_reset,
    n_remove_clicks,
    preset_values,
    scope_values,
    plants_values,
    preset_ids,
    scope_ids,
    plants_ids,
    remove_ids,
    store,
):
    """Maintain settings→plants assignment rows in a store.

    Each row has:
      - row (int): stable row id
      - preset (str|None): preset name
      - apply_remaining (bool): True => applies to all remaining plants
      - plants (list[str]): selected plants (only if apply_remaining=False)
    """
    store = store or {"next_id": 1, "rows": [{"row": 0, "preset": None, "plants": [], "apply_remaining": True}]}
    store.setdefault("next_id", 1)
    store.setdefault("rows", [{"row": 0, "preset": None, "plants": [], "apply_remaining": True}])

    trig = ctx.triggered_id

    if trig == "btn-reset-setting-rows":
        return {"next_id": 1, "rows": [{"row": 0, "preset": None, "plants": [], "apply_remaining": True}]}

    if trig == "btn-add-setting-row":
        rid = int(store.get("next_id") or 1)
        store["next_id"] = rid + 1
        rows = list(store.get("rows") or [])
        rows.append({"row": rid, "preset": None, "plants": [], "apply_remaining": False})
        store["rows"] = rows
        return store

    if isinstance(trig, dict) and trig.get("type") == "setting-row-remove":
        rid = int(trig.get("row", -1))
        rows = [r for r in (store.get("rows") or []) if int(r.get("row", -2)) != rid]
        if not rows:
            rows = [{"row": 0, "preset": None, "plants": [], "apply_remaining": True}]
            store["next_id"] = 1
        store["rows"] = rows
        return store

    # Otherwise: sync row fields from the current component values
    preset_map: dict[int, str | None] = {}
    scope_map: dict[int, str | None] = {}
    plants_map: dict[int, list[str]] = {}

    for pid, val in zip(preset_ids or [], preset_values or []):
        rid = int((pid or {}).get("row", -1))
        preset_map[rid] = val

    for sid, val in zip(scope_ids or [], scope_values or []):
        rid = int((sid or {}).get("row", -1))
        scope_map[rid] = val

    for plid, val in zip(plants_ids or [], plants_values or []):
        rid = int((plid or {}).get("row", -1))
        if val is None:
            plants_map[rid] = []
        elif isinstance(val, list):
            plants_map[rid] = [str(x) for x in val if x]
        else:
            plants_map[rid] = [str(val)]

    updated_rows = []
    for r in (store.get("rows") or []):
        rid = int(r.get("row", 0))
        preset = preset_map.get(rid, r.get("preset"))
        scope = scope_map.get(rid, "REMAINING" if r.get("apply_remaining") else "SELECTED")
        apply_remaining = (scope == "REMAINING")
        plants_sel = [] if apply_remaining else plants_map.get(rid, r.get("plants") or [])
        updated_rows.append({"row": rid, "preset": preset, "plants": plants_sel, "apply_remaining": apply_remaining})

    store["rows"] = updated_rows
    return store


@app.callback(
    Output("plant-setting-mapping-container", "children"),
    Input("custom-setting-mode", "value"),
    Input("setting-assignments-store", "data"),
    Input("plant-dd", "value"),
    Input("presets-store", "data"),
    State("region-dd", "value"),
    State("state-dd", "value"),
    State("resource-type-dd", "value"),
    State("qca-dd", "value"),
    State("pooling-dd", "value"),
    State("exclude-plant-dd", "value"),
)
def build_mapping_table(mode, assignments_store, plants, presets_store, regions, states, resources, qcas, pools, excluded_plants):
    """Build the settings→plants assignment UI dynamically (PLANT_WISE mode)."""
    if mode != "PLANT_WISE":
        return dash.no_update
    if not plants:
        return html.Div("Select plants first to assign custom settings.", className="text-muted")
    
    # Resolve actual plant list (handle SELECT_ALL)
    plant_list = _normalize_multi(plants)
    is_select_all = "SELECT_ALL" in plant_list
    
    if is_select_all:
        # Get all plants from master data based on filters
        dfm = _filter_master(regions, states, resources, qcas, pools)
        all_plants = sorted({p for p in dfm.get("plant_name", pd.Series(dtype=str)).dropna().astype(str).tolist() if p.strip()})
        excluded_list = excluded_plants if isinstance(excluded_plants, list) else ([excluded_plants] if excluded_plants else [])
        excluded_list = [str(x) for x in excluded_list if x]
        resolved_plants = [p for p in all_plants if p not in excluded_list]
    else:
        resolved_plants = [str(p) for p in plant_list if p and str(p) != "SELECT_ALL"]
    
    if not resolved_plants:
        return html.Div("No plants selected.", className="text-muted")
    
    # Get available preset options
    presets = presets_store or []
    preset_options = [{"label": p.get("name", "Unnamed"), "value": p.get("name", "Unnamed")} for p in presets]

    store = assignments_store or {}
    rows_state = list(store.get("rows") or [])
    if not rows_state:
        rows_state = [{"row": 0, "preset": None, "plants": [], "apply_remaining": True}]

    # Build assignment rows, showing only remaining plants per row
    assigned: set[str] = set()
    ui_rows = []
    for r in rows_state:
        rid = int(r.get("row", 0))
        preset_val = r.get("preset")
        apply_remaining = bool(r.get("apply_remaining", False))

        remaining = [p for p in resolved_plants if p not in assigned]

        # Plants value should always be subset of remaining
        plants_val = [p for p in (r.get("plants") or []) if p in remaining]

        scope_value = "REMAINING" if apply_remaining else "SELECTED"

        # If apply_remaining, we will mark all remaining as assigned for next rows (preview only)
        preview_selected = remaining if apply_remaining else plants_val
        for p in preview_selected:
            assigned.add(p)

        ui_rows.append(
            dbc.Card(
                dbc.CardBody(
                    [
                        dbc.Row(
                            [
                                dbc.Col(
                                    [
                                        html.Label("Custom Setting", style={"fontWeight": 600, "fontSize": "0.9rem"}),
                                        dcc.Dropdown(
                                            id={"type": "setting-row-preset", "row": rid},
                                            options=preset_options,
                                            value=preset_val,
                                            placeholder="Select custom setting",
                                            clearable=False,
                                        ),
                                    ],
                                    md=4,
                                ),
                                dbc.Col(
                                    [
                                        html.Label("Apply To", style={"fontWeight": 600, "fontSize": "0.9rem"}),
                                        dcc.RadioItems(
                                            id={"type": "setting-row-scope", "row": rid},
                                            options=[
                                                {"label": f"All remaining plants ({len(remaining)})", "value": "REMAINING"},
                                                {"label": "Choose specific plants", "value": "SELECTED"},
                                            ],
                                            value=scope_value,
                                            inline=True,
                                        ),
                                    ],
                                    md=5,
                                ),
                                dbc.Col(
                                    [
                                        html.Label(" ", style={"display": "block"}),
                                        dbc.Button(
                                            "Remove",
                                            id={"type": "setting-row-remove", "row": rid},
                                            color="secondary",
                                            outline=True,
                                            size="sm",
                                        ),
                                    ],
                                    md=3,
                                    className="text-end",
                                ),
                            ],
                            className="g-2",
                        ),
                        html.Div(style={"height": "8px"}),
                        html.Div(
                            dmc.MultiSelect(
                                id={"type": "setting-row-plants", "row": rid},
                                data=[{"label": p, "value": p} for p in remaining],
                                value=plants_val,
                                searchable=True,
                                clearable=True,
                                comboboxProps={"keepOpened": True},
                                placeholder="Select plants for this custom setting...",
                                disabled=apply_remaining,
                                maxDropdownHeight=260,
                            )
                        ),
                        html.Div(
                            f"Will apply to: {', '.join(preview_selected) if preview_selected else '(none)'}",
                            className="text-muted",
                            style={"marginTop": "6px", "fontSize": "0.85rem"},
                        ),
                    ]
                ),
                className="mb-2 shadow-sm",
            )
        )

    # Remaining plants after all rows (preview)
    remaining_final = [p for p in resolved_plants if p not in assigned]

    return dbc.Card(
        [
            dbc.CardHeader(
                dbc.Row(
                    [
                        dbc.Col(html.Strong("Assign Custom Settings to Plants"), md=6),
                        dbc.Col(
                            html.Div(
                                [
                                    dbc.Button("Add another custom setting", id="btn-add-setting-row", color="primary", size="sm"),
                                    dbc.Button("Reset", id="btn-reset-setting-rows", color="secondary", outline=True, size="sm", className="ms-2"),
                                ],
                                className="text-end",
                            ),
                            md=6,
                        ),
                    ],
                    align="center",
                ),
                style={"backgroundColor": "#f8f9fa"},
            ),
            dbc.CardBody(
                ui_rows
                + (
                    [
                        dbc.Alert(
                            f"Unassigned plants: {', '.join(remaining_final)}",
                            color="warning",
                            className="mt-2",
                        )
                    ]
                    if remaining_final
                    else [
                        dbc.Alert("All selected plants are assigned to a custom setting.", color="success", className="mt-2")
                    ]
                )
            ),
        ],
        className="mb-2",
    )


@app.callback(
    Output("agg-analysis-preset-select", "data"),
    Input("presets-store", "data"),
)
def load_agg_preset_options(presets):
    presets = presets or []
    return [{"label": p.get("name", "Unnamed"), "value": p.get("name", "Unnamed")} for p in presets]


@app.callback(
    Output("presets-store", "data", allow_duplicate=True),
    Input("btn-delete-preset", "n_clicks"),
    State("analysis-preset-select", "value"),
    State("presets-store", "data"),
    prevent_initial_call=True
)
def delete_presets(n_clicks, selected_names, presets):
    if not n_clicks or not selected_names:
        raise PreventUpdate
    presets = list(presets or [])
    keep = [p for p in presets if p.get("name") not in (selected_names or [])]
    print(f"DEBUG - Deleted presets: {set(selected_names) - {p.get('name') for p in keep}}; remaining: {len(keep)}")
    return keep

@app.callback(
    Output("results-store", "data"),
    Output("results-section", "style"),
    Output("progress-container", "style"),
    Input("plot-now", "n_clicks"),
    State("region-dd", "value"),
    State("state-dd", "value"),
    State("resource-type-dd", "value"),
    State("qca-dd", "value"),
    State("pooling-dd", "value"),
    State("plant-dd", "value"),
    State("exclude-plant-dd", "value"),
    State("date-range", "start_date"),
    State("date-range", "end_date"),
    State("err-mode", "value"),
    State("x-pct", "value"),
    State("bands-table", "data"),
    State("saved-settings-store", "data"),
    State("analysis-preset-select", "value"),
    State("presets-store", "data"),
    State("custom-setting-mode", "value"),
    State("setting-assignments-store", "data"),
    prevent_initial_call=True
)
def compute_on_click(n, regions, states, resources, qcas, pools, plants, excluded_plants, start_date, end_date, err_mode, x_pct, bands_rows, saved_settings, selected_preset_names, presets_store, mode, assignments_store):
    if not n:
        raise PreventUpdate
    
    try:
        # Resolve "Select All" using the Excel master (single source of truth)
        resolved_plants = plants
        plant_list = _normalize_multi(plants)
        is_select_all = "SELECT_ALL" in plant_list
        if is_select_all:
            dfm = _filter_master(regions, states, resources, qcas, pools)
            all_plants = sorted({p for p in dfm.get("plant_name", pd.Series(dtype=str)).dropna().astype(str).tolist() if p.strip()})
            excluded_list = excluded_plants if isinstance(excluded_plants, list) else ([excluded_plants] if excluded_plants else [])
            excluded_list = [str(x) for x in excluded_list if x]
            remaining = [p for p in all_plants if p not in excluded_list]
            if not remaining:
                return ({"error": "All plants have been excluded. Please select at least one plant."}, {"display": "block"}, {"display": "none"})
            resolved_plants = remaining
            excluded_plants = None  # already applied
        else:
            resolved_plants = [str(p) for p in plant_list if p and str(p) != "SELECT_ALL"]

        # Determine mode (default to GLOBAL for backward compatibility)
        application_mode = mode if mode else "GLOBAL"
        
        name_to_settings = {p["name"]: p.get("settings", {}) for p in (presets_store or []) if isinstance(p, dict) and "name" in p}

        # Helper to run once with a given settings dict
        def run_once(with_settings: dict, target_plants: list):
            final_err_mode = with_settings.get("err_mode", err_mode)
            final_x_pct = with_settings.get("x_pct", x_pct)
            saved_bands = with_settings.get("bands", None)
            final_bands = saved_bands if (saved_bands and len(saved_bands) > 0) else (bands_rows or DEFAULT_BANDS.copy())
            unpaid_threshold = float(with_settings.get("unpaid_oi_threshold", 15.0))
            print(f"DEBUG - Analysis using: err_mode={final_err_mode}, x_pct={final_x_pct}, bands_count={len(final_bands)}")
            print(f"DEBUG - Bands: {[b.get('label','no-label') for b in final_bands[:3]]}...")
            res = _compute_pipeline(regions, target_plants, start_date, end_date, final_err_mode, final_x_pct, final_bands, unpaid_threshold, excluded_plants)
            return res, final_err_mode, final_x_pct, final_bands

        multi = []
        all_detail_dfs = []  # Initialize for PLANT_WISE mode
        
        if application_mode == "PLANT_WISE":
            # PLANT_WISE mode: settings→plants assignment rows
            store = assignments_store or {}
            rows_state = list(store.get("rows") or [])
            if not rows_state:
                return ({"error": "Please assign custom settings to plants (PLANT_WISE mode)."}, {"display": "block"}, {"display": "none"})

            assigned: set[str] = set()
            for r in rows_state:
                setting_name = r.get("preset")
                if not setting_name:
                    return ({"error": "Please select a Custom Setting for each assignment row."}, {"display": "block"}, {"display": "none"})
                st = name_to_settings.get(setting_name)
                if not st:
                    return ({"error": f"Preset '{setting_name}' not found. Please reselect it."}, {"display": "block"}, {"display": "none"})

                apply_remaining = bool(r.get("apply_remaining", False))
                remaining = [p for p in resolved_plants if p not in assigned]
                if apply_remaining:
                    target_plants = remaining
                else:
                    raw_plants = r.get("plants") or []
                    target_plants = [p for p in raw_plants if p in remaining]

                if not target_plants:
                    return ({"error": f"No plants selected for Custom Setting '{setting_name}'."}, {"display": "block"}, {"display": "none"})

                # Mark assigned before running (prevents overlaps across later rows)
                for p in target_plants:
                    assigned.add(p)

                res, fm, fx, fb = run_once(st, target_plants)
                if "error" in res:
                    return (res, {"display": "block"}, {"display": "none"})

                df_detail = res.get("df", pd.DataFrame())
                if isinstance(df_detail, pd.DataFrame) and not df_detail.empty:
                    df_detail = df_detail.copy()
                    df_detail["Custom Setting"] = setting_name
                    all_detail_dfs.append(df_detail)

                ps = res.get("plant_summary", pd.DataFrame())
                if isinstance(ps, pd.DataFrame) and not ps.empty:
                    ps = ps.copy()
                    ps["Custom Setting"] = setting_name
                    res["plant_summary"] = ps

                res["_preset_name"] = setting_name
                res["_final_err_mode"] = fm
                res["_final_x_pct"] = fx
                res["_final_bands"] = fb
                res["_plants"] = target_plants
                multi.append(res)

            missing = [p for p in resolved_plants if p not in assigned]
            if missing:
                return ({"error": f"Unassigned plants: {', '.join(missing)}"}, {"display": "block"}, {"display": "none"})
        elif selected_preset_names:
            # GLOBAL mode with multiple presets: Each preset applies to all plants
            print(f"DEBUG - Running multi-preset analysis for: {selected_preset_names}")
            for nm in selected_preset_names:
                st = name_to_settings.get(nm)
                if not st:
                    print(f"DEBUG - Preset '{nm}' not found in store; skipping")
                    continue
                res, fm, fx, fb = run_once(st, resolved_plants)
                if "error" in res:
                    return (res, {"display": "block"}, {"display": "none"})
                ps = res.get("plant_summary", pd.DataFrame())
                if isinstance(ps, pd.DataFrame) and not ps.empty:
                    ps = ps.copy()
                    ps["Custom Setting"] = nm
                    res["plant_summary"] = ps
                res["_preset_name"] = nm
                res["_final_err_mode"] = fm
                res["_final_x_pct"] = fx
                res["_final_bands"] = fb
                multi.append(res)
        else:
            # GLOBAL mode: Backward-compatible single setting path (prefer saved settings if present)
            base_settings = saved_settings if (isinstance(saved_settings, dict) and saved_settings) else {
                "err_mode": err_mode,
                "x_pct": x_pct,
                "bands": bands_rows
            }
            res, fm, fx, fb = run_once(base_settings, resolved_plants)
            if "error" in res:
                return (res, {"display": "block"}, {"display": "none"})
            ps = res.get("plant_summary", pd.DataFrame())
            if isinstance(ps, pd.DataFrame) and not ps.empty:
                ps = ps.copy()
                ps["Custom Setting"] = (base_settings.get("name") or base_settings.get("label") or "Current")
                res["plant_summary"] = ps
            res["_preset_name"] = ps["Custom Setting"].iloc[0] if isinstance(ps, pd.DataFrame) and not ps.empty else "Current"
            res["_final_err_mode"] = fm
            res["_final_x_pct"] = fx
            res["_final_bands"] = fb
            multi.append(res)

        combined_ps = pd.concat([
            m["plant_summary"] for m in multi if "plant_summary" in m and isinstance(m["plant_summary"], pd.DataFrame)
        ], ignore_index=True) if multi else pd.DataFrame()

        # Combine detail DataFrames for PLANT_WISE mode or use first for GLOBAL mode
        if application_mode == "PLANT_WISE" and all_detail_dfs:
            combined_detail = pd.concat(all_detail_dfs, ignore_index=True)
            combined_data_quality = compute_data_quality_stats(combined_detail)
        else:
            combined_detail = multi[0].get("df", pd.DataFrame()) if multi else pd.DataFrame()
            combined_data_quality = multi[0].get("data_quality", {}) if multi else {}

        first = multi[0] if multi else {}
        df_pack = _pack_df_for_store(combined_detail if isinstance(combined_detail, pd.DataFrame) else pd.DataFrame())
        out_payload = {
            **df_pack,
            "kpis": first.get("kpis", {}),
            "blockwise": first.get("blockwise", pd.DataFrame()).to_dict("records"),
            "plant_summary": combined_ps.to_dict("records"),
            "data_quality": combined_data_quality,
            "used_settings": {
                "multi": [m["_preset_name"] for m in multi],
            },
            "band_labels": [band.get("label", "no-label") for band in (first.get("final_bands") or first.get("_final_bands") or [])[:5]],
            "final_bands": first.get("final_bands", first.get("_final_bands", [])),
            # IMPORTANT: do NOT send per-run detail previews to the browser (can cause "Aw, Snap" OOM).
            # For downloads we store a cache-key per run.
            "_all_runs": [
                {
                    "name": m["_preset_name"],
                    "plants": m.get("_plants", []),
                    **_pack_df_for_store_cache_only(m.get("df", pd.DataFrame()) if isinstance(m, dict) else pd.DataFrame()),
                    "final_bands": m.get("_final_bands", []),
                    "err_mode": m.get("_final_err_mode", "default"),
                    "x_pct": m.get("_final_x_pct", 0),
                }
                for m in multi
            ],
        }

        return (out_payload, {"display": "block"}, {"display": "none"})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return ({"error": str(e)}, {"display": "block"}, {"display": "none"})


@app.callback(
    Output("agg-results-store", "data"),
    Output("agg-results-section", "style"),
    Output("agg-progress-container", "style"),
    Input("agg-plot-now", "n_clicks"),
    State("agg-region-dd", "value"),
    State("agg-resource-type-dd", "value"),
    State("agg-state-dd", "value"),
    State("agg-qca-dd", "value"),
    State("agg-pooling-dd", "value"),
    State("agg-plant-dd", "value"),
    State("agg-exclude-plant-dd", "value"),
    State("agg-date-range", "start_date"),
    State("agg-date-range", "end_date"),
    State("err-mode", "value"),
    State("x-pct", "value"),
    State("bands-table", "data"),
    State("saved-settings-store", "data"),
    State("agg-analysis-preset-select", "value"),
    State("presets-store", "data"),
    State("agg-ppa-mode", "value"),
    State("agg-ppa-value", "value"),
    prevent_initial_call=True
)
def compute_agg_on_click(
    n_plot,
    regions,
    resource_type,
    states,
    qcas,
    pools,
    plants,
    excluded_plants,
    start_date,
    end_date,
    err_mode,
    x_pct,
    bands_rows,
    saved_settings,
    selected_preset_names,
    presets_store,
    agg_ppa_mode,
    agg_ppa_value,
):
    if not n_plot:
        raise PreventUpdate

    try:
        # Apply exclusion logic: Final Plants = Selected Plants - Excluded Plants
        final_plants = plants
        if isinstance(plants, list):
            is_select_all = "SELECT_ALL" in plants
        else:
            is_select_all = (plants == "SELECT_ALL")
        
        if is_select_all:
            # Resolve SELECT_ALL using Excel master (single source of truth)
            dfm = _filter_master(regions, states, resource_type, qcas, pools)
            all_plants = sorted({p for p in dfm.get("plant_name", pd.Series(dtype=str)).dropna().astype(str).tolist() if p.strip()})
            # Apply exclusion: remove excluded plants
            excluded_list = excluded_plants if (excluded_plants and isinstance(excluded_plants, list)) else (excluded_plants if excluded_plants else [])
            final_plants = [p for p in all_plants if p not in excluded_list]
            
            # Validation: ensure at least one plant remains
            if not final_plants:
                return (
                    {"error": "All plants have been excluded. Please revise your exclusion selection."},
                    {"display": "block"},
                    {"display": "none"}
                )
        else:
            # Manual selection - no exclusion applied
            final_plants = plants if isinstance(plants, list) else [plants]
        
        # Normalize PPA mode selection (multi-select dropdown)
        if agg_ppa_mode is None or agg_ppa_mode == "":
            ppa_modes = ["mean"]
        elif isinstance(agg_ppa_mode, list):
            ppa_modes = agg_ppa_mode if len(agg_ppa_mode) > 0 else ["mean"]
        else:
            ppa_modes = [str(agg_ppa_mode)]

        def _ppa_label(pm: str) -> str:
            pm = (pm or "").strip().lower()
            if pm == "median":
                return "PPA Median"
            if pm == "mode":
                return "PPA Mode"
            if pm == "weighted":
                return "PPA Weighted"
            if pm == "numeric":
                v = agg_ppa_value
                return f"PPA Numeric ({v})" if v is not None else "PPA Numeric"
            return "PPA Mean"

        def run_once(with_settings: dict, ppa_mode_one: str):
            final_err_mode = with_settings.get("err_mode", err_mode)
            final_x_pct = with_settings.get("x_pct", x_pct)
            saved_bands = with_settings.get("bands", None)
            final_bands = saved_bands if (saved_bands and len(saved_bands) > 0) else (bands_rows or DEFAULT_BANDS.copy())
            unpaid_threshold = float(with_settings.get("unpaid_oi_threshold", 15.0))
            ppa_mode_one = (ppa_mode_one or "mean").strip().lower()
            if ppa_mode_one == "numeric" and agg_ppa_value is None:
                return {"error": "Please enter a numeric PPA value for 'Numeric' mode."}, final_err_mode, final_x_pct, final_bands
            print(f"DEBUG - Aggregation Analysis using: err_mode={final_err_mode}, x_pct={final_x_pct}, bands_count={len(final_bands)}, ppa_mode={ppa_mode_one}")
            res = _compute_pipeline_aggregated(
                regions,
                final_plants,  # Use final_plants (with exclusions applied)
                start_date,
                end_date,
                final_err_mode,
                final_x_pct,
                final_bands,
                ppa_mode_one,
                agg_ppa_value,
                unpaid_threshold,
            )
            return res, final_err_mode, final_x_pct, final_bands

        multi = []
        if selected_preset_names:
            name_to_settings = {p["name"]: p.get("settings", {}) for p in (presets_store or []) if isinstance(p, dict) and "name" in p}
            print(f"DEBUG - Running multi-preset aggregation analysis for: {selected_preset_names}")
            for nm in selected_preset_names:
                st = name_to_settings.get(nm)
                if not st:
                    print(f"DEBUG - Preset '{nm}' not found in store; skipping (agg)")
                    continue
                for pm in ppa_modes:
                    res, fm, fx, fb = run_once(st, pm)
                    if "error" in res:
                        return (res, {"display": "block"}, {"display": "none"})
                    ps = res.get("plant_summary", pd.DataFrame())
                    scenario = f"{nm} - {_ppa_label(pm)}"
                    if isinstance(ps, pd.DataFrame) and not ps.empty:
                        ps = ps.copy()
                        ps["Custom Setting"] = scenario
                        res["plant_summary"] = ps
                    res["_preset_name"] = scenario
                    res["_final_err_mode"] = fm
                    res["_final_x_pct"] = fx
                    res["_final_bands"] = fb
                    res["_ppa_mode"] = pm
                    multi.append(res)
        else:
            base_settings = saved_settings if (isinstance(saved_settings, dict) and saved_settings) else {
                "err_mode": err_mode,
                "x_pct": x_pct,
                "bands": bands_rows
            }
            base_label = (base_settings.get("name") or base_settings.get("label") or "Current")
            for pm in ppa_modes:
                res, fm, fx, fb = run_once(base_settings, pm)
                if "error" in res:
                    return (res, {"display": "block"}, {"display": "none"})
                ps = res.get("plant_summary", pd.DataFrame())
                scenario = f"{base_label} - {_ppa_label(pm)}" if len(ppa_modes) > 1 else base_label
                if isinstance(ps, pd.DataFrame) and not ps.empty:
                    ps = ps.copy()
                    ps["Custom Setting"] = scenario
                    res["plant_summary"] = ps
                res["_preset_name"] = scenario
                res["_final_err_mode"] = fm
                res["_final_x_pct"] = fx
                res["_final_bands"] = fb
                res["_ppa_mode"] = pm
                multi.append(res)

        combined_ps = pd.concat([
            m["plant_summary"] for m in multi if "plant_summary" in m and isinstance(m["plant_summary"], pd.DataFrame)
        ], ignore_index=True) if multi else pd.DataFrame()

        first = multi[0]
        used_settings = {
            "multi": [m["_preset_name"] for m in multi],
            "aggregated": True,
            "aggregated_plants": first.get("aggregated_plants", []),
            "aggregated_region": first.get("aggregated_region", ""),
            "ppa_modes": ppa_modes,
        }
        out_payload = {
            "df": first["df"].to_json(date_format="iso", orient="records"),
            "kpis": first.get("kpis", {}),
            "blockwise": first.get("blockwise", pd.DataFrame()).to_dict("records"),
            "plant_summary": combined_ps.to_dict("records"),
            "data_quality": first.get("data_quality", {}),
            "used_settings": used_settings,
            "band_labels": [band.get("label", "no-label") for band in (first.get("final_bands") or first.get("_final_bands") or [])[:5]],
            "final_bands": first.get("final_bands", first.get("_final_bands", [])),
            "_all_runs": [
                {
                    "name": m["_preset_name"],
                    "df": m["df"].to_json(date_format="iso", orient="records"),
                    "final_bands": m.get("_final_bands", []),
                    "err_mode": m.get("_final_err_mode", "default"),
                    "x_pct": m.get("_final_x_pct", 0),
                    "ppa_mode": m.get("_ppa_mode", None),
                    "ppa_value": agg_ppa_value,
                } for m in multi
            ]
        }

        return (out_payload, {"display": "block"}, {"display": "none"})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return ({"error": str(e)}, {"display": "block"}, {"display": "none"})

# --------- Stats availability compute ---------
@app.callback(
    Output("stats-results", "children"),
    Input("stats-run", "n_clicks"),
    State("stats-region-dd", "value"),
    State("stats-plant-dd", "value"),
    State("stats-date-range", "start_date"),
    State("stats-date-range", "end_date"),
    prevent_initial_call=True
)
def run_stats(n, regions, plants, start_date, end_date):
    if not n:
        raise PreventUpdate
    if not regions or not plants:
        return dbc.Alert("Select regions and plants", color="warning")

    if "SELECT_ALL" in (plants or []):
        plants = get_plants_from_duckdb(regions)

    region = regions[0]
    df = load_region_data(region, str(start_date), str(end_date), plants)
    if df.empty:
        return dbc.Alert("No data for selection", color="warning")

    # availability logic: consider rows where Scheduled_MW>0 then check PPA and Actual_MW present
    df["Scheduled_MW"] = pd.to_numeric(df["Scheduled_MW"], errors="coerce").fillna(0)
    df["Actual_MW"] = pd.to_numeric(df["Actual_MW"], errors="coerce")
    df["PPA"] = pd.to_numeric(df["PPA"], errors="coerce")

    df["is_candidate"] = df["Scheduled_MW"] > 0
    df["is_ok"] = df["is_candidate"] & df["Actual_MW"].notna() & df["PPA"].notna()

    def summarize_availability(pdf):
        cand = int(pdf["is_candidate"].sum())
        ok = int(pdf["is_ok"].sum())
        pct = (ok / cand * 100.0) if cand > 0 else 0.0
        missing = cand - ok
        return pd.Series({"availability": round(pct, 2), "missing": int(missing)})

    stats = df.groupby("Plant").apply(summarize_availability).reset_index().rename(columns={"Plant": "Plant name"})
    stats["Date Range"] = f"{start_date} → {end_date}"
    stats = stats[["Plant name", "Date Range", "availability", "missing"]]

    table = dash_table.DataTable(
        columns=[{"name": c, "id": c} for c in stats.columns],
        data=stats.to_dict("records"),
        sort_action="native",
        style_table={"overflowX": "auto"},
        style_cell={"padding": "6px"},
        style_header={"fontWeight": "600"},
        page_size=15,
    )
    return table

# Callback to show progress bar when plot button is clicked
@app.callback(
    Output("progress-container", "style", allow_duplicate=True),
    Input("plot-now", "n_clicks"),
    prevent_initial_call=True
)
def show_progress_bar(n):
    """Show progress bar when analysis starts"""
    if not n:
        raise PreventUpdate
    return {"display": "block", "marginTop": "1.5rem"}


@app.callback(
    Output("agg-progress-container", "style", allow_duplicate=True),
    Input("agg-plot-now", "n_clicks"),
    prevent_initial_call=True
)
def show_agg_progress_bar(n_plot):
    """Show progress bar when aggregation analysis starts."""
    if not n_plot:
        raise PreventUpdate
    return {"display": "block", "marginTop": "1.5rem"}

## Removed separate nav-store updater to avoid duplicate outputs; handled in switch_nav_tabs

@app.callback(
    Output("tab-content", "children"),
    Input("results-store", "data"),
)
def render_tabs(stored):
    try:
        if not stored:
            return dbc.Alert("Click Plot Now to compute.", color="info")
        
        if not isinstance(stored, dict):
            return dbc.Alert(f"Unexpected data format: {type(stored)}", color="danger")
        
        if stored.get("error"):
            err = stored.get("error", "Unknown error")
            # Format error message for display
            if isinstance(err, str):
                msg = err
            else:
                msg = str(err)
            return dbc.Alert([
                html.H5("Error", className="alert-heading"),
                html.Pre(msg, style={"whiteSpace": "pre-wrap", "marginBottom": 0})
            ], color="danger")

        # UI uses the preview df from store; large runs keep full df server-side in cache.
        try:
            df_data = stored.get("df")
            if isinstance(df_data, str):
                df = pd.DataFrame(json.loads(df_data))
            elif isinstance(df_data, list):
                df = pd.DataFrame(df_data)
            elif isinstance(df_data, dict):
                df = pd.DataFrame([df_data])
            else:
                df = pd.DataFrame()
        except Exception as e:
            return dbc.Alert([
                html.H5("Data Loading Error", className="alert-heading"),
                html.P(f"Failed to load data: {str(e)}", style={"marginBottom": 0})
            ], color="danger")
        
        try:
            plant_summary_data = stored.get("plant_summary", [])
            if isinstance(plant_summary_data, list):
                plant_summary_df = pd.DataFrame(plant_summary_data)
            elif isinstance(plant_summary_data, dict):
                plant_summary_df = pd.DataFrame([plant_summary_data])
            else:
                plant_summary_df = pd.DataFrame()
        except Exception as e:
            print(f"WARNING - Failed to load plant summary: {e}")
            plant_summary_df = pd.DataFrame()
        # Enrich Plant Summary with master metadata (Analysis tab only)
        plant_summary_df = enrich_plant_summary_with_master(plant_summary_df)
        used = stored.get("used_settings", {})

        # We can render results from plant_summary alone (detail df may be cached-only for large runs).
        if plant_summary_df.empty:
            return dbc.Alert("No data found for the selected filters.", color="warning")

        # Data Quality section (use precomputed stats from pipeline to avoid recomputing on preview df)
        dq = stored.get("data_quality", {}) if isinstance(stored, dict) else {}
        dq_card = None
        dq_pie_chart = None
        dq_table = None
        try:
            if dq and dq.get("total_ts", 0):
                dq_card = dbc.Card(
                    [
                        dbc.CardHeader("Data Quality Stats"),
                        dbc.CardBody(
                            [
                                html.Div(f"Total slots checked: {dq.get('total_ts', 0)}", className="text-muted"),
                                html.Div(f"Data Available: {dq.get('available_pct', 0)}%  (count: {dq.get('available_count', 0)})"),
                                html.Div(f"Actual Power Missing: {dq.get('actual_missing_pct', 0)}%  (count: {dq.get('actual_missing_count', 0)})"),
                                html.Div(f"PPA / AVC Missing: {dq.get('ppa_avc_missing_pct', 0)}%  (count: {dq.get('ppa_avc_missing_count', 0)})"),
                                html.Div(
                                    f"DSM not calculated (zero Schedule/AVC/PPA): {dq.get('dsm_skipped_zero_inputs_pct', 0)}%  "
                                    f"(count: {dq.get('dsm_skipped_zero_inputs_count', 0)})"
                                ),
                            ]
                        ),
                    ],
                    className="mb-3",
                )
                
                # Data Quality Pie Chart (always sums to 100%)
                available_pct = dq.get('available_pct', 0.0)
                actual_missing_pct = dq.get('actual_missing_pct', 0.0)
                ppa_avc_missing_pct = dq.get('ppa_avc_missing_pct', 0.0)
                unknown_pct = round(100.0 - (available_pct + actual_missing_pct + ppa_avc_missing_pct), 2)
                
                pie_data = {
                    "Category": ["Data Available", "Actual Power Missing", "PPA / AVC Missing", "Unknown"],
                    "Percentage": [available_pct, actual_missing_pct, ppa_avc_missing_pct, unknown_pct]
                }
                pie_df = pd.DataFrame(pie_data)
                
                fig_pie = px.pie(
                    pie_df,
                    values="Percentage",
                    names="Category",
                    title="Data Quality Distribution",
                    color="Category",
                    color_discrete_map={
                        "Data Available": "#28a745",
                        "Actual Power Missing": "#dc3545",
                        "PPA / AVC Missing": "#ffc107",
                        "Unknown": "#6c757d"
                    },
                    hole=0.4  # Donut chart
                )
                fig_pie.update_traces(
                    textposition="inside",
                    textinfo="percent+label",
                    hovertemplate="<b>%{label}</b><br>Percentage: %{percent}<br>Value: %{value:.2f}%<extra></extra>"
                )
                fig_pie.update_layout(
                    template="plotly_white",
                    height=400,
                    showlegend=True,
                    legend=dict(orientation="v", yanchor="middle", y=0.5, xanchor="left", x=1.05)
                )
                
                dq_pie_chart = dbc.Card(
                    [
                        dbc.CardHeader("Data Quality Pie Chart"),
                        dbc.CardBody(
                            dcc.Graph(figure=fig_pie, config={"displayModeBar": True})
                        ),
                    ],
                    className="mb-3",
                )
                
                issues = dq.get("issues_rows", []) if isinstance(dq.get("issues_rows", []), list) else []
                issues_total = int(dq.get("issues_total", len(issues) if issues else 0) or 0)
                if issues_total > 0:
                    show_n = len(issues)
                    dq_table = html.Details(
                        [
                            html.Summary(f"Data Quality Issues (showing {show_n} of {issues_total})"),
                            html.Div(
                                dash_table.DataTable(
                                    columns=[{"name": c, "id": c} for c in (issues[0].keys() if issues else [])],
                                    data=issues,
                                    sort_action="native",
                                    filter_action="native",
                                    page_size=15,
                                    style_table={"overflowX": "auto"},
                                    style_cell={"padding": "6px", "textAlign": "left"},
                                    style_header={"fontWeight": "600", "backgroundColor": "#f8f9fa"},
                                ),
                                className="mt-2",
                            ),
                        ],
                        className="mb-3",
                    )
        except Exception:
            dq_card = None
            dq_pie_chart = None
            dq_table = None

        # Large run note (prevents "Aw, Snap" and explains why only some rows show in UI)
        df_note = None
        if isinstance(stored, dict) and stored.get("df_is_truncated"):
            df_note = dbc.Alert(
                f"Large selection detected: UI is showing a preview of {stored.get('df_preview_rows', 0)} / {stored.get('df_full_rows', 0)} rows. "
                f"All calculations and downloads use the full dataset.",
                color="info",
                className="mb-3",
            )

        # Analysis view simplified to only Plant Summary

        # Plant summary table (responsive + downloadable)
        plant_summary_table = dash_table.DataTable(
            columns=[
                {"name": c, "id": c, "type": "numeric", "format": {"specifier": ".2f"}} 
                if c in ["PPA", "Data Availability %", "Revenue Loss (%)", "Revenue Loss (p/k)", "Plant Capacity", "DSM Loss"]
                else {"name": c, "id": c}
                for c in plant_summary_df.columns
            ] if not plant_summary_df.empty else [],
            data=plant_summary_df.to_dict("records"),
            sort_action="native",
            filter_action="native",
            fill_width=True,
            fixed_rows={"headers": True},
            style_as_list_view=True,
            style_table={
                "overflowX": "auto",
                "overflowY": "auto",
                "maxWidth": "100%",
                "height": "60vh",
                "borderRadius": "12px",
                "border": "1px solid #eee",
            },
            style_cell={
                "padding": "10px",
                "textAlign": "left",
                "minWidth": "120px",
                "width": "140px",
                "maxWidth": "260px",
                "whiteSpace": "normal",
                "height": "auto",
                "fontSize": "0.92rem",
            },
            style_header={
                "fontWeight": "700",
                "backgroundColor": "#f8f9fa",
                "borderBottom": "1px solid #e9ecef",
            },
            style_data={
                "borderBottom": "1px solid #f1f3f5",
            },
        )

        # Optional comparison chart
        graph_div = None
        chart_heading = None
        try:
            if not plant_summary_df.empty and ("Plant name" in plant_summary_df.columns) and ("Revenue Loss (%)" in plant_summary_df.columns):
                ps = plant_summary_df.copy()
                ps = ps.sort_values("Revenue Loss (%)")
                print(f"DEBUG - Creating chart with {len(ps)} rows, columns: {list(ps.columns)}")
                
                # Get X threshold values for each preset from stored data
                preset_x_map = {}
                all_runs = stored.get("_all_runs", []) if isinstance(stored, dict) else []
                for run in all_runs:
                    preset_name = run.get("name", "")
                    err_mode = str(run.get("err_mode", "default")).lower()
                    x_pct = float(run.get("x_pct", 100))
                    # For default mode, X is effectively 100%
                    x_threshold = 100.0 if err_mode == "default" else x_pct
                    preset_x_map[preset_name] = {
                        "x_value": x_threshold,
                        "err_mode": err_mode,
                        "x_pct": x_pct
                    }
                print(f"DEBUG - Preset X thresholds: {preset_x_map}")
                
                # Build per-row X label for bar text
                def label_for_setting(setting_name: str) -> str:
                    info = preset_x_map.get(setting_name, None)
                    if not info:
                        return ""
                    x_val = int(round(float(info.get("x_value", 0))))
                    return f"X = {x_val} %"

                if "Custom Setting" in ps.columns:
                    ps["X Label"] = ps["Custom Setting"].apply(label_for_setting)
                    fig = px.bar(
                        ps,
                        x="Revenue Loss (%)",
                        y="Plant name",
                        color="Custom Setting",
                        text="X Label",
                        orientation="h",
                        barmode="group",
                    )
                    fig.update_layout(legend_title_text="Setting")
                else:
                    # Single setting path - use first run's X from preset_x_map if present
                    if len(preset_x_map) > 0:
                        single_name = list(preset_x_map.keys())[0]
                        x_val = int(round(float(preset_x_map[single_name].get("x_value", 0))))
                        ps["X Label"] = [f"X = {x_val} %"] * len(ps)
                    else:
                        ps["X Label"] = [""] * len(ps)
                    fig = px.bar(
                        ps,
                        x="Revenue Loss (%)",
                        y="Plant name",
                        text="X Label",
                        orientation="h",
                    )
                
                # (Removed) Vertical threshold lines and top annotations per request
                
                # Ensure text labels are outside and colored like the bar
                for tr in getattr(fig, "data", []):
                    try:
                        tr.textposition = "outside"
                        if hasattr(tr, "marker") and hasattr(tr.marker, "color"):
                            tr.textfont.color = tr.marker.color
                            tr.textfont.size = 12
                    except Exception:
                        pass

                n_rows = len(ps.index)
                fig.update_layout(
                    template="plotly_white",
                    title="Revenue Loss % Comparison",
                    xaxis_title="Revenue Loss (%)",
                    yaxis_title="Plant Name",
                    margin=dict(l=150, r=60, t=60, b=60),  # Reduced top margin; no annotations/threshold lines
                    height=max(450, 220 + 40 * n_rows),
                    bargap=0.25,
                )
                chart_height = max(450, 220 + 40 * n_rows)
                graph_div = dcc.Graph(
                    id="revenue-loss-chart",
                    figure=fig,
                    config={"displayModeBar": True},
                    style={"width": "100%", "height": f"{chart_height}px"}
                )
                chart_heading = html.H6("Revenue Loss % — Comparison", 
                                       style={"marginTop": "1rem", "marginBottom": "0.5rem", "color": "#333", "fontWeight": "600"})
                print(f"DEBUG - Chart created successfully, height={chart_height}")
        except Exception as e:
            import traceback
            print(f"DEBUG - Chart build failed: {e}")
            traceback.print_exc()

        download_btn = dbc.Button("Download Full Calculation (Excel)", id="btn-download", color="primary", className="mt-2")
        plant_summary_download_btn = dbc.Button(
            "Download Plant Summary (Excel)",
            id="btn-download-plant-summary",
            color="secondary",
            outline=True,
            className="mt-2 ms-2",
        )

        # Debug: check if graph was created
        print(f"DEBUG - graph_div is None: {graph_div is None}")
        print(f"DEBUG - chart_heading is None: {chart_heading is None}")
        if graph_div:
            print(f"DEBUG - graph_div type: {type(graph_div)}")
            print(f"DEBUG - graph_div.id: {getattr(graph_div, 'id', 'no id')}")

        # Build children list - put chart BEFORE table
        children_list = [
            *( [df_note] if df_note is not None else [] ),
            *( [dq_card] if dq_card is not None else [] ),
            *( [dq_pie_chart] if dq_pie_chart is not None else [] ),
            *( [dq_table] if dq_table is not None else [] ),
        ]
        
        # Add chart heading and graph if available (BEFORE table)
        if graph_div is not None:
            print(f"DEBUG - Adding chart to display...")
            if chart_heading:
                children_list.append(chart_heading)
            # Add graph with minimal wrapper
            children_list.append(graph_div)
            print(f"DEBUG - Chart added! Total children before table: {len(children_list)}")
        
        # Plant Summary (modern responsive card + buttons always visible)
        children_list.append(
            dbc.Card(
                [
                    dbc.CardHeader(
                        dbc.Row(
                            [
                                dbc.Col(
                                    html.H6("Plant Summary", style={"margin": 0, "fontWeight": "700", "color": "#333"}),
                                    md=6,
                                ),
                                dbc.Col(
                                    html.Div([download_btn, plant_summary_download_btn], className="text-end"),
                                    md=6,
                                ),
                            ],
                            align="center",
                        ),
                        style={"backgroundColor": "#ffffff"},
                    ),
                    dbc.CardBody(plant_summary_table, style={"padding": "0.75rem"}),
                ],
                className="shadow-sm mb-3",
                style={"borderRadius": "14px"},
            )
        )

        # Band-wise DSM section (per-plant)
        children_list.append(bandwise_section_layout("analysis"))

        return dcc.Loading(
            type="circle",
            children=html.Div(children_list)
        )
    except Exception as e:
        return dbc.Alert(f"Failed to render results: {e}", color="danger")


@app.callback(
    Output("agg-tab-content", "children"),
    Input("agg-results-store", "data"),
)
def render_agg_tabs(stored):
    """Render results for Aggregation Analysis using the same layout as main Analysis,
    with an additional note indicating the plants included in the aggregation.
    """
    try:
        if not stored or (isinstance(stored, dict) and stored.get("error")):
            err = stored.get("error") if isinstance(stored, dict) else ""
            msg = err or "Click Plot Now to compute."
            return dbc.Alert(msg, color="warning")

        df = pd.DataFrame(json.loads(stored["df"])) if isinstance(stored.get("df"), str) else pd.DataFrame(stored.get("df", []))
        plant_summary_df = pd.DataFrame(stored.get("plant_summary", []))
        used = stored.get("used_settings", {}) if isinstance(stored, dict) else {}

        if df.empty:
            return dbc.Alert("No data found for the selected filters.", color="warning")

        # Data Quality section (computed from raw pre-aggregation slot series)
        # IMPORTANT: Always compute Data Quality from the same slot-level detail dataframe
        # used by this view (stored["df"] -> df) to avoid mismatches.
        dq = compute_data_quality_stats(df) if not df.empty else {}
        dq_card = None
        dq_pie_chart = None
        dq_table = None
        try:
            if dq and dq.get("total_ts", 0):
                dq_card = dbc.Card(
                    [
                        dbc.CardHeader("Data Quality Stats"),
                        dbc.CardBody(
                            [
                                html.Div(f"Total slots checked: {dq.get('total_ts', 0)}", className="text-muted"),
                                html.Div(f"Data Available: {dq.get('available_pct', 0)}%  (count: {dq.get('available_count', 0)})"),
                                html.Div(f"Actual Power Missing: {dq.get('actual_missing_pct', 0)}%  (count: {dq.get('actual_missing_count', 0)})"),
                                html.Div(f"PPA / AVC Missing: {dq.get('ppa_avc_missing_pct', 0)}%  (count: {dq.get('ppa_avc_missing_count', 0)})"),
                            ]
                        ),
                    ],
                    className="mb-3",
                )
                
                # Data Quality Pie Chart (always sums to 100%)
                available_pct = dq.get('available_pct', 0.0)
                actual_missing_pct = dq.get('actual_missing_pct', 0.0)
                ppa_avc_missing_pct = dq.get('ppa_avc_missing_pct', 0.0)
                unknown_pct = round(100.0 - (available_pct + actual_missing_pct + ppa_avc_missing_pct), 2)
                
                pie_data = {
                    "Category": ["Data Available", "Actual Power Missing", "PPA / AVC Missing", "Unknown"],
                    "Percentage": [available_pct, actual_missing_pct, ppa_avc_missing_pct, unknown_pct]
                }
                pie_df = pd.DataFrame(pie_data)
                
                fig_pie = px.pie(
                    pie_df,
                    values="Percentage",
                    names="Category",
                    title="Data Quality Distribution",
                    color="Category",
                    color_discrete_map={
                        "Data Available": "#28a745",
                        "Actual Power Missing": "#dc3545",
                        "PPA / AVC Missing": "#ffc107",
                        "Unknown": "#6c757d"
                    },
                    hole=0.4  # Donut chart
                )
                fig_pie.update_traces(
                    textposition="inside",
                    textinfo="percent+label",
                    hovertemplate="<b>%{label}</b><br>Percentage: %{percent}<br>Value: %{value:.2f}%<extra></extra>"
                )
                fig_pie.update_layout(
                    template="plotly_white",
                    height=400,
                    showlegend=True,
                    legend=dict(orientation="v", yanchor="middle", y=0.5, xanchor="left", x=1.05)
                )
                
                dq_pie_chart = dbc.Card(
                    [
                        dbc.CardHeader("Data Quality Pie Chart"),
                        dbc.CardBody(
                            dcc.Graph(figure=fig_pie, config={"displayModeBar": True})
                        ),
                    ],
                    className="mb-3",
                )
                
                issues = dq.get("issues_rows", []) if isinstance(dq.get("issues_rows", []), list) else []
                issues_total = int(dq.get("issues_total", len(issues) if issues else 0) or 0)
                if issues_total > 0:
                    show_n = len(issues)
                    dq_table = html.Details(
                        [
                            html.Summary(f"Data Quality Issues (showing {show_n} of {issues_total})"),
                            html.Div(
                                dash_table.DataTable(
                                    columns=[{"name": c, "id": c} for c in (issues[0].keys() if issues else [])],
                                    data=issues,
                                    sort_action="native",
                                    filter_action="native",
                                    page_size=15,
                                    style_table={"overflowX": "auto"},
                                    style_cell={"padding": "6px", "textAlign": "left"},
                                    style_header={"fontWeight": "600", "backgroundColor": "#f8f9fa"},
                                ),
                                className="mt-2",
                            ),
                        ],
                        className="mb-3",
                    )
        except Exception:
            dq_card = None
            dq_pie_chart = None
            dq_table = None

        # Plant summary table (aggregated)
        plant_summary_table = dash_table.DataTable(
            columns=[
                {"name": c, "id": c, "type": "numeric", "format": {"specifier": ".2f"}} 
                if c in ["PPA", "Data Availability %", "Revenue Loss (%)", "Revenue Loss (p/k)", "Plant Capacity", "DSM Loss"]
                else {"name": c, "id": c}
                for c in plant_summary_df.columns
            ] if not plant_summary_df.empty else [],
            data=plant_summary_df.to_dict("records"),
            sort_action="native",
            style_table={"overflowX": "auto"},
            style_cell={"padding": "6px", "textAlign": "left"},
            style_header={"fontWeight": "600", "backgroundColor": "#f8f9fa"},
        )

        # Optional comparison chart (still valid for multiple presets, but on aggregated profile)
        graph_div = None
        chart_heading = None
        try:
            if not plant_summary_df.empty and ("Plant name" in plant_summary_df.columns) and ("Revenue Loss (%)" in plant_summary_df.columns):
                ps = plant_summary_df.copy()
                ps = ps.sort_values("Revenue Loss (%)")
                print(f"DEBUG - Creating aggregation chart with {len(ps)} rows, columns: {list(ps.columns)}")

                preset_x_map = {}
                all_runs = stored.get("_all_runs", []) if isinstance(stored, dict) else []
                for run in all_runs:
                    preset_name = run.get("name", "")
                    err_mode = str(run.get("err_mode", "default")).lower()
                    x_pct = float(run.get("x_pct", 100))
                    x_threshold = 100.0 if err_mode == "default" else x_pct
                    preset_x_map[preset_name] = {
                        "x_value": x_threshold,
                        "err_mode": err_mode,
                        "x_pct": x_pct
                    }

                def label_for_setting(setting_name: str) -> str:
                    info = preset_x_map.get(setting_name, None)
                    if not info:
                        return ""
                    x_val = int(round(float(info.get("x_value", 0))))
                    return f"X = {x_val} %"

                if "Custom Setting" in ps.columns:
                    ps["X Label"] = ps["Custom Setting"].apply(label_for_setting)
                    fig = px.bar(
                        ps,
                        x="Revenue Loss (%)",
                        y="Plant name",
                        color="Custom Setting",
                        text="X Label",
                        orientation="h",
                        barmode="group",
                    )
                    fig.update_layout(legend_title_text="Setting")
                else:
                    if len(preset_x_map) > 0:
                        single_name = list(preset_x_map.keys())[0]
                        x_val = int(round(float(preset_x_map[single_name].get("x_value", 0))))
                        ps["X Label"] = [f"X = {x_val} %"] * len(ps)
                    else:
                        ps["X Label"] = [""] * len(ps)
                    fig = px.bar(
                        ps,
                        x="Revenue Loss (%)",
                        y="Plant name",
                        text="X Label",
                        orientation="h",
                    )

                for tr in getattr(fig, "data", []):
                    try:
                        tr.textposition = "outside"
                        if hasattr(tr, "marker") and hasattr(tr.marker, "color"):
                            tr.textfont.color = tr.marker.color
                            tr.textfont.size = 12
                    except Exception:
                        pass

                n_rows = len(ps.index)
                fig.update_layout(
                    template="plotly_white",
                    title="Revenue Loss % Comparison (Aggregated Profile)",
                    xaxis_title="Revenue Loss (%)",
                    yaxis_title="Plant Name",
                    margin=dict(l=150, r=60, t=60, b=60),
                    height=max(450, 220 + 40 * n_rows),
                    bargap=0.25,
                )
                chart_height = max(450, 220 + 40 * n_rows)
                graph_div = dcc.Graph(
                    id="agg-revenue-loss-chart",
                    figure=fig,
                    config={"displayModeBar": True},
                    style={"width": "100%", "height": f"{chart_height}px"}
                )
                chart_heading = html.H6(
                    "Revenue Loss % — Aggregated Comparison", 
                    style={"marginTop": "1rem", "marginBottom": "0.5rem", "color": "#333", "fontWeight": "600"}
                )
        except Exception as e:
            import traceback
            print(f"DEBUG - Aggregation chart build failed: {e}")
            traceback.print_exc()

        # Aggregation note
        agg_plants = used.get("aggregated_plants") or []
        agg_region = used.get("aggregated_region") or ""
        note_children = []
        if agg_plants:
            plants_text = ", ".join(map(str, agg_plants))
            note_text = f"Aggregated output for {len(agg_plants)} plants: {plants_text}"
            if agg_region:
                note_text = f"{note_text} (Region: {agg_region})"
            note_children.append(
                html.Div(
                    note_text,
                    style={
                        "fontSize": "0.9rem",
                        "color": "#555",
                        "backgroundColor": "#f8f9fa",
                        "borderRadius": "6px",
                        "padding": "8px 12px",
                        "marginBottom": "0.75rem",
                        "border": "1px dashed #ddd",
                    },
                )
            )

        children_list = note_children + [
            *( [dq_card] if dq_card is not None else [] ),
            *( [dq_pie_chart] if dq_pie_chart is not None else [] ),
            *( [dq_table] if dq_table is not None else [] ),
            html.H5("Plant Summary (Aggregated)", style={"marginTop": "0.5rem", "marginBottom": "1rem", "color": "#333"}),
        ]

        if graph_div is not None:
            if chart_heading:
                children_list.append(chart_heading)
            children_list.append(graph_div)

        children_list.append(
            dbc.Row([dbc.Col(plant_summary_table, md=12)])
        )
        children_list.append(
            html.Div(
                dbc.Button(
                    "Download Full Calculation (Excel)",
                    id={"type": "agg-btn-download", "index": "main"},
                    color="primary",
                    className="mt-2",
                ),
                className="text-end mt-2",
            )
        )

        return dcc.Loading(
            type="circle",
            children=html.Div(children_list)
        )
    except Exception as e:
        return dbc.Alert(f"Failed to render aggregated results: {e}", color="danger")


# =========================
# Band-wise DSM callbacks
# =========================
@app.callback(
    Output("analysis-band-plant-dd", "options"),
    Output("analysis-band-plant-dd", "value"),
    Input("results-store", "data"),
)
def _bandwise_analysis_plants(stored):
    if not stored or (isinstance(stored, dict) and stored.get("error")):
        return [], None
    df = _detail_df_from_store(stored)
    if df.empty:
        return [], None
    if "plant_name" not in df.columns and "Plant" in df.columns:
        df["plant_name"] = df["Plant"]
    plants = sorted({str(p) for p in df.get("plant_name", pd.Series(dtype=str)).dropna().unique()})
    opts = [{"label": p, "value": p} for p in plants]
    return opts, (plants[0] if plants else None)


@app.callback(
    Output("analysis-band-chart", "figure"),
    Output("analysis-band-table", "data"),
    Input("analysis-band-plant-dd", "value"),
    Input("analysis-band-metric", "value"),
    State("results-store", "data"),
)
def _bandwise_analysis_update(plant_name, metric, stored):
    if not stored or not plant_name:
        raise PreventUpdate
    df = _detail_df_from_store(stored)
    if df.empty:
        raise PreventUpdate
    bands_rows = stored.get("final_bands", []) if isinstance(stored, dict) else []
    band_df, band_order = compute_bandwise_for_dashboard(df, bands_rows)
    plant_df = band_df[band_df["plant_name"] == str(plant_name)].copy()
    # Ensure all bands exist in output with zeros (keeps X-axis stable)
    if band_order:
        all_rows = []
        for b in band_order:
            d = "UI" if str(b).strip().upper().startswith("UI") else ("OI" if str(b).strip().upper().startswith("OI") else "")
            all_rows.append({"plant_name": plant_name, "band": b, "direction": d, "energy_kwh": 0.0, "revenue_loss": 0.0})
        base = pd.DataFrame(all_rows)
        if not plant_df.empty:
            plant_df = base.merge(
                plant_df,
                on=["plant_name", "band", "direction"],
                how="left",
                suffixes=("_base", ""),
            )
            for col in ["energy_kwh", "revenue_loss"]:
                if f"{col}_base" in plant_df.columns:
                    plant_df[col] = plant_df[col].fillna(0.0) + plant_df[f"{col}_base"].fillna(0.0)
            plant_df = plant_df[["plant_name", "band", "direction", "energy_kwh", "revenue_loss"]]
        else:
            plant_df = base
    fig = make_bandwise_bar_chart(plant_df, band_order, metric, str(plant_name))
    return fig, plant_df.to_dict("records")


@app.callback(
    Output("agg-band-plant-dd", "options"),
    Output("agg-band-plant-dd", "value"),
    Input("agg-results-store", "data"),
)
def _bandwise_agg_plants(stored):
    if not stored or (isinstance(stored, dict) and stored.get("error")):
        return [], None
    df = _detail_df_from_store(stored)
    if df.empty:
        return [], None
    if "plant_name" not in df.columns and "Plant" in df.columns:
        df["plant_name"] = df["Plant"]
    plants = sorted({str(p) for p in df.get("plant_name", pd.Series(dtype=str)).dropna().unique()})
    opts = [{"label": p, "value": p} for p in plants]
    return opts, (plants[0] if plants else None)


@app.callback(
    Output("agg-band-chart", "figure"),
    Output("agg-band-table", "data"),
    Input("agg-band-plant-dd", "value"),
    Input("agg-band-metric", "value"),
    State("agg-results-store", "data"),
)
def _bandwise_agg_update(plant_name, metric, stored):
    if not stored or not plant_name:
        raise PreventUpdate
    df = _detail_df_from_store(stored)
    if df.empty:
        raise PreventUpdate
    bands_rows = stored.get("final_bands", []) if isinstance(stored, dict) else []
    band_df, band_order = compute_bandwise_for_dashboard(df, bands_rows)
    plant_df = band_df[band_df["plant_name"] == str(plant_name)].copy()
    if band_order:
        all_rows = []
        for b in band_order:
            d = "UI" if str(b).strip().upper().startswith("UI") else ("OI" if str(b).strip().upper().startswith("OI") else "")
            all_rows.append({"plant_name": plant_name, "band": b, "direction": d, "energy_kwh": 0.0, "revenue_loss": 0.0})
        base = pd.DataFrame(all_rows)
        if not plant_df.empty:
            plant_df = base.merge(
                plant_df,
                on=["plant_name", "band", "direction"],
                how="left",
                suffixes=("_base", ""),
            )
            for col in ["energy_kwh", "revenue_loss"]:
                if f"{col}_base" in plant_df.columns:
                    plant_df[col] = plant_df[col].fillna(0.0) + plant_df[f"{col}_base"].fillna(0.0)
            plant_df = plant_df[["plant_name", "band", "direction", "energy_kwh", "revenue_loss"]]
        else:
            plant_df = base
    fig = make_bandwise_bar_chart(plant_df, band_order, metric, str(plant_name))
    return fig, plant_df.to_dict("records")


@app.callback(
    Output("custom-band-plant-dd", "options"),
    Output("custom-band-plant-dd", "value"),
    Input("custom-results-store", "data"),
)
def _bandwise_custom_plants(stored):
    if not stored or (isinstance(stored, dict) and stored.get("error")):
        return [], None
    df = pd.DataFrame(stored.get("df", [])) if isinstance(stored, dict) else pd.DataFrame()
    if df.empty:
        return [], None
    if "plant_name" not in df.columns and "Plant" in df.columns:
        df["plant_name"] = df["Plant"]
    plants = sorted({str(p) for p in df.get("plant_name", pd.Series(dtype=str)).dropna().unique()})
    opts = [{"label": p, "value": p} for p in plants]
    return opts, (plants[0] if plants else None)


@app.callback(
    Output("custom-band-chart", "figure"),
    Output("custom-band-table", "data"),
    Input("custom-band-plant-dd", "value"),
    Input("custom-band-metric", "value"),
    State("custom-results-store", "data"),
)
def _bandwise_custom_update(plant_name, metric, stored):
    if not stored or not plant_name:
        raise PreventUpdate
    df = pd.DataFrame(stored.get("df", [])) if isinstance(stored, dict) else pd.DataFrame()
    if df.empty:
        raise PreventUpdate
    bands_rows = stored.get("final_bands", []) if isinstance(stored, dict) else []
    err_mode = stored.get("err_mode", "default") if isinstance(stored, dict) else "default"
    x_pct = stored.get("x_pct", 50.0) if isinstance(stored, dict) else 50.0
    band_df, band_order = compute_bandwise_for_dashboard(df, bands_rows, err_mode=str(err_mode), x_pct=float(x_pct))
    plant_df = band_df[band_df["plant_name"] == str(plant_name)].copy()
    if band_order:
        all_rows = []
        for b in band_order:
            d = "UI" if str(b).strip().upper().startswith("UI") else ("OI" if str(b).strip().upper().startswith("OI") else "")
            all_rows.append({"plant_name": plant_name, "band": b, "direction": d, "energy_kwh": 0.0, "revenue_loss": 0.0})
        base = pd.DataFrame(all_rows)
        if not plant_df.empty:
            plant_df = base.merge(
                plant_df,
                on=["plant_name", "band", "direction"],
                how="left",
                suffixes=("_base", ""),
            )
            for col in ["energy_kwh", "revenue_loss"]:
                if f"{col}_base" in plant_df.columns:
                    plant_df[col] = plant_df[col].fillna(0.0) + plant_df[f"{col}_base"].fillna(0.0)
            plant_df = plant_df[["plant_name", "band", "direction", "energy_kwh", "revenue_loss"]]
        else:
            plant_df = base
    fig = make_bandwise_bar_chart(plant_df, band_order, metric, str(plant_name))
    return fig, plant_df.to_dict("records")

@app.callback(
    Output("custom-upload-preset-select", "options"),
    Input("presets-store", "data"),
)
def load_custom_preset_options(presets):
    presets = presets or []
    return [{"label": p.get("name", "Unnamed"), "value": p.get("name", "Unnamed")} for p in presets]

@app.callback(
    Output("dl-sample-xlsx", "data"),
    Input("btn-download-sample-xlsx", "n_clicks"),
    prevent_initial_call=True
)
def dl_sample_xlsx(n):
    if not n:
        raise PreventUpdate
    import pandas as pd
    from io import BytesIO
    # Build sample dataframe
    def block_to_times(block):
        start = (int(block) - 1) * 15
        h1, m1 = divmod(start, 60)
        end = start + 15
        h2, m2 = divmod(end, 60)
        return f"{h1:02d}:{m1:02d}", f"{h2:02d}:{m2:02d}"
    rows = []
    for blk in [21, 22, 23, 24]:
        ft, tt = block_to_times(blk)
        rows.append({
            "region": "NRPC",
            "plant_name": "Plant-A",
            "date": "2025-01-01",
            "time_block": blk,
            "from_time": ft,
            "to_time": tt,
            "AvC_MW": 50,
            "Scheduled_MW": 45,
            "Actual_MW": 40,
            "PPA": 3.0,
        })
    df = pd.DataFrame(rows)
    out = BytesIO()
    with pd.ExcelWriter(out, engine="xlsxwriter") as xw:
        df.to_excel(xw, sheet_name="Template", index=False)
    return dcc.send_bytes(out.getvalue(), filename="custom_upload_template.xlsx")

@app.callback(
    Output("dl-sample-csv", "data"),
    Input("btn-download-sample-csv", "n_clicks"),
    prevent_initial_call=True
)
def dl_sample_csv(n):
    if not n:
        raise PreventUpdate
    import pandas as pd
    from io import StringIO
    sio = StringIO()
    df = pd.DataFrame([{
        "region": "NRPC",
        "plant_name": "Plant-A",
        "date": "2025-01-01",
        "time_block": 21,
        "from_time": "05:00",
        "to_time": "05:15",
        "AvC_MW": 50,
        "Scheduled_MW": 45,
        "Actual_MW": 40,
        "PPA": 3.0,
    }])
    df.to_csv(sio, index=False)
    return dict(content=sio.getvalue(), filename="custom_upload_template.csv", type="text/csv")

@app.callback(
    Output("custom-upload-store", "data"),
    Output("custom-upload-validate", "children"),
    Input("upload-custom-file", "contents"),
    State("upload-custom-file", "filename"),
    prevent_initial_call=True
)
def handle_upload(contents, filename):
    if not contents:
        raise PreventUpdate
    import base64, io
    header, b64 = contents.split(",", 1)
    data = base64.b64decode(b64)
    try:
        if filename.lower().endswith(".csv"):
            df = pd.read_csv(io.BytesIO(data))
        else:
            df = pd.read_excel(io.BytesIO(data))
    except Exception as e:
        return None, dbc.Alert(f"Unable to read file: {e}", color="danger")

    def canon(s):
        return str(s).strip()
    df.columns = [canon(c) for c in df.columns]

    required = ["region","plant_name","date","time_block","AvC_MW","Scheduled_MW","Actual_MW","PPA"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        return None, dbc.Alert(f"Missing required columns: {', '.join(missing)}", color="warning")

    for opt in ["from_time","to_time"]:
        if opt not in df.columns:
            df[opt] = ""

    for c in ["time_block","AvC_MW","Scheduled_MW","Actual_MW","PPA"]:
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)

    return df.to_json(orient="records"), dbc.Alert(f"Loaded {len(df)} rows from {filename}", color="success")

@app.callback(
    Output("btn-run-custom", "disabled"),
    Input("custom-upload-store", "data"),
    Input("custom-upload-preset-select", "value"),
)
def toggle_run(data, presets):
    return not (data and (presets or []))

@app.callback(
    Output("custom-results", "children"),
    Output("custom-results-store", "data"),
    Input("btn-run-custom", "n_clicks"),
    State("custom-upload-store", "data"),
    State("custom-upload-preset-select", "value"),
    State("presets-store", "data"),
    prevent_initial_call=True
)
def run_custom(n, data_json, preset_names, presets_store):
    if not n:
        raise PreventUpdate
    if not data_json:
        raise PreventUpdate
    df = pd.DataFrame(json.loads(data_json))
    if df.empty:
        return dbc.Alert("No rows found in upload.", color="warning"), None

    # Data quality will be computed from the same slot-level detail dataframe used for results
    # (detail_first), to avoid mismatches.

    presets_store = presets_store or []
    name_to_settings = {p.get("name"): p.get("settings", {}) for p in presets_store if isinstance(p, dict)}
    selected = list(preset_names or [])
    if not selected:
        return dbc.Alert("Select at least one preset.", color="warning"), None

    combined_ps = []
    first_settings = None
    first_bands = None
    detail_first = None
    
    # Extract date range from uploaded data for Run Date Range column
    try:
        start_date = pd.to_datetime(df["date"]).min().strftime('%Y-%m-%d')
        end_date = pd.to_datetime(df["date"]).max().strftime('%Y-%m-%d')
    except Exception:
        start_date = ""
        end_date = ""

    for nm in selected:
        st = name_to_settings.get(nm)
        if not st:
            print(f"DEBUG - Preset '{nm}' not found; skipping")
            continue
        err_mode = str(st.get("err_mode", "default")).lower()
        mode_upper = MODE_DEFAULT if err_mode == "default" else MODE_DYNAMIC
        try:
            x_pct = float(st.get("x_pct", 50))
        except Exception:
            x_pct = 50.0
        dyn_x = (x_pct / 100.0) if mode_upper == MODE_DYNAMIC else 0.0
        bands_rows = st.get("bands", DEFAULT_BANDS.copy())
        bands_list, bands_table = parse_bands_from_settings(bands_rows)

        out_rows = []
        for slot in df.to_dict("records"):
            # Ensure numeric coercion
            slot_local = {
                **slot,
                "AvC_MW": float(pd.to_numeric(slot.get("AvC_MW", 0), errors="coerce") or 0),
                "Scheduled_MW": float(pd.to_numeric(slot.get("Scheduled_MW", 0), errors="coerce") or 0),
                "Actual_MW": float(pd.to_numeric(slot.get("Actual_MW", 0), errors="coerce") or 0),
                "PPA": float(pd.to_numeric(slot.get("PPA", 0), errors="coerce") or 0),
            }
            calc = compute_slot_row(slot_local, bands_list, mode_upper, dyn_x)
            out_rows.append({**slot_local, **calc})
        out_df = pd.DataFrame(out_rows)

        # Normalize plant_name column to avoid duplicates caused by whitespace/NBSP variations
        if "plant_name" in out_df.columns:
            out_df["plant_name"] = out_df["plant_name"].apply(lambda x: _norm_plant_name(str(x)) if pd.notna(x) else x)

        # Compute Data Availability % per plant before aggregation
        plant_data_avail = {}
        for (reg, plant), group_df in out_df.groupby(["region", "plant_name"]):
            plant_data_avail[(reg, plant)] = compute_plant_data_availability_pct(group_df)

        # Custom aggregation: PPA/AvC median only from "active" blocks (Schedule > 0)
        def safe_median_avc(series):
            """Median AvC from blocks where Scheduled_MW > 0"""
            group_df = series.reset_index(drop=True)
            parent_df = out_df.loc[series.index]
            active_mask = pd.to_numeric(parent_df.get("Scheduled_MW", 0), errors="coerce").fillna(0) > 0
            active_values = series[active_mask]
            return float(active_values.median()) if len(active_values) > 0 else 0.0
        
        def safe_median_ppa(series):
            """Median PPA from blocks where Scheduled_MW > 0 AND PPA > 0"""
            parent_df = out_df.loc[series.index]
            sch = pd.to_numeric(parent_df.get("Scheduled_MW", 0), errors="coerce").fillna(0)
            ppa_vals = pd.to_numeric(series, errors="coerce").fillna(0)
            active_mask = (sch > 0) & (ppa_vals > 0)
            active_values = series[active_mask]
            return float(active_values.median()) if len(active_values) > 0 else 0.0

        plant_summary = (
            out_df.groupby(["region","plant_name"], as_index=False)
            .agg({
                "PPA": safe_median_ppa,
                "AvC_MW": safe_median_avc,
                "Revenue_Loss":"sum",
                "Revenue_as_per_generation":"sum",
                "Total_DSM":"sum"
            })
            .rename(columns={"plant_name":"Plant name","AvC_MW":"Plant Capacity","Total_DSM":"DSM Loss"})
        )
        # Avoid division by zero
        plant_summary["Revenue Loss (%)"] = plant_summary.apply(
            lambda r: (float(r["Revenue_Loss"]) / float(r["Revenue_as_per_generation"])) * 100.0 if float(r["Revenue_as_per_generation"]) > 0 else 0.0,
            axis=1
        )
        plant_summary["Revenue Loss (p/k)"] = plant_summary["Revenue_Loss"] / 1000.0
        # Add Data Availability % per plant
        plant_summary["Data Availability %"] = plant_summary.apply(
            lambda r: plant_data_avail.get((r["region"], r["Plant name"]), 0.0),
            axis=1
        )
        
        # Add Run Date Range (user-selected) and Data Date Range (active power > 0)
        def format_date_range(start_str, end_str):
            """Format date range as 'DD-MMM-YYYY to DD-MMM-YYYY'"""
            try:
                start_dt = pd.to_datetime(start_str)
                end_dt = pd.to_datetime(end_str)
                return f"{start_dt.strftime('%d-%b-%Y')} to {end_dt.strftime('%d-%b-%Y')}"
            except Exception:
                return f"{start_str} to {end_str}"
        
        run_date_range_str = format_date_range(start_date, end_date)
        plant_summary["Run Date Range"] = run_date_range_str
        
        # Calculate Data Date Range per plant from FULL database (not filtered by user date range)
        data_date_ranges = {}
        for (region, plant_name), _ in out_df.groupby(["region", "plant_name"]):
            first_date, last_date = get_plant_full_active_date_range(plant_name, region)
            if first_date == "No Active Data":
                data_date_ranges[(region, plant_name)] = "No Active Data"
            else:
                data_date_ranges[(region, plant_name)] = f"{first_date} to {last_date}"
        
        plant_summary["Data Date Range"] = plant_summary.apply(
            lambda r: data_date_ranges.get((r["region"], r["Plant name"]), "No Active Data"),
            axis=1
        )
        
        plant_summary["Custom Setting"] = nm
        combined_ps.append(plant_summary)

        if first_settings is None:
            first_settings = {"err_mode": err_mode, "x_pct": x_pct}
            first_bands = bands_rows
            detail_first = out_df

    if not combined_ps:
        return dbc.Alert("No valid preset selected.", color="warning"), None

    combined_df = pd.concat(combined_ps, ignore_index=True) if len(combined_ps) > 1 else combined_ps[0]
    table = dash_table.DataTable(
        columns=[
            {"name": c, "id": c, "type": "numeric", "format": {"specifier": ".2f"}} 
            if c in ["PPA", "Data Availability %", "Revenue Loss (%)", "Revenue Loss (p/k)", "Plant Capacity", "DSM Loss"]
            else {"name": c, "id": c}
            for c in combined_df.columns
        ],
        data=combined_df.to_dict("records"),
        sort_action="native",
        style_table={"overflowX":"auto"},
        style_header={"fontWeight":"600"},
    )
    stored = {
        "df": detail_first.to_dict("records") if isinstance(detail_first, pd.DataFrame) else [],
        "final_bands": first_bands or [],
        "err_mode": (first_settings or {}).get("err_mode", "default"),
        "x_pct": (first_settings or {}).get("x_pct", 50.0),
        "data_quality": compute_data_quality_stats(detail_first) if isinstance(detail_first, pd.DataFrame) and not detail_first.empty else {},
    }
    # Render Data Quality section above the plant summary table
    dq = stored.get("data_quality", {}) or {}
    dq_card = None
    dq_pie_chart = None
    dq_table = None
    try:
        if dq and dq.get("total_ts", 0):
            dq_card = dbc.Card(
                [
                    dbc.CardHeader("Data Quality Stats"),
                    dbc.CardBody(
                        [
                            html.Div(f"Total slots checked: {dq.get('total_ts', 0)}", className="text-muted"),
                            html.Div(f"Data Available: {dq.get('available_pct', 0)}%  (count: {dq.get('available_count', 0)})"),
                            html.Div(f"Actual Power Missing: {dq.get('actual_missing_pct', 0)}%  (count: {dq.get('actual_missing_count', 0)})"),
                            html.Div(f"PPA / AVC Missing: {dq.get('ppa_avc_missing_pct', 0)}%  (count: {dq.get('ppa_avc_missing_count', 0)})"),
                            html.Div(
                                f"DSM not calculated (zero Schedule/AVC/PPA): {dq.get('dsm_skipped_zero_inputs_pct', 0)}%  "
                                f"(count: {dq.get('dsm_skipped_zero_inputs_count', 0)})"
                            ),
                        ]
                    ),
                ],
                className="mb-3",
            )
            
            # Data Quality Pie Chart (always sums to 100%)
            available_pct = dq.get('available_pct', 0.0)
            actual_missing_pct = dq.get('actual_missing_pct', 0.0)
            ppa_avc_missing_pct = dq.get('ppa_avc_missing_pct', 0.0)
            unknown_pct = round(100.0 - (available_pct + actual_missing_pct + ppa_avc_missing_pct), 2)
            
            pie_data = {
                "Category": ["Data Available", "Actual Power Missing", "PPA / AVC Missing", "Unknown"],
                "Percentage": [available_pct, actual_missing_pct, ppa_avc_missing_pct, unknown_pct]
            }
            pie_df = pd.DataFrame(pie_data)
            
            fig_pie = px.pie(
                pie_df,
                values="Percentage",
                names="Category",
                title="Data Quality Distribution",
                color="Category",
                color_discrete_map={
                    "Data Available": "#28a745",
                    "Actual Power Missing": "#dc3545",
                    "PPA / AVC Missing": "#ffc107",
                    "Unknown": "#6c757d"
                },
                hole=0.4  # Donut chart
            )
            fig_pie.update_traces(
                textposition="inside",
                textinfo="percent+label",
                hovertemplate="<b>%{label}</b><br>Percentage: %{percent}<br>Value: %{value:.2f}%<extra></extra>"
            )
            fig_pie.update_layout(
                template="plotly_white",
                height=400,
                showlegend=True,
                legend=dict(orientation="v", yanchor="middle", y=0.5, xanchor="left", x=1.05)
            )
            
            dq_pie_chart = dbc.Card(
                [
                    dbc.CardHeader("Data Quality Pie Chart"),
                    dbc.CardBody(
                        dcc.Graph(figure=fig_pie, config={"displayModeBar": True})
                    ),
                ],
                className="mb-3",
            )
            
            issues = dq.get("issues_rows", []) if isinstance(dq.get("issues_rows", []), list) else []
            issues_total = int(dq.get("issues_total", len(issues) if issues else 0) or 0)
            if issues_total > 0:
                show_n = len(issues)
                dq_table = html.Details(
                    [
                        html.Summary(f"Data Quality Issues (showing {show_n} of {issues_total})"),
                        html.Div(
                            dash_table.DataTable(
                                columns=[{"name": c, "id": c} for c in (issues[0].keys() if issues else [])],
                                data=issues,
                                sort_action="native",
                                filter_action="native",
                                page_size=15,
                                style_table={"overflowX": "auto"},
                                style_cell={"padding": "6px", "textAlign": "left"},
                                style_header={"fontWeight": "600", "backgroundColor": "#f8f9fa"},
                            ),
                            className="mt-2",
                        ),
                    ],
                    className="mb-3",
                )
    except Exception:
        dq_card = None
        dq_pie_chart = None
        dq_table = None

    children = []
    if dq_card is not None:
        children.append(dq_card)
    if dq_pie_chart is not None:
        children.append(dq_pie_chart)
    if dq_table is not None:
        children.append(dq_table)

    # Band-wise DSM section (per-plant)
    children.append(bandwise_section_layout("custom"))

    children.append(table)
    return html.Div(children), stored

@app.callback(
    Output("download-custom-output", "data"),
    Output("dl-done-store", "data", allow_duplicate=True),
    Input("download-custom-output-btn", "n_clicks"),
    State("custom-results-store", "data"),
    State("dl-done-store", "data"),
    prevent_initial_call=True
)
def download_custom(n, stored, _done):
    if not n:
        raise PreventUpdate
    if not stored:
        raise PreventUpdate
    try:
        df_main = pd.DataFrame(stored.get("df", []))
    except Exception:
        df_main = pd.DataFrame()
    if df_main.empty:
        raise PreventUpdate

    err_mode = str(stored.get("err_mode", "default")).lower()
    mode_upper = MODE_DEFAULT if err_mode == "default" else MODE_DYNAMIC
    try:
        x_pct = float(stored.get("x_pct", 50))
    except Exception:
        x_pct = 50.0
    dyn_x = (x_pct / 100.0) if mode_upper == MODE_DYNAMIC else 0.0

    bands_rows = stored.get("final_bands", [])
    bands_list, bands_table = parse_bands_from_settings(bands_rows)

    base_cols = [
        "region","plant_name","date","time_block","from_time","to_time",
        "AvC_MW","Scheduled_MW","Actual_MW","PPA"
    ]
    missing = [c for c in base_cols if c not in df_main.columns]
    if missing:
        df_exp = df_main.copy()
        if "plant_name" not in df_exp.columns and "Plant" in df_exp.columns:
            df_exp["plant_name"] = df_exp["Plant"]
        if "time_block" not in df_exp.columns and "block" in df_exp.columns:
            df_exp["time_block"] = df_exp["block"]
        missing2 = [c for c in base_cols if c not in df_exp.columns]
        if missing2:
            raise PreventUpdate
        detail_for_excel = df_exp[base_cols].copy()
    else:
        detail_for_excel = df_main[base_cols].copy()

    # Calculate per-slot metrics as in main export
    detail_calculated_rows = []
    for _, row in detail_for_excel.iterrows():
        slot = {
            "region": row["region"],
            "plant_name": row["plant_name"],
            "date": row["date"],
            "time_block": row["time_block"],
            "from_time": row["from_time"],
            "to_time": row["to_time"],
            "AvC_MW": float(row["AvC_MW"]),
            "Scheduled_MW": float(row["Scheduled_MW"]),
            "Actual_MW": float(row["Actual_MW"]),
            "PPA": float(row["PPA"]),
        }
        calc = compute_slot_row(slot, bands_list, mode_upper, dyn_x)
        detail_calculated_rows.append({**slot, **calc})
    detail_calculated_df = pd.DataFrame(detail_calculated_rows)

    # Build workbook (mirrors main export structure)
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.utils import column_index_from_string
    from openpyxl.workbook.defined_name import DefinedName

    buf = BytesIO()
    wb = Workbook()
    wb.remove(wb.active)

    # Config
    ws_config = wb.create_sheet("Config")
    ws_config.cell(row=1, column=1, value="Key")
    ws_config.cell(row=1, column=2, value="Value")
    ws_config.cell(row=2, column=1, value="MODE")
    ws_config.cell(row=2, column=2, value=mode_upper)
    ws_config.cell(row=3, column=1, value="DYN_X")
    ws_config.cell(row=3, column=2, value=float(dyn_x))
    try:
        wb.defined_names.append(DefinedName(name="CFG_MODE", attr_text="Config!$B$2"))
        wb.defined_names.append(DefinedName(name="CFG_DYNX", attr_text="Config!$B$3"))
    except Exception:
        pass

    # Bands
    ws_bands = wb.create_sheet("Bands")
    bands_cols = ["direction","lower_pct","upper_pct","rate_type","rate_value","rate_slope","loss_zone"]
    for col_idx, col_name in enumerate(bands_cols, start=1):
        ws_bands.cell(row=1, column=col_idx, value=col_name)
    for row_idx, row in enumerate(bands_table.itertuples(index=False), start=2):
        ws_bands.cell(row=row_idx, column=1, value=str(getattr(row, "direction", "")))
        ws_bands.cell(row=row_idx, column=2, value=float(getattr(row, "lower_pct", 0)))
        ws_bands.cell(row=row_idx, column=3, value=float(getattr(row, "upper_pct", 0)))
        ws_bands.cell(row=row_idx, column=4, value=str(getattr(row, "rate_type", "")))
        ws_bands.cell(row=row_idx, column=5, value=float(getattr(row, "rate_value", 0)))
        ws_bands.cell(row=row_idx, column=6, value=float(getattr(row, "rate_slope", 0)))
        ws_bands.cell(row=row_idx, column=7, value=bool(getattr(row, "loss_zone", False)))
    n_bands = max(len(bands_table), 1)
    bands_end_ref = 200 if n_bands < 199 else (1 + n_bands)
    try:
        wb.defined_names.append(DefinedName(name="Bands_Dir", attr_text=f"Bands!$A$2:$A${bands_end_ref}"))
        wb.defined_names.append(DefinedName(name="Bands_Lower", attr_text=f"Bands!$B$2:$B${bands_end_ref}"))
        wb.defined_names.append(DefinedName(name="Bands_Upper", attr_text=f"Bands!$C$2:$C${bands_end_ref}"))
        wb.defined_names.append(DefinedName(name="Bands_RateType", attr_text=f"Bands!$D$2:$D${bands_end_ref}"))
        wb.defined_names.append(DefinedName(name="Bands_RateVal", attr_text=f"Bands!$E$2:$E${bands_end_ref}"))
        wb.defined_names.append(DefinedName(name="Bands_RateSlope", attr_text=f"Bands!$F$2:$F${bands_end_ref}"))
        wb.defined_names.append(DefinedName(name="Bands_LossZone", attr_text=f"Bands!$G$2:$G${bands_end_ref}"))
    except Exception:
        pass

    # Detail
    ws_detail = wb.create_sheet("Detail")
    detail_headers = {
        'A': 'Region',
        'B': 'Plant Name',
        'C': 'Date',
        'D': 'Block',
        'E': 'From Time',
        'F': 'To Time',
        'G': 'Schedule Power (MW)',
        'H': 'AvC (MW)',
        'I': 'Injected Power (MW)',
        'J': 'PPA or MCP',
        'K': 'Error %',
        'L': 'Absolute error %',
        'M': 'Direction',
        'N': 'Deviation (MW)',
        'U': 'Revenue as per Generation (INR)',
        'V': 'Scheduled Revenue (INR)',
        'AE': '_basis'
    }
    for col_letter, header in detail_headers.items():
        col_idx = column_index_from_string(col_letter)
        ws_detail.cell(row=1, column=col_idx, value=header)
    ws_detail.column_dimensions['AE'].hidden = True

    n_rows = len(detail_for_excel)
    for row_idx, row_data in enumerate(detail_for_excel.itertuples(index=False), start=2):
        calc_row = detail_calculated_df.iloc[row_idx - 2] if row_idx - 2 < len(detail_calculated_df) else {}
        ws_detail.cell(row=row_idx, column=1, value=getattr(row_data, "region", ""))
        ws_detail.cell(row=row_idx, column=2, value=getattr(row_data, "plant_name", ""))
        ws_detail.cell(row=row_idx, column=3, value=getattr(row_data, "date", ""))
        ws_detail.cell(row=row_idx, column=4, value=getattr(row_data, "time_block", ""))
        ws_detail.cell(row=row_idx, column=5, value=getattr(row_data, "from_time", ""))
        ws_detail.cell(row=row_idx, column=6, value=getattr(row_data, "to_time", ""))
        ws_detail.cell(row=row_idx, column=7, value=float(getattr(row_data, "Scheduled_MW", 0)))
        ws_detail.cell(row=row_idx, column=8, value=float(getattr(row_data, "AvC_MW", 0)))
        ws_detail.cell(row=row_idx, column=9, value=float(getattr(row_data, "Actual_MW", 0)))
        ws_detail.cell(row=row_idx, column=10, value=float(getattr(row_data, "PPA", 0)))
        basis = denominator_and_basis(float(getattr(row_data, "AvC_MW", 0)), float(getattr(row_data, "Scheduled_MW", 0)), mode_upper, dyn_x)
        ws_detail.cell(row=row_idx, column=31, value=float(basis))
        ws_detail.cell(row=row_idx, column=11, value=float(calc_row.get("error_pct", 0.0)))
        ws_detail.cell(row=row_idx, column=12, value=float(calc_row.get("abs_err", 0.0)))
        ws_detail.cell(row=row_idx, column=13, value=str(calc_row.get("direction", "")))
        deviation = float(getattr(row_data, "Actual_MW", 0)) - float(getattr(row_data, "Scheduled_MW", 0))
        ws_detail.cell(row=row_idx, column=14, value=float(deviation))
        ws_detail.cell(row=row_idx, column=21, value=float(calc_row.get("Revenue_as_per_generation", 0.0)))
        ws_detail.cell(row=row_idx, column=22, value=float(calc_row.get("Scheduled_Revenue_as_per_generation", 0.0)))

    # ===== BAND-WISE ENERGY DEVIATION COLUMNS (O-T = columns 15-20) =====
    ui_bands = sorted([b for b in bands_list if b.direction == "UI"], key=lambda x: x.lower_pct)
    oi_bands = sorted([b for b in bands_list if b.direction == "OI"], key=lambda x: x.lower_pct)
    all_bands_for_energy = (ui_bands + oi_bands)[:6]

    n_rows = len(detail_for_excel)
    for col_idx, band in enumerate(all_bands_for_energy, start=15):
        dir_label = "UI" if band.direction == "UI" else "OI"
        if band.upper_pct >= 999:
            header = f"{dir_label} Energy >{int(band.lower_pct)}% (MW)"
        else:
            header = f"{dir_label} Energy {int(band.lower_pct)}-{int(band.upper_pct)}% (MW)"
        ws_detail.cell(row=1, column=col_idx, value=header)

        for data_row in range(2, n_rows + 2):
            calc_row = detail_calculated_df.iloc[data_row - 2] if data_row - 2 < len(detail_calculated_df) else {}
            abs_err = calc_row.get("abs_err", 0.0)
            direction = calc_row.get("direction", "")
            avc = detail_for_excel.iloc[data_row - 2]["AvC_MW"]
            sch = detail_for_excel.iloc[data_row - 2]["Scheduled_MW"]
            act = detail_for_excel.iloc[data_row - 2]["Actual_MW"]
            basis = denominator_and_basis(avc, sch, mode_upper, dyn_x)
            deviation = act - sch
            energy_mw = 0.0
            if direction == dir_label:
                sp = slice_pct(abs_err, band.lower_pct, band.upper_pct)
                if sp > 0:
                    sign = -1.0 if deviation < 0 else (1.0 if deviation > 0 else 0.0)
                    energy_mw = sign * sp / 100.0 * basis
            ws_detail.cell(row=data_row, column=col_idx, value=float(energy_mw))

    # ===== PER-BAND DSM COLUMNS (W+ = starting column 23) =====
    current_col = 23

    # UI DSM columns
    for band in ui_bands:
        if band.upper_pct >= 999:
            header = f"UI >{int(band.lower_pct)}% DSM"
        else:
            header = f"UI {int(band.lower_pct)}-{int(band.upper_pct)}% DSM"
        ws_detail.cell(row=1, column=current_col, value=header)
        for data_row in range(2, n_rows + 2):
            calc_row = detail_calculated_df.iloc[data_row - 2] if data_row - 2 < len(detail_calculated_df) else {}
            abs_err = calc_row.get("abs_err", 0.0)
            direction = calc_row.get("direction", "")
            avc = detail_for_excel.iloc[data_row - 2]["AvC_MW"]
            sch = detail_for_excel.iloc[data_row - 2]["Scheduled_MW"]
            ppa = detail_for_excel.iloc[data_row - 2]["PPA"]
            basis = denominator_and_basis(avc, sch, mode_upper, dyn_x)
            dsm_value = 0.0
            if direction == "UI":
                sp = slice_pct(abs_err, band.lower_pct, band.upper_pct)
                if sp > 0:
                    kwh = kwh_from_slice(sp, basis)
                    rate = band_rate(ppa, band.rate_type, band.rate_value, band.rate_slope, abs_err)
                    dsm_value = kwh * rate
            ws_detail.cell(row=data_row, column=current_col, value=float(dsm_value))
        current_col += 1

    # OI DSM columns (non-loss-zone)
    for band in oi_bands:
        if not band.loss_zone:
            if band.upper_pct >= 999:
                header = f"OI >{int(band.lower_pct)}% DSM"
            else:
                header = f"OI {int(band.lower_pct)}-{int(band.upper_pct)}% DSM"
            ws_detail.cell(row=1, column=current_col, value=header)
            for data_row in range(2, n_rows + 2):
                calc_row = detail_calculated_df.iloc[data_row - 2] if data_row - 2 < len(detail_calculated_df) else {}
                abs_err = calc_row.get("abs_err", 0.0)
                direction = calc_row.get("direction", "")
                avc = detail_for_excel.iloc[data_row - 2]["AvC_MW"]
                sch = detail_for_excel.iloc[data_row - 2]["Scheduled_MW"]
                ppa = detail_for_excel.iloc[data_row - 2]["PPA"]
                basis = denominator_and_basis(avc, sch, mode_upper, dyn_x)
                dsm_value = 0.0
                if direction == "OI":
                    sp = slice_pct(abs_err, band.lower_pct, band.upper_pct)
                    if sp > 0:
                        kwh = kwh_from_slice(sp, basis)
                        rate = band_rate(ppa, band.rate_type, band.rate_value, band.rate_slope, abs_err)
                        dsm_value = kwh * rate
                ws_detail.cell(row=data_row, column=current_col, value=float(dsm_value))
            current_col += 1

    # OI Loss columns (loss-zone only)
    for band in oi_bands:
        if band.loss_zone:
            if band.upper_pct >= 999:
                header = f"OI >{int(band.lower_pct)}% Loss"
            else:
                header = f"OI {int(band.lower_pct)}-{int(band.upper_pct)}% Loss"
            ws_detail.cell(row=1, column=current_col, value=header)
            for data_row in range(2, n_rows + 2):
                calc_row = detail_calculated_df.iloc[data_row - 2] if data_row - 2 < len(detail_calculated_df) else {}
                abs_err = calc_row.get("abs_err", 0.0)
                direction = calc_row.get("direction", "")
                avc = detail_for_excel.iloc[data_row - 2]["AvC_MW"]
                sch = detail_for_excel.iloc[data_row - 2]["Scheduled_MW"]
                ppa = detail_for_excel.iloc[data_row - 2]["PPA"]
                basis = denominator_and_basis(avc, sch, mode_upper, dyn_x)
                loss_value = 0.0
                if direction == "OI":
                    sp = slice_pct(abs_err, band.lower_pct, band.upper_pct)
                    if sp > 0:
                        kwh = kwh_from_slice(sp, basis)
                        rate = band_rate(ppa, band.rate_type, band.rate_value, band.rate_slope, abs_err)
                        loss_value = kwh * rate
                ws_detail.cell(row=data_row, column=current_col, value=float(loss_value))
            current_col += 1

    # ===== TOTALS (AC=col 29, AD=col 30) =====
    ws_detail.cell(row=1, column=29, value='Total DSM (INR)')
    ws_detail.cell(row=1, column=30, value='Revenue Loss (INR)')
    for data_row in range(2, n_rows + 2):
        calc_row = detail_calculated_df.iloc[data_row - 2]
        ws_detail.cell(row=data_row, column=29, value=float(calc_row.get("Total_DSM", 0.0)))
        ws_detail.cell(row=data_row, column=30, value=float(calc_row.get("Revenue_Loss", 0.0)))

    wb.save(buf)
    buf.seek(0)
    return dcc.send_bytes(lambda x: x.write(buf.read()), "custom_analysis.xlsx"), (_done or 0) + 1

@app.callback(
    Output("download-excel", "data"),
    Output("dl-done-store", "data", allow_duplicate=True),
    Input("btn-download", "n_clicks"),
    State("results-store", "data"),
    State("dl-done-store", "data"),
    prevent_initial_call=True
)
def download_full(n, stored, _done):
    # Only download if button was actually clicked
    if not n or n == 0:
        raise PreventUpdate
    if not stored or stored.get("error"):
        raise PreventUpdate

    # Multi-preset export path (only when >1 runs; single-run should use the detailed export below)
    all_runs = stored.get("_all_runs") if isinstance(stored, dict) else None
    if all_runs and isinstance(all_runs, list) and len(all_runs) > 1:
        from io import BytesIO
        output = BytesIO()
        with pd.ExcelWriter(output, engine="xlsxwriter") as xw:
            # Combined Plant Summary
            try:
                plant_summary_df = pd.DataFrame(stored.get("plant_summary", []))
                if not plant_summary_df.empty:
                    plant_summary_df.to_excel(xw, sheet_name="Plant_Summary", index=False)
            except Exception as e:
                print(f"DEBUG - Failed writing Plant_Summary: {e}")

            # One Detail/Config/Bands per preset
            for r in all_runs:
                nm = str(r.get("name", "Preset"))
                # Load full df from cache if present (large runs)
                df_detail = _detail_df_from_store(r if isinstance(r, dict) else {})
                bands_df = _normalize_bands_df(pd.DataFrame(r.get("final_bands", [])))
                # Detail
                try:
                    df_detail.to_excel(xw, sheet_name=f"Detail_{nm}", index=False)
                except Exception as e:
                    print(f"DEBUG - Writing Detail_{nm} failed: {e}")
                # Config
                try:
                    pd.DataFrame({"Key":["MODE","DYN_X"], "Value":[r.get("err_mode"), r.get("x_pct")]}) \
                        .to_excel(xw, sheet_name=f"Config_{nm}", index=False)
                except Exception as e:
                    print(f"DEBUG - Writing Config_{nm} failed: {e}")
                # Bands
                try:
                    bands_df.to_excel(xw, sheet_name=f"Bands_{nm}", index=False)
                except Exception as e:
                    print(f"DEBUG - Writing Bands_{nm} failed: {e}")
        return dcc.send_bytes(output.getvalue(), filename="DSM_Full_Calculation_MultiPresets.xlsx"), (_done or 0) + 1

    df_main = _detail_df_from_store(stored)
    if df_main.empty:
        raise PreventUpdate

    used = stored.get("used_settings", {}) if isinstance(stored, dict) else {}
    err_mode = str(used.get("err_mode", "default")).lower()
    mode_upper = MODE_DEFAULT if err_mode == "default" else MODE_DYNAMIC
    try:
        x_pct = float(used.get("x_pct", 50))
    except Exception:
        x_pct = 50.0
    dyn_x = (x_pct / 100.0) if mode_upper == MODE_DYNAMIC else 0.0

    bands_rows = stored.get("final_bands", []) if isinstance(stored, dict) else []
    bands_list, bands_table = parse_bands_from_settings(bands_rows)

    # Select input columns for formula-driven export
    base_cols = [
        "region","plant_name","date","time_block","from_time","to_time",
        "AvC_MW","Scheduled_MW","Actual_MW","PPA"
    ]
    missing = [c for c in base_cols if c not in df_main.columns]
    if missing:
        # Try to map Plant->plant_name and block->time_block if present
        df_exp = df_main.copy()
        if "plant_name" not in df_exp.columns and "Plant" in df_exp.columns:
            df_exp["plant_name"] = df_exp["Plant"]
        if "time_block" not in df_exp.columns and "block" in df_exp.columns:
            df_exp["time_block"] = df_exp["block"]
        missing2 = [c for c in base_cols if c not in df_exp.columns]
        if missing2:
            raise PreventUpdate
        detail_for_excel = df_exp[base_cols].copy()
    else:
        detail_for_excel = df_main[base_cols].copy()

    # Pre-calculate all slot values and build complete detail with per-band columns
    detail_calculated_rows = []
    for _, row in detail_for_excel.iterrows():
        slot = {
            "region": row["region"],
            "plant_name": row["plant_name"],
            "date": row["date"],
            "time_block": row["time_block"],
            "from_time": row["from_time"],
            "to_time": row["to_time"],
            "AvC_MW": float(row["AvC_MW"]),
            "Scheduled_MW": float(row["Scheduled_MW"]),
            "Actual_MW": float(row["Actual_MW"]),
            "PPA": float(row["PPA"]),
        }
        calc = compute_slot_row(slot, bands_list, mode_upper, dyn_x)
        # Merge input + calculated
        detail_calculated_rows.append({**slot, **calc})
    detail_calculated_df = pd.DataFrame(detail_calculated_rows)

    # Build per-band columns: one column per band showing deviation energy (MW) AND DSM amount (INR)
    per_band_cols = {}
    per_band_dsm_cols = {}
    for band in bands_list:
        dir_label = "UI" if band.direction == "UI" else "OI"
        if band.upper_pct >= 999:
            band_label = f"{dir_label} Energy deviation >{int(band.lower_pct)}%"
            dsm_label = f"{dir_label} DSM due to Deviation >{int(band.lower_pct)}% (INR)"
        else:
            band_label = f"{dir_label} Energy deviation between {int(band.lower_pct)}-{int(band.upper_pct)}%"
            dsm_label = f"{dir_label} DSM between {int(band.lower_pct)}-{int(band.upper_pct)}% (INR)"
        per_band_cols[band_label] = []
        per_band_dsm_cols[dsm_label] = []

    # Calculate per-band deviation and DSM for each row
    for _, row in detail_calculated_df.iterrows():
        abs_err = row.get("abs_err", 0.0)
        direction = row.get("direction", "")
        avc = row.get("AvC_MW", 0.0)
        sch = row.get("Scheduled_MW", 0.0)
        ppa = row.get("PPA", 0.0)
        denom = denominator_and_basis(avc, sch, mode_upper, dyn_x)

        for band_idx, band in enumerate(bands_list):
            energy_col_name = list(per_band_cols.keys())[band_idx]
            dsm_col_name = list(per_band_dsm_cols.keys())[band_idx]
            if band.direction != direction:
                per_band_cols[energy_col_name].append(0.0)
                per_band_dsm_cols[dsm_col_name].append(0.0)
                continue
            sp = slice_pct(abs_err, band.lower_pct, band.upper_pct)
            if sp > 0:
                kwh = kwh_from_slice(sp, denom)
                rate = band_rate(ppa, band.rate_type, band.rate_value, band.rate_slope, abs_err)
                amount = kwh * rate
                per_band_cols[energy_col_name].append(kwh / 1000.0)  # Convert to MW for display
                per_band_dsm_cols[dsm_col_name].append(amount)
            else:
                per_band_cols[energy_col_name].append(0.0)
                per_band_dsm_cols[dsm_col_name].append(0.0)

    # Add per-band columns to detail dataframe
    for col_name, values in per_band_cols.items():
        detail_calculated_df[col_name] = values
    for col_name, values in per_band_dsm_cols.items():
        detail_calculated_df[col_name] = values

    # Reorder columns: base columns, then calculated standard columns, then per-band columns (energy + DSM)
    # Removed generic UI_Energy_deviation_bands and OI_Energy_deviation_bands - replaced with per-band columns
    standard_calc_cols = [
        "error_pct", "direction", "abs_err", "band_level",
        "Revenue_as_per_generation", "Scheduled_Revenue_as_per_generation",
        "UI_DSM", "OI_DSM", "OI_Loss", "Total_DSM", "Revenue_Loss"
    ]
    # Interleave per-band columns: energy then DSM for each band
    # Sort by direction and lower_pct for proper ordering
    def get_band_sort_key(col_name):
        # Extract direction and range for sorting
        if "UI" in col_name:
            dir_val = 0
        elif "OI" in col_name:
            dir_val = 1
        else:
            dir_val = 2
        # Extract lower percentage
        import re
        nums = re.findall(r'\d+', col_name)
        lower_val = int(nums[0]) if nums else 9999
        return (dir_val, lower_val)
    
    per_band_col_names = sorted(list(per_band_cols.keys()), key=get_band_sort_key)
    per_band_dsm_col_names = sorted(list(per_band_dsm_cols.keys()), key=get_band_sort_key)
    # Group by direction and range to interleave properly
    interleaved_per_band = []
    for energy_col in per_band_col_names:
        interleaved_per_band.append(energy_col)
        # Find corresponding DSM column by matching the range
        # Extract the range portion (e.g., "0-10%" or ">20%")
        import re
        energy_range = re.search(r'(\d+-?\d*%|>\d+%)', energy_col)
        if energy_range:
            range_str = energy_range.group(1)
            for dsm_col in per_band_dsm_col_names:
                if range_str in dsm_col and energy_col.split()[0] in dsm_col:  # Same direction and range
                    interleaved_per_band.append(dsm_col)
                    break
    final_cols = base_cols + [c for c in standard_calc_cols if c in detail_calculated_df.columns] + interleaved_per_band
    detail_calculated_df = detail_calculated_df[[c for c in final_cols if c in detail_calculated_df.columns]]

    # Export with Office 365 dynamic array formulas
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter, column_index_from_string
    
    buf = BytesIO()
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # ===== CONFIG SHEET =====
    ws_config = wb.create_sheet("Config")
    ws_config.cell(row=1, column=1, value="Key")
    ws_config.cell(row=1, column=2, value="Value")
    ws_config.cell(row=2, column=1, value="MODE")
    ws_config.cell(row=2, column=2, value=mode_upper)
    ws_config.cell(row=3, column=1, value="DYN_X")
    ws_config.cell(row=3, column=2, value=float(dyn_x))
    
    # Named cells for Config
    from openpyxl.workbook.defined_name import DefinedName
    try:
        wb.defined_names.append(DefinedName(name="CFG_MODE", attr_text="Config!$B$2"))
        wb.defined_names.append(DefinedName(name="CFG_DYNX", attr_text="Config!$B$3"))
    except Exception:
        pass
    
    # ===== BANDS SHEET =====
    ws_bands = wb.create_sheet("Bands")
    bands_cols = ["direction","lower_pct","upper_pct","rate_type","rate_value","rate_slope","loss_zone"]
    for col_idx, col_name in enumerate(bands_cols, start=1):
        ws_bands.cell(row=1, column=col_idx, value=col_name)
    
    for row_idx, row in enumerate(bands_table.itertuples(index=False), start=2):
        ws_bands.cell(row=row_idx, column=1, value=str(getattr(row, "direction", "")))
        ws_bands.cell(row=row_idx, column=2, value=float(getattr(row, "lower_pct", 0)))
        ws_bands.cell(row=row_idx, column=3, value=float(getattr(row, "upper_pct", 0)))
        ws_bands.cell(row=row_idx, column=4, value=str(getattr(row, "rate_type", "")))
        ws_bands.cell(row=row_idx, column=5, value=float(getattr(row, "rate_value", 0)))
        ws_bands.cell(row=row_idx, column=6, value=float(getattr(row, "rate_slope", 0)))
        ws_bands.cell(row=row_idx, column=7, value=bool(getattr(row, "loss_zone", False)))
    
    n_bands = max(len(bands_table), 1)
    bands_end_row = 1 + n_bands
    # Named ranges with headroom (2:200)
    bands_end_ref = 200 if n_bands < 199 else bands_end_row
    
    try:
        wb.defined_names.append(DefinedName(name="Bands_Dir", attr_text=f"Bands!$A$2:$A${bands_end_ref}"))
        wb.defined_names.append(DefinedName(name="Bands_Lower", attr_text=f"Bands!$B$2:$B${bands_end_ref}"))
        wb.defined_names.append(DefinedName(name="Bands_Upper", attr_text=f"Bands!$C$2:$C${bands_end_ref}"))
        wb.defined_names.append(DefinedName(name="Bands_RateType", attr_text=f"Bands!$D$2:$D${bands_end_ref}"))
        wb.defined_names.append(DefinedName(name="Bands_RateVal", attr_text=f"Bands!$E$2:$E${bands_end_ref}"))
        wb.defined_names.append(DefinedName(name="Bands_RateSlope", attr_text=f"Bands!$F$2:$F${bands_end_ref}"))
        wb.defined_names.append(DefinedName(name="Bands_LossZone", attr_text=f"Bands!$G$2:$G${bands_end_ref}"))
    except Exception:
        pass
    
    # ===== DETAIL SHEET =====
    ws_detail = wb.create_sheet("Detail")
    
    # Column mapping per spec
    # A=Region, B=Plant Name, C=Date, D=Block, E=From Time, F=To Time,
    # G=Schedule Power (MW), H=AvC (MW), I=Injected Power (MW), J=PPA or MCP,
    # K=Error %, L=Absolute error %, M=Direction, N=Deviation (MW),
    # U=Revenue as per Generation, V=Scheduled Revenue, AE=_basis
    detail_headers = {
        'A': 'Region',
        'B': 'Plant Name',
        'C': 'Date',
        'D': 'Block',
        'E': 'From Time',
        'F': 'To Time',
        'G': 'Schedule Power (MW)',
        'H': 'AvC (MW)',
        'I': 'Injected Power (MW)',
        'J': 'PPA or MCP',
        'K': 'Error %',
        'L': 'Absolute error %',
        'M': 'Direction',
        'N': 'Deviation (MW)',
        'U': 'Revenue as per Generation (INR)',
        'V': 'Scheduled Revenue (INR)',
        'AE': '_basis'
    }
    
    # Write headers - use column_index_from_string to handle multi-letter columns
    for col_letter, header in detail_headers.items():
        col_idx = column_index_from_string(col_letter)
        ws_detail.cell(row=1, column=col_idx, value=header)
    
    # Hide _basis column
    ws_detail.column_dimensions['AE'].hidden = True
    
    # Write data rows with calculated values (no formulas)
    # Excel limit: 1,048,576 rows total (1 header + 1,048,575 data rows)
    MAX_EXCEL_DATA_ROWS = 1_048_575
    n_rows = len(detail_for_excel)
    
    # Check if data exceeds Excel row limit - split into multiple sheets if needed
    if n_rows > MAX_EXCEL_DATA_ROWS:
        # Split data into chunks
        num_sheets = (n_rows + MAX_EXCEL_DATA_ROWS - 1) // MAX_EXCEL_DATA_ROWS  # Ceiling division
        print(f"DEBUG - Data exceeds Excel limit ({n_rows} rows). Splitting into {num_sheets} sheets.")
        
        for sheet_num in range(num_sheets):
            start_idx = sheet_num * MAX_EXCEL_DATA_ROWS
            end_idx = min(start_idx + MAX_EXCEL_DATA_ROWS, n_rows)
            chunk_df = detail_for_excel.iloc[start_idx:end_idx]
            chunk_calc_df = detail_calculated_df.iloc[start_idx:end_idx]
            
            # Create new sheet for this chunk
            sheet_name = f"Detail_{sheet_num + 1}" if sheet_num > 0 else "Detail"
            ws_detail_chunk = wb.create_sheet(sheet_name) if sheet_num > 0 else ws_detail
            
            # Write headers
            for col_letter, header in detail_headers.items():
                col_idx = column_index_from_string(col_letter)
                ws_detail_chunk.cell(row=1, column=col_idx, value=header)
            
            # Hide _basis column
            ws_detail_chunk.column_dimensions['AE'].hidden = True
            
            # Write chunk data rows
            chunk_n_rows = len(chunk_df)
            for local_idx, row_data in enumerate(chunk_df.itertuples(index=False), start=0):
                row_idx = local_idx + 2  # Start from row 2 (row 1 is header)
                calc_row = chunk_calc_df.iloc[local_idx] if local_idx < len(chunk_calc_df) else {}
                
                # Input values
                ws_detail_chunk.cell(row=row_idx, column=1, value=getattr(row_data, "region", ""))
                ws_detail_chunk.cell(row=row_idx, column=2, value=getattr(row_data, "plant_name", ""))
                ws_detail_chunk.cell(row=row_idx, column=3, value=getattr(row_data, "date", ""))
                ws_detail_chunk.cell(row=row_idx, column=4, value=getattr(row_data, "time_block", ""))
                ws_detail_chunk.cell(row=row_idx, column=5, value=getattr(row_data, "from_time", ""))
                ws_detail_chunk.cell(row=row_idx, column=6, value=getattr(row_data, "to_time", ""))
                ws_detail_chunk.cell(row=row_idx, column=7, value=float(getattr(row_data, "Scheduled_MW", 0)))
                ws_detail_chunk.cell(row=row_idx, column=8, value=float(getattr(row_data, "AvC_MW", 0)))
                ws_detail_chunk.cell(row=row_idx, column=9, value=float(getattr(row_data, "Actual_MW", 0)))
                ws_detail_chunk.cell(row=row_idx, column=10, value=float(getattr(row_data, "PPA", 0)))
                
                # Calculated values
                avc = float(getattr(row_data, "AvC_MW", 0))
                sch = float(getattr(row_data, "Scheduled_MW", 0))
                act = float(getattr(row_data, "Actual_MW", 0))
                ppa = float(getattr(row_data, "PPA", 0))
                
                basis = denominator_and_basis(avc, sch, mode_upper, dyn_x)
                ws_detail_chunk.cell(row=row_idx, column=31, value=float(basis))
                
                error_pct = calc_row.get("error_pct", 0.0) if isinstance(calc_row, dict) else (getattr(calc_row, "error_pct", 0.0) if hasattr(calc_row, "error_pct") else 0.0)
                ws_detail_chunk.cell(row=row_idx, column=11, value=float(error_pct))
                
                abs_err = calc_row.get("abs_err", 0.0) if isinstance(calc_row, dict) else (getattr(calc_row, "abs_err", 0.0) if hasattr(calc_row, "abs_err") else 0.0)
                ws_detail_chunk.cell(row=row_idx, column=12, value=float(abs_err))
                
                direction = calc_row.get("direction", "") if isinstance(calc_row, dict) else (getattr(calc_row, "direction", "") if hasattr(calc_row, "direction") else "")
                ws_detail_chunk.cell(row=row_idx, column=13, value=str(direction))
                
                deviation = act - sch
                ws_detail_chunk.cell(row=row_idx, column=14, value=float(deviation))
                
                rev_act = calc_row.get("Revenue_as_per_generation", act * 0.25 * 1000 * ppa) if isinstance(calc_row, dict) else (getattr(calc_row, "Revenue_as_per_generation", act * 0.25 * 1000 * ppa) if hasattr(calc_row, "Revenue_as_per_generation") else act * 0.25 * 1000 * ppa)
                ws_detail_chunk.cell(row=row_idx, column=21, value=float(rev_act))
                
                rev_sch = calc_row.get("Scheduled_Revenue_as_per_generation", sch * 0.25 * 1000 * ppa) if isinstance(calc_row, dict) else (getattr(calc_row, "Scheduled_Revenue_as_per_generation", sch * 0.25 * 1000 * ppa) if hasattr(calc_row, "Scheduled_Revenue_as_per_generation") else sch * 0.25 * 1000 * ppa)
                ws_detail_chunk.cell(row=row_idx, column=22, value=float(rev_sch))
            
            # Write band-wise columns for this chunk (reuse same logic as single-sheet path)
            # Note: Band columns are written after the main data rows in the single-sheet path below
            # For split sheets, we'll write a simplified version or skip for now to avoid complexity
            # The main data is already written above, which is the critical part
        
        # Skip the original single-sheet writing since we've split into multiple sheets
        wb.save(buf)
        buf.seek(0)
        return dcc.send_bytes(lambda x: x.write(buf.read()), "DSM_calculation.xlsx")
    
    # Original single-sheet path (when data fits in one sheet)
    for row_idx, row_data in enumerate(detail_for_excel.itertuples(index=False), start=2):
        calc_row = detail_calculated_df.iloc[row_idx - 2] if row_idx - 2 < len(detail_calculated_df) else {}
        
        # Input values
        ws_detail.cell(row=row_idx, column=1, value=getattr(row_data, "region", ""))  # A
        ws_detail.cell(row=row_idx, column=2, value=getattr(row_data, "plant_name", ""))  # B
        ws_detail.cell(row=row_idx, column=3, value=getattr(row_data, "date", ""))  # C
        ws_detail.cell(row=row_idx, column=4, value=getattr(row_data, "time_block", ""))  # D
        ws_detail.cell(row=row_idx, column=5, value=getattr(row_data, "from_time", ""))  # E
        ws_detail.cell(row=row_idx, column=6, value=getattr(row_data, "to_time", ""))  # F
        ws_detail.cell(row=row_idx, column=7, value=float(getattr(row_data, "Scheduled_MW", 0)))  # G
        ws_detail.cell(row=row_idx, column=8, value=float(getattr(row_data, "AvC_MW", 0)))  # H
        ws_detail.cell(row=row_idx, column=9, value=float(getattr(row_data, "Actual_MW", 0)))  # I
        ws_detail.cell(row=row_idx, column=10, value=float(getattr(row_data, "PPA", 0)))  # J
        
        # Calculated values (not formulas)
        avc = float(getattr(row_data, "AvC_MW", 0))
        sch = float(getattr(row_data, "Scheduled_MW", 0))
        act = float(getattr(row_data, "Actual_MW", 0))
        ppa = float(getattr(row_data, "PPA", 0))
        
        # _basis (AE) - calculated value
        basis = denominator_and_basis(avc, sch, mode_upper, dyn_x)
        ws_detail.cell(row=row_idx, column=31, value=float(basis))
        
        # Error % (K)
        error_pct = calc_row.get("error_pct", 0.0)
        ws_detail.cell(row=row_idx, column=11, value=float(error_pct))
        
        # Absolute error % (L)
        abs_err = calc_row.get("abs_err", 0.0)
        ws_detail.cell(row=row_idx, column=12, value=float(abs_err))
        
        # Direction (M)
        direction = calc_row.get("direction", "")
        ws_detail.cell(row=row_idx, column=13, value=str(direction))
        
        # Deviation (MW) (N)
        deviation = act - sch
        ws_detail.cell(row=row_idx, column=14, value=float(deviation))
        
        # Revenue as per Generation (U)
        rev_act = calc_row.get("Revenue_as_per_generation", act * 0.25 * 1000 * ppa)
        ws_detail.cell(row=row_idx, column=21, value=float(rev_act))
        
        # Scheduled Revenue (V)
        rev_sch = calc_row.get("Scheduled_Revenue_as_per_generation", sch * 0.25 * 1000 * ppa)
        ws_detail.cell(row=row_idx, column=22, value=float(rev_sch))
    
    # ===== BAND-WISE ENERGY DEVIATION COLUMNS (O-T) =====
    # Use user-defined bands directly with proper cumulative slicing
    ui_bands = sorted([b for b in bands_list if b.direction == "UI"], key=lambda x: x.lower_pct)
    oi_bands = sorted([b for b in bands_list if b.direction == "OI"], key=lambda x: x.lower_pct)
    
    # Combine bands: UI first, then OI, limit to 6 columns (O-T = columns 15-20)
    # Show all UI bands first, then OI bands
    all_bands_for_energy = (ui_bands + oi_bands)[:6]
    
    # Write band-wise energy deviation headers and values (columns O-T)
    for col_idx, band in enumerate(all_bands_for_energy, start=15):  # O=15, P=16, ..., T=20
        dir_label = "UI" if band.direction == "UI" else "OI"
        
        if band.upper_pct >= 999:
            header = f"{dir_label} Energy >{int(band.lower_pct)}% (MW)"
        else:
            header = f"{dir_label} Energy {int(band.lower_pct)}-{int(band.upper_pct)}% (MW)"
        ws_detail.cell(row=1, column=col_idx, value=header)
        
        # Write calculated energy values for each data row with proper cumulative slicing
        # Formula: SIGN(deviation) * MAX(0, MIN(abs_err, upper) - lower) / 100 * basis
        # Ensure we don't exceed Excel's row limit
        max_data_row = min(n_rows + 2, MAX_EXCEL_DATA_ROWS + 2)
        for data_row in range(2, max_data_row):
            calc_row = detail_calculated_df.iloc[data_row - 2] if data_row - 2 < len(detail_calculated_df) else {}
            abs_err = calc_row.get("abs_err", 0.0)
            direction = calc_row.get("direction", "")
            avc = detail_for_excel.iloc[data_row - 2]["AvC_MW"]
            sch = detail_for_excel.iloc[data_row - 2]["Scheduled_MW"]
            act = detail_for_excel.iloc[data_row - 2]["Actual_MW"]
            basis = denominator_and_basis(avc, sch, mode_upper, dyn_x)
            deviation = act - sch  # N2 = I2 - G2 (can be negative for UI, positive for OI)
            
            # Calculate energy deviation for this specific band with cumulative slicing
            # Formula: SIGN(deviation) * MAX(0, MIN(abs_err, upper) - lower) / 100 * basis
            # This preserves sign: negative for UI, positive for OI
            energy_mw = 0.0
            if direction == dir_label:
                # Calculate slice percentage
                sp = slice_pct(abs_err, band.lower_pct, band.upper_pct)
                if sp > 0:
                    # SIGN(deviation) preserves the correct sign
                    sign = -1.0 if deviation < 0 else (1.0 if deviation > 0 else 0.0)
                    # Convert slice percentage to MW: SIGN * slice% / 100 * basis
                    energy_mw = sign * sp / 100.0 * basis
            
            ws_detail.cell(row=data_row, column=col_idx, value=float(energy_mw))
    
    # ===== PER-BAND DSM COLUMNS (after revenue columns) =====
    # Use actual user-defined bands for DSM columns too (for consistency)
    # Start after column V (Revenue columns)
    current_col = 23  # Column W
    
    # UI DSM columns - one column per UI band
    ui_dsm_start_col = current_col
    for band in ui_bands:
        if band.upper_pct >= 999:
            header = f"UI >{int(band.lower_pct)}% DSM"
        else:
            header = f"UI {int(band.lower_pct)}-{int(band.upper_pct)}% DSM"
        ws_detail.cell(row=1, column=current_col, value=header)
        
        # Write calculated values for each data row
        # Ensure we don't exceed Excel's row limit
        max_data_row = min(n_rows + 2, MAX_EXCEL_DATA_ROWS + 2)
        for data_row in range(2, max_data_row):
            calc_row = detail_calculated_df.iloc[data_row - 2] if data_row - 2 < len(detail_calculated_df) else {}
            abs_err = calc_row.get("abs_err", 0.0)
            direction = calc_row.get("direction", "")
            avc = detail_for_excel.iloc[data_row - 2]["AvC_MW"]
            sch = detail_for_excel.iloc[data_row - 2]["Scheduled_MW"]
            ppa = detail_for_excel.iloc[data_row - 2]["PPA"]
            basis = denominator_and_basis(avc, sch, mode_upper, dyn_x)
            
            # Calculate DSM for this specific band with cumulative slicing
            dsm_value = 0.0
            if direction == "UI":
                sp = slice_pct(abs_err, band.lower_pct, band.upper_pct)
                if sp > 0:
                    kwh = kwh_from_slice(sp, basis)
                    rate = band_rate(ppa, band.rate_type, band.rate_value, band.rate_slope, abs_err)
                    dsm_value = kwh * rate
            
            ws_detail.cell(row=data_row, column=current_col, value=float(dsm_value))
        
        current_col += 1
    ui_dsm_end_col = current_col - 1
    
    # OI DSM columns (LossZone=FALSE) - one column per OI band
    oi_dsm_start_col = current_col
    for band in oi_bands:
        if not band.loss_zone:  # Only create columns for non-loss-zone OI bands
            if band.upper_pct >= 999:
                header = f"OI >{int(band.lower_pct)}% DSM"
            else:
                header = f"OI {int(band.lower_pct)}-{int(band.upper_pct)}% DSM"
            ws_detail.cell(row=1, column=current_col, value=header)
            
            max_data_row = min(n_rows + 2, MAX_EXCEL_DATA_ROWS + 2)
            for data_row in range(2, max_data_row):
                calc_row = detail_calculated_df.iloc[data_row - 2] if data_row - 2 < len(detail_calculated_df) else {}
                abs_err = calc_row.get("abs_err", 0.0)
                direction = calc_row.get("direction", "")
                avc = detail_for_excel.iloc[data_row - 2]["AvC_MW"]
                sch = detail_for_excel.iloc[data_row - 2]["Scheduled_MW"]
                ppa = detail_for_excel.iloc[data_row - 2]["PPA"]
                basis = denominator_and_basis(avc, sch, mode_upper, dyn_x)
                
                dsm_value = 0.0
                if direction == "OI":
                    sp = slice_pct(abs_err, band.lower_pct, band.upper_pct)
                    if sp > 0:
                        kwh = kwh_from_slice(sp, basis)
                        rate = band_rate(ppa, band.rate_type, band.rate_value, band.rate_slope, abs_err)
                        dsm_value = kwh * rate
                
                ws_detail.cell(row=data_row, column=current_col, value=float(dsm_value))
            
            current_col += 1
    oi_dsm_end_col = current_col - 1
    
    # OI Loss columns (LossZone=TRUE) - one column per loss-zone OI band
    oi_loss_start_col = current_col
    for band in oi_bands:
        if band.loss_zone:  # Only create columns for loss-zone OI bands
            if band.upper_pct >= 999:
                header = f"OI >{int(band.lower_pct)}% Loss"
            else:
                header = f"OI {int(band.lower_pct)}-{int(band.upper_pct)}% Loss"
            ws_detail.cell(row=1, column=current_col, value=header)
            
            max_data_row = min(n_rows + 2, MAX_EXCEL_DATA_ROWS + 2)
            for data_row in range(2, max_data_row):
                calc_row = detail_calculated_df.iloc[data_row - 2] if data_row - 2 < len(detail_calculated_df) else {}
                abs_err = calc_row.get("abs_err", 0.0)
                direction = calc_row.get("direction", "")
                avc = detail_for_excel.iloc[data_row - 2]["AvC_MW"]
                sch = detail_for_excel.iloc[data_row - 2]["Scheduled_MW"]
                ppa = detail_for_excel.iloc[data_row - 2]["PPA"]
                basis = denominator_and_basis(avc, sch, mode_upper, dyn_x)
                
                loss_value = 0.0
                if direction == "OI":
                    sp = slice_pct(abs_err, band.lower_pct, band.upper_pct)
                    if sp > 0:
                        kwh = kwh_from_slice(sp, basis)
                        rate = band_rate(ppa, band.rate_type, band.rate_value, band.rate_slope, abs_err)
                        loss_value = kwh * rate
                
                ws_detail.cell(row=data_row, column=current_col, value=float(loss_value))
            
            current_col += 1
    oi_loss_end_col = current_col - 1
    
    # ===== TOTAL COLUMNS (AC and AD) =====
    # AC = column 29, AD = column 30
    ws_detail.cell(row=1, column=29, value='Total DSM (INR)')
    ws_detail.cell(row=1, column=30, value='Revenue Loss (INR)')
    
    # Ensure we don't exceed Excel's row limit
    max_data_row = min(n_rows + 2, MAX_EXCEL_DATA_ROWS + 2)
    for data_row in range(2, max_data_row):
        calc_row = detail_calculated_df.iloc[data_row - 2] if data_row - 2 < len(detail_calculated_df) else {}
        
        # Total DSM = calculated value
        total_dsm = calc_row.get("Total_DSM", 0.0)
        ws_detail.cell(row=data_row, column=29, value=float(total_dsm))
        
        # Revenue Loss = calculated value
        revenue_loss = calc_row.get("Revenue_Loss", 0.0)
        ws_detail.cell(row=data_row, column=30, value=float(revenue_loss))
    
    # ===== SUMMARY AREA (optional, at bottom of Detail sheet) =====
    summary_start = n_rows + 5
    ws_detail.cell(row=summary_start, column=1, value='Summary')
    
    # Calculate summary values
    if len(detail_calculated_df) > 0:
        # Get summary from plant_summary if available, otherwise calculate
        region_val = detail_calculated_df["region"].mode().iloc[0] if "region" in detail_calculated_df.columns else ""
        plant_val = detail_calculated_df["plant_name"].mode().iloc[0] if "plant_name" in detail_calculated_df.columns else ""
        avc_mode = safe_mode(detail_for_excel["AvC_MW"].tolist())
        ppa_mode = safe_mode(detail_for_excel["PPA"].tolist())
        rev_loss_sum = detail_calculated_df["Revenue_Loss"].sum() if "Revenue_Loss" in detail_calculated_df.columns else 0.0
        rev_act_sum = detail_calculated_df["Revenue_as_per_generation"].sum() if "Revenue_as_per_generation" in detail_calculated_df.columns else 0.0
        rev_loss_pct = (rev_loss_sum / rev_act_sum * 100.0) if rev_act_sum > 0 else 0.0
        dsm_loss_sum = detail_calculated_df["Total_DSM"].sum() if "Total_DSM" in detail_calculated_df.columns else 0.0
        
        ws_detail.cell(row=summary_start + 1, column=1, value='Region')
        ws_detail.cell(row=summary_start + 1, column=2, value=str(region_val))
        ws_detail.cell(row=summary_start + 2, column=1, value='Plant Name')
        ws_detail.cell(row=summary_start + 2, column=2, value=str(plant_val))
        ws_detail.cell(row=summary_start + 3, column=1, value='Plant Capacity (MODE AvC)')
        ws_detail.cell(row=summary_start + 3, column=2, value=round(float(avc_mode), 2))
        ws_detail.cell(row=summary_start + 4, column=1, value='PPA (MODE)')
        ws_detail.cell(row=summary_start + 4, column=2, value=round(float(ppa_mode), 2))
        ws_detail.cell(row=summary_start + 5, column=1, value='Revenue Loss %')
        ws_detail.cell(row=summary_start + 5, column=2, value=round(float(rev_loss_pct), 2))
        ws_detail.cell(row=summary_start + 6, column=1, value='DSM Loss')
        ws_detail.cell(row=summary_start + 6, column=2, value=round(float(dsm_loss_sum), 2))
    
    wb.save(buf)
    buf.seek(0)
    return dcc.send_bytes(lambda x: x.write(buf.read()), "DSM_calculation.xlsx"), (_done or 0) + 1


@app.callback(
    Output("download-plant-summary", "data"),
    Output("dl-done-store", "data", allow_duplicate=True),
    Input("btn-download-plant-summary", "n_clicks"),
    State("results-store", "data"),
    State("dl-done-store", "data"),
    prevent_initial_call=True
)
def download_plant_summary(n, stored, _done):
    """Download the Analysis Plant Summary table (including master metadata columns) as an Excel file."""
    if not n:
        raise PreventUpdate
    if not stored or (isinstance(stored, dict) and stored.get("error")):
        raise PreventUpdate
    ps = pd.DataFrame(stored.get("plant_summary", [])) if isinstance(stored, dict) else pd.DataFrame()
    if ps.empty:
        raise PreventUpdate
    ps = enrich_plant_summary_with_master(ps)

    out = BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as xw:
        ps.to_excel(xw, sheet_name="Plant Summary", index=False)
    out.seek(0)
    return dcc.send_bytes(lambda x: x.write(out.read()), "Plant_Summary.xlsx"), (_done or 0) + 1


@app.callback(
    Output("download-excel", "data", allow_duplicate=True),
    Input({"type": "agg-btn-download", "index": ALL}, "n_clicks"),
    State("agg-results-store", "data"),
    prevent_initial_call=True
)
def download_full_aggregated(n_clicks_list, stored):
    """Download full calculation for Aggregation Analysis, reusing the same
    Excel structure as the main Analysis export."""
    if not n_clicks_list or not any(n_clicks_list):
        raise PreventUpdate
    if not stored or stored.get("error"):
        raise PreventUpdate

    # Multi-preset aggregated export (only when >1 runs; single-run should use the detailed export below)
    all_runs = stored.get("_all_runs") if isinstance(stored, dict) else None
    if all_runs and isinstance(all_runs, list) and len(all_runs) > 1:
        from io import BytesIO
        output = BytesIO()
        with pd.ExcelWriter(output, engine="xlsxwriter") as xw:
            # Combined Plant Summary
            try:
                plant_summary_df = pd.DataFrame(stored.get("plant_summary", []))
                if not plant_summary_df.empty:
                    plant_summary_df.to_excel(xw, sheet_name="Plant_Summary", index=False)
            except Exception as e:
                print(f"DEBUG - (agg) Failed writing Plant_Summary: {e}")

            # One Detail/Config/Bands per preset
            for r in all_runs:
                nm = str(r.get("name", "Preset"))
                df_detail = _detail_df_from_store(r if isinstance(r, dict) else {})
                bands_df = _normalize_bands_df(pd.DataFrame(r.get("final_bands", [])))
                # Detail
                try:
                    df_detail.to_excel(xw, sheet_name=f"Detail_{nm}", index=False)
                except Exception as e:
                    print(f"DEBUG - (agg) Writing Detail_{nm} failed: {e}")
                # Config
                try:
                    pd.DataFrame({"Key":["MODE","DYN_X"], "Value":[r.get("err_mode"), r.get("x_pct")]}) \
                        .to_excel(xw, sheet_name=f"Config_{nm}", index=False)
                except Exception as e:
                    print(f"DEBUG - (agg) Writing Config_{nm} failed: {e}")
                # Bands
                try:
                    bands_df.to_excel(xw, sheet_name=f"Bands_{nm}", index=False)
                except Exception as e:
                    print(f"DEBUG - (agg) Writing Bands_{nm} failed: {e}")
        return dcc.send_bytes(output.getvalue(), filename="DSM_Full_Calculation_Aggregated_MultiPresets.xlsx")

    # Single-setting path – identical to main download_full, but reading from agg-results-store
    df_main = _detail_df_from_store(stored)
    if df_main.empty:
        raise PreventUpdate

    used = stored.get("used_settings", {}) if isinstance(stored, dict) else {}
    err_mode = str(used.get("err_mode", "default")).lower()
    mode_upper = MODE_DEFAULT if err_mode == "default" else MODE_DYNAMIC
    try:
        x_pct = float(used.get("x_pct", 50))
    except Exception:
        x_pct = 50.0
    dyn_x = (x_pct / 100.0) if mode_upper == MODE_DYNAMIC else 0.0

    bands_rows = stored.get("final_bands", []) if isinstance(stored, dict) else []
    bands_list, bands_table = parse_bands_from_settings(bands_rows)

    # Select input columns for export
    base_cols = [
        "region","plant_name","date","time_block","from_time","to_time",
        "AvC_MW","Scheduled_MW","Actual_MW","PPA"
    ]
    missing = [c for c in base_cols if c not in df_main.columns]
    if missing:
        df_exp = df_main.copy()
        if "plant_name" not in df_exp.columns and "Plant" in df_exp.columns:
            df_exp["plant_name"] = df_exp["Plant"]
        if "time_block" not in df_exp.columns and "block" in df_exp.columns:
            df_exp["time_block"] = df_exp["block"]
        missing2 = [c for c in base_cols if c not in df_exp.columns]
        if missing2:
            raise PreventUpdate
        detail_for_excel = df_exp[base_cols].copy()
    else:
        detail_for_excel = df_main[base_cols].copy()

    # Reuse same detail_calculated_df logic
    detail_calculated_rows = []
    for _, row in detail_for_excel.iterrows():
        slot = {
            "region": row["region"],
            "plant_name": row["plant_name"],
            "date": row["date"],
            "time_block": row["time_block"],
            "from_time": row["from_time"],
            "to_time": row["to_time"],
            "AvC_MW": float(row["AvC_MW"]),
            "Scheduled_MW": float(row["Scheduled_MW"]),
            "Actual_MW": float(row["Actual_MW"]),
            "PPA": float(row["PPA"]),
        }
        calc = compute_slot_row(slot, bands_list, mode_upper, dyn_x)
        detail_calculated_rows.append({**slot, **calc})
    detail_calculated_df = pd.DataFrame(detail_calculated_rows)

    # The remainder of the export logic (per-band columns, headers, summary, etc.)
    # is identical to the main download_full implementation and is reused here
    # by following the same steps.

    # Build per-band columns
    per_band_cols = {}
    per_band_dsm_cols = {}
    for band in bands_list:
        dir_label = "UI" if band.direction == "UI" else "OI"
        if band.upper_pct >= 999:
            band_label = f"{dir_label} Energy deviation >{int(band.lower_pct)}%"
            dsm_label = f"{dir_label} DSM due to Deviation >{int(band.lower_pct)}% (INR)"
        else:
            band_label = f"{dir_label} Energy deviation between {int(band.lower_pct)}-{int(band.upper_pct)}%"
            dsm_label = f"{dir_label} DSM between {int(band.lower_pct)}-{int(band.upper_pct)}% (INR)"
        per_band_cols[band_label] = []
        per_band_dsm_cols[dsm_label] = []

    for _, row in detail_calculated_df.iterrows():
        abs_err = row.get("abs_err", 0.0)
        direction = row.get("direction", "")
        avc = row.get("AvC_MW", 0.0)
        sch = row.get("Scheduled_MW", 0.0)
        ppa = row.get("PPA", 0.0)
        denom = denominator_and_basis(avc, sch, mode_upper, dyn_x)

        for band_idx, band in enumerate(bands_list):
            energy_col_name = list(per_band_cols.keys())[band_idx]
            dsm_col_name = list(per_band_dsm_cols.keys())[band_idx]
            if band.direction != direction:
                per_band_cols[energy_col_name].append(0.0)
                per_band_dsm_cols[dsm_col_name].append(0.0)
                continue
            sp = slice_pct(abs_err, band.lower_pct, band.upper_pct)
            if sp > 0:
                kwh = kwh_from_slice(sp, denom)
                rate = band_rate(ppa, band.rate_type, band.rate_value, band.rate_slope, abs_err)
                amount = kwh * rate
                per_band_cols[energy_col_name].append(kwh / 1000.0)
                per_band_dsm_cols[dsm_col_name].append(amount)
            else:
                per_band_cols[energy_col_name].append(0.0)
                per_band_dsm_cols[dsm_col_name].append(0.0)

    for col_name, values in per_band_cols.items():
        detail_calculated_df[col_name] = values
    for col_name, values in per_band_dsm_cols.items():
        detail_calculated_df[col_name] = values

    standard_calc_cols = [
        "error_pct", "direction", "abs_err", "band_level",
        "Revenue_as_per_generation", "Scheduled_Revenue_as_per_generation",
        "UI_DSM", "OI_DSM", "OI_Loss", "Total_DSM", "Revenue_Loss"
    ]

    def get_band_sort_key(col_name):
        if "UI" in col_name:
            dir_val = 0
        elif "OI" in col_name:
            dir_val = 1
        else:
            dir_val = 2
        import re
        nums = re.findall(r'\d+', col_name)
        lower_val = int(nums[0]) if nums else 9999
        return (dir_val, lower_val)

    per_band_col_names = sorted(list(per_band_cols.keys()), key=get_band_sort_key)
    per_band_dsm_col_names = sorted(list(per_band_dsm_cols.keys()), key=get_band_sort_key)
    interleaved_per_band = []
    for energy_col in per_band_col_names:
        interleaved_per_band.append(energy_col)
        import re
        energy_range = re.search(r'(\d+-?\d*%|>\d+%)', energy_col)
        if energy_range:
            range_str = energy_range.group(1)
            for dsm_col in per_band_dsm_col_names:
                if range_str in dsm_col and energy_col.split()[0] in dsm_col:
                    interleaved_per_band.append(dsm_col)
                    break
    final_cols = base_cols + [c for c in standard_calc_cols if c in detail_calculated_df.columns] + interleaved_per_band
    detail_calculated_df = detail_calculated_df[[c for c in final_cols if c in detail_calculated_df.columns]]

    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter, column_index_from_string

    buf = BytesIO()
    wb = Workbook()
    wb.remove(wb.active)

    # Reuse same Config/Bands/Detail sheet build as main export
    ws_config = wb.create_sheet("Config")
    ws_config.cell(row=1, column=1, value="Key")
    ws_config.cell(row=1, column=2, value="Value")
    ws_config.cell(row=2, column=1, value="MODE")
    ws_config.cell(row=2, column=2, value=mode_upper)
    ws_config.cell(row=3, column=1, value="DYN_X")
    ws_config.cell(row=3, column=2, value=float(dyn_x))

    from openpyxl.workbook.defined_name import DefinedName
    try:
        wb.defined_names.append(DefinedName(name="CFG_MODE", attr_text="Config!$B$2"))
        wb.defined_names.append(DefinedName(name="CFG_DYNX", attr_text="Config!$B$3"))
    except Exception:
        pass

    ws_bands = wb.create_sheet("Bands")
    bands_cols = ["direction","lower_pct","upper_pct","rate_type","rate_value","rate_slope","loss_zone"]
    for col_idx, col_name in enumerate(bands_cols, start=1):
        ws_bands.cell(row=1, column=col_idx, value=col_name)
    for row_idx, row in enumerate(bands_table.itertuples(index=False), start=2):
        ws_bands.cell(row=row_idx, column=1, value=str(getattr(row, "direction", "")))
        ws_bands.cell(row=row_idx, column=2, value=float(getattr(row, "lower_pct", 0)))
        ws_bands.cell(row=row_idx, column=3, value=float(getattr(row, "upper_pct", 0)))
        ws_bands.cell(row=row_idx, column=4, value=str(getattr(row, "rate_type", "")))
        ws_bands.cell(row=row_idx, column=5, value=float(getattr(row, "rate_value", 0)))
        ws_bands.cell(row=row_idx, column=6, value=float(getattr(row, "rate_slope", 0)))
        ws_bands.cell(row=row_idx, column=7, value=bool(getattr(row, "loss_zone", False)))

    n_bands = max(len(bands_table), 1)
    bands_end_row = 1 + n_bands
    bands_end_ref = 200 if n_bands < 199 else bands_end_row

    try:
        wb.defined_names.append(DefinedName(name="Bands_Dir", attr_text=f"Bands!$A$2:$A${bands_end_ref}"))
        wb.defined_names.append(DefinedName(name="Bands_Lower", attr_text=f"Bands!$B$2:$B${bands_end_ref}"))
        wb.defined_names.append(DefinedName(name="Bands_Upper", attr_text=f"Bands!$C$2:$C${bands_end_ref}"))
        wb.defined_names.append(DefinedName(name="Bands_RateType", attr_text=f"Bands!$D$2:$D${bands_end_ref}"))
        wb.defined_names.append(DefinedName(name="Bands_RateVal", attr_text=f"Bands!$E$2:$E${bands_end_ref}"))
        wb.defined_names.append(DefinedName(name="Bands_RateSlope", attr_text=f"Bands!$F$2:$F${bands_end_ref}"))
        wb.defined_names.append(DefinedName(name="Bands_LossZone", attr_text=f"Bands!$G$2:$G${bands_end_ref}"))
    except Exception:
        pass

    ws_detail = wb.create_sheet("Detail")
    detail_headers = {
        'A': 'Region',
        'B': 'Plant Name',
        'C': 'Date',
        'D': 'Block',
        'E': 'From Time',
        'F': 'To Time',
        'G': 'Schedule Power (MW)',
        'H': 'AvC (MW)',
        'I': 'Injected Power (MW)',
        'J': 'PPA or MCP',
        'K': 'Error %',
        'L': 'Absolute error %',
        'M': 'Direction',
        'N': 'Deviation (MW)',
        'U': 'Revenue as per Generation (INR)',
        'V': 'Scheduled Revenue (INR)',
        'AE': '_basis'
    }

    for col_letter, header in detail_headers.items():
        col_idx = column_index_from_string(col_letter)
        ws_detail.cell(row=1, column=col_idx, value=header)

    ws_detail.column_dimensions['AE'].hidden = True

    n_rows = len(detail_for_excel)
    for row_idx, row_data in enumerate(detail_for_excel.itertuples(index=False), start=2):
        calc_row = detail_calculated_df.iloc[row_idx - 2] if row_idx - 2 < len(detail_calculated_df) else {}

        ws_detail.cell(row=row_idx, column=1, value=getattr(row_data, "region", ""))
        ws_detail.cell(row=row_idx, column=2, value=getattr(row_data, "plant_name", ""))
        ws_detail.cell(row=row_idx, column=3, value=getattr(row_data, "date", ""))
        ws_detail.cell(row=row_idx, column=4, value=getattr(row_data, "time_block", ""))
        ws_detail.cell(row=row_idx, column=5, value=getattr(row_data, "from_time", ""))
        ws_detail.cell(row=row_idx, column=6, value=getattr(row_data, "to_time", ""))
        ws_detail.cell(row=row_idx, column=7, value=float(getattr(row_data, "Scheduled_MW", 0)))
        ws_detail.cell(row=row_idx, column=8, value=float(getattr(row_data, "AvC_MW", 0)))
        ws_detail.cell(row=row_idx, column=9, value=float(getattr(row_data, "Actual_MW", 0)))
        ws_detail.cell(row=row_idx, column=10, value=float(getattr(row_data, "PPA", 0)))

        avc = float(getattr(row_data, "AvC_MW", 0))
        sch = float(getattr(row_data, "Scheduled_MW", 0))
        act = float(getattr(row_data, "Actual_MW", 0))
        ppa = float(getattr(row_data, "PPA", 0))

        basis = denominator_and_basis(avc, sch, mode_upper, dyn_x)
        ws_detail.cell(row=row_idx, column=31, value=float(basis))

        error_pct = calc_row.get("error_pct", 0.0)
        ws_detail.cell(row=row_idx, column=11, value=float(error_pct))

        abs_err = calc_row.get("abs_err", 0.0)
        ws_detail.cell(row=row_idx, column=12, value=float(abs_err))

        direction = calc_row.get("direction", "")
        ws_detail.cell(row=row_idx, column=13, value=str(direction))

        deviation = act - sch
        ws_detail.cell(row=row_idx, column=14, value=float(deviation))

        rev_act = calc_row.get("Revenue_as_per_generation", 0.0)
        ws_detail.cell(row=row_idx, column=21, value=float(rev_act))

        rev_sch = calc_row.get("Scheduled_Revenue_as_per_generation", 0.0)
        ws_detail.cell(row=row_idx, column=22, value=float(rev_sch))

    # Reuse final summary section pattern if desired, or keep as-is
    wb.save(buf)
    buf.seek(0)
    return dcc.send_bytes(lambda x: x.write(buf.read()), "DSM_calculation_aggregated.xlsx")


# ==========================================
# ------- DOWNLOAD PROGRESS CALLBACKS ------
# ==========================================

_DL_OVERLAY_HIDDEN = {
    "display": "none",
    "position": "fixed",
    "top": 0, "left": 0, "right": 0, "bottom": 0,
    "zIndex": 9999,
    "backgroundColor": "rgba(15, 20, 40, 0.55)",
    "alignItems": "center",
    "justifyContent": "center",
    "backdropFilter": "blur(4px)",
}
_DL_OVERLAY_VISIBLE = {**_DL_OVERLAY_HIDDEN, "display": "flex"}


@app.callback(
    Output("dl-progress-store", "data"),
    Input("btn-download", "n_clicks"),
    Input("btn-download-plant-summary", "n_clicks"),
    Input("download-custom-output-btn", "n_clicks"),
    State("results-store", "data"),
    State("custom-results-store", "data"),
    prevent_initial_call=True,
)
def start_download_progress(n1, n2, n3, results, custom_results):
    triggered_id = ctx.triggered_id
    if not triggered_id:
        raise PreventUpdate

    if triggered_id == "btn-download":
        label = "Full Calculation (Excel)"
        stored = results
    elif triggered_id == "btn-download-plant-summary":
        label = "Plant Summary (Excel)"
        stored = results
    elif triggered_id == "download-custom-output-btn":
        label = "Custom Upload Output (Excel)"
        stored = custom_results
    else:
        raise PreventUpdate

    # Estimate export time from data size
    est_sec = 8.0
    if stored and isinstance(stored, dict):
        all_runs = stored.get("_all_runs") or []
        ps = stored.get("plant_summary") or []
        n_plants = len(ps) if isinstance(ps, list) else 0
        n_presets = max(1, len(all_runs))
        est_sec = max(5.0, 3.0 + n_plants * 0.4 + n_presets * 4.0)
        if triggered_id == "btn-download-plant-summary":
            est_sec = max(3.0, 2.0 + n_plants * 0.1)

    return {"active": True, "t0": time.time(), "label": label, "est_sec": float(est_sec)}


@app.callback(
    Output("dl-progress-overlay", "style"),
    Output("dl-progress-label", "children"),
    Output("dl-progress-bar", "value"),
    Output("dl-progress-pct", "children"),
    Output("dl-progress-time", "children"),
    Input("dl-progress-interval", "n_intervals"),
    State("dl-progress-store", "data"),
)
def update_download_progress(_, store):
    if not store or not store.get("active"):
        raise PreventUpdate

    t0 = float(store.get("t0", time.time()))
    est_sec = float(store.get("est_sec", 10))
    label = store.get("label", "")
    elapsed = time.time() - t0

    if elapsed >= est_sec * 1.4:
        # Beyond timeout: show 100% (file should be downloading by now)
        pct = 100
        time_str = "Download started!"
    elif elapsed >= est_sec:
        # Overshoot zone: creep from 95→99%
        overshoot = elapsed - est_sec
        pct = min(99, 95 + int(overshoot / (est_sec * 0.4) * 4))
        time_str = "Almost done\u2026"
    else:
        pct = int((elapsed / est_sec) * 95)
        remaining = est_sec - elapsed
        if remaining > 60:
            time_str = f"~{int(remaining // 60)}m {int(remaining % 60)}s remaining"
        elif remaining > 2:
            time_str = f"~{int(remaining)}s remaining"
        else:
            time_str = "Almost done\u2026"

    return _DL_OVERLAY_VISIBLE, label, pct, f"{pct}%", time_str


@app.callback(
    Output("dl-progress-store", "data", allow_duplicate=True),
    Output("dl-progress-overlay", "style", allow_duplicate=True),
    Input("dl-done-store", "data"),
    prevent_initial_call=True,
)
def reset_download_progress(done_count):
    if not done_count:
        raise PreventUpdate
    return {"active": False, "t0": 0, "label": "", "est_sec": 10}, _DL_OVERLAY_HIDDEN


@app.callback(
    Output("dl-progress-store", "data", allow_duplicate=True),
    Output("dl-progress-overlay", "style", allow_duplicate=True),
    Input("plot-now", "n_clicks"),
    Input("agg-plot-now", "n_clicks"),
    Input("btn-run-custom", "n_clicks"),
    Input("stats-run", "n_clicks"),
    Input("nav-store", "data"),
    State("dl-progress-store", "data"),
    prevent_initial_call=True,
)
def clear_download_progress_on_non_download_actions(_, __, ___, ____, nav, store):
    # The download overlay should never appear for analysis/run actions.
    # If a previous download left the progress store active, clear it as soon as
    # the user starts a normal analysis flow or changes sections.
    if not store or not store.get("active"):
        raise PreventUpdate
    return {"active": False, "t0": 0, "label": "", "est_sec": 10}, _DL_OVERLAY_HIDDEN


@app.callback(
    Output("dl-progress-store", "data", allow_duplicate=True),
    Output("dl-progress-overlay", "style", allow_duplicate=True),
    Input("dl-progress-interval", "n_intervals"),
    State("dl-progress-store", "data"),
    prevent_initial_call=True,
)
def fail_safe_reset_download_progress(_, store):
    if not store or not store.get("active"):
        raise PreventUpdate
    t0 = float(store.get("t0", time.time()))
    est_sec = float(store.get("est_sec", 10))
    # If the browser never receives the done signal, do not let the overlay leak
    # into later actions indefinitely.
    if (time.time() - t0) < max(15.0, est_sec * 2.5):
        raise PreventUpdate
    return {"active": False, "t0": 0, "label": "", "est_sec": 10}, _DL_OVERLAY_HIDDEN


# ==========================================
# ------- MANAGE PRESETS CALLBACKS ---------
# ==========================================

def _make_bands_preview(bands):
    """Return a compact, read-only DataTable for a list of band dicts."""
    if not bands:
        return html.Div(
            "No bands configured in this preset.",
            style={"color": "#aaa", "fontStyle": "italic", "fontSize": "0.85rem", "padding": "8px 0"}
        )
    return dash_table.DataTable(
        data=bands,
        columns=[
            {"name": "Dir", "id": "direction"},
            {"name": "Lower %", "id": "lower_pct"},
            {"name": "Upper %", "id": "upper_pct"},
            {"name": "Rate Type", "id": "rate_type"},
            {"name": "Rate Value", "id": "rate_value"},
            {"name": "Excess Slope", "id": "rate_slope"},
            {"name": "Loss Zone", "id": "loss_zone"},
        ],
        style_table={"overflowX": "auto"},
        style_cell={
            "padding": "7px 12px", "fontSize": "0.82rem",
            "fontFamily": "inherit", "border": "none",
            "borderBottom": "1px solid #f2f2f2",
        },
        style_header={
            "backgroundColor": "#f7f8fa", "fontWeight": "700",
            "fontSize": "0.74rem", "color": "#666",
            "textTransform": "uppercase", "letterSpacing": "0.4px",
            "border": "none", "borderBottom": "2px solid #e4e8ef",
        },
        style_data_conditional=[{"if": {"row_index": "odd"}, "backgroundColor": "#fafbfc"}],
    )


@app.callback(
    Output("preset-manage-select", "options"),
    Input("presets-store", "data"),
)
def update_preset_manage_options(presets):
    presets = presets or []
    return [{"label": p.get("name", "Unnamed"), "value": p.get("name", "Unnamed")} for p in presets]


@app.callback(
    Output("preset-detail-view", "children"),
    Output("btn-load-preset", "disabled"),
    Output("btn-delete-preset-manage", "disabled"),
    Output("preset-mode-badge", "children"),
    Output("save-mode-hint", "children"),
    Input("preset-manage-select", "value"),
    State("presets-store", "data"),
)
def view_preset_detail(selected_name, presets):
    _empty = dbc.Badge("New / Unsaved", color="secondary", pill=True)
    _hint_new = "Fill the form below and save as a new preset."
    if not selected_name or not presets:
        return html.Div(), True, True, _empty, _hint_new

    preset = next((p for p in presets if p.get("name") == selected_name), None)
    if not preset:
        return html.Div(), True, True, _empty, _hint_new

    s = preset.get("settings", {})
    err_mode = s.get("err_mode", "default")
    x_pct = s.get("x_pct", 50)
    zero_basis_guard = s.get("zero_basis_guard", False)
    bands = s.get("bands") or []
    n_bands = len(bands)

    mode_label = (
        f"Default \u2014 (Actual \u2212 Scheduled) / AvC \u00d7 100"
        if err_mode == "default"
        else f"Dynamic \u2014 X = {x_pct}%"
    )

    # Compact 3-stat summary strip
    summary_strip = dbc.Row([
        dbc.Col([
            html.Div("ERROR MODE", style=_stat_label),
            html.Div(mode_label, style={**_stat_value, "fontSize": "0.82rem"}),
        ], md=5),
        dbc.Col([
            html.Div("X %", style=_stat_label),
            html.Div(str(x_pct) if err_mode == "dynamic" else "\u2014", style=_stat_value),
        ], md=2),
        dbc.Col([
            html.Div("ZERO BASIS GUARD", style=_stat_label),
            dbc.Badge("ON", color="success", pill=True, style={"marginTop": "2px"}) if zero_basis_guard
            else dbc.Badge("OFF", color="secondary", pill=True, style={"marginTop": "2px"}),
        ], md=3),
        dbc.Col([
            html.Div("BANDS", style=_stat_label),
            html.Div(str(n_bands), style=_stat_value),
        ], md=2),
    ], className="mb-0")

    detail_view = html.Div([
        html.Hr(style={"borderColor": "#eef0f4", "margin": "16px 0 14px"}),
        summary_strip,
        html.Div([
            html.Div("DSM BANDS", style={**_stat_label, "marginBottom": "8px", "marginTop": "14px"}),
            _make_bands_preview(bands),
        ]),
    ])

    badge = dbc.Badge(f"\u2713  {selected_name}", color="primary", pill=True)
    hint = f"Editing \u201c{selected_name}\u201d \u2014 load into form, adjust, then update."
    return detail_view, False, False, badge, hint


# Shared style constants for stat labels / values inside the detail view
_stat_label = {
    "fontSize": "0.68rem", "color": "#aaa",
    "textTransform": "uppercase", "letterSpacing": "0.5px", "marginBottom": "3px",
}
_stat_value = {"fontWeight": "700", "fontSize": "0.92rem", "color": "#222"}


@app.callback(
    Output("err-mode", "value", allow_duplicate=True),
    Output("x-pct", "value", allow_duplicate=True),
    Output("bands-table", "data", allow_duplicate=True),
    Output("zero-basis-guard", "value", allow_duplicate=True),
    Output("preset-name", "value"),
    Output("preset-manage-message", "children"),
    Input("btn-load-preset", "n_clicks"),
    State("preset-manage-select", "value"),
    State("presets-store", "data"),
    prevent_initial_call=True,
)
def load_preset_into_form(n_clicks, selected_name, presets):
    if not n_clicks or not selected_name:
        raise PreventUpdate
    preset = next((p for p in (presets or []) if p.get("name") == selected_name), None)
    if not preset:
        raise PreventUpdate
    s = preset.get("settings", {})
    err_mode = s.get("err_mode", "default")
    x_pct = s.get("x_pct", 50)
    bands = s.get("bands") or DEFAULT_BANDS.copy()
    zbg_val = ["on"] if s.get("zero_basis_guard", False) else []
    msg = dbc.Alert(
        ["\u2705  ", html.Strong(f"'{selected_name}'"), " loaded into the form. Edit below, then click Update Selected."],
        color="success", dismissable=True, duration=5000,
        style={"fontSize": "0.86rem", "borderRadius": "8px"},
    )
    return err_mode, x_pct, bands, zbg_val, selected_name, msg


@app.callback(
    Output("presets-store", "data", allow_duplicate=True),
    Output("preset-manage-select", "value"),
    Output("preset-manage-message", "children", allow_duplicate=True),
    Input("btn-delete-preset-manage", "n_clicks"),
    State("preset-manage-select", "value"),
    State("presets-store", "data"),
    prevent_initial_call=True,
)
def delete_preset_manage(n_clicks, selected_name, presets):
    if not n_clicks or not selected_name:
        raise PreventUpdate
    presets = [p for p in (presets or []) if p.get("name") != selected_name]
    msg = dbc.Alert(
        ["\U0001f5d1\ufe0f  ", html.Strong(f"'{selected_name}'"), " deleted."],
        color="warning", dismissable=True, duration=4000,
        style={"fontSize": "0.86rem", "borderRadius": "8px"},
    )
    return presets, None, msg


@app.callback(
    Output("err-mode", "value", allow_duplicate=True),
    Output("x-pct", "value", allow_duplicate=True),
    Output("bands-table", "data", allow_duplicate=True),
    Output("zero-basis-guard", "value", allow_duplicate=True),
    Output("preset-name", "value", allow_duplicate=True),
    Output("preset-manage-select", "value", allow_duplicate=True),
    Input("btn-preset-new", "n_clicks"),
    prevent_initial_call=True,
)
def create_new_preset(n_clicks):
    if not n_clicks:
        raise PreventUpdate
    return "default", 50, DEFAULT_BANDS.copy(), ["on"], "", None


if __name__ == "__main__":
    # Fixed port + no reloader (Windows socket / duplicate-process issues with debug reload)
    app.run(
        debug=True,
        host="127.0.0.1",
        port=DASH_PORT,
        use_reloader=False,
        threaded=True,
    )